"""
migrador.py — Criador de experimentos CSV + TOML para o Natalia Time

Modo Automático: lê diretamente o .xlsx de dados (Dados_O2_100eV.xlsx)
Modo Manual:     usuário cola os dados copiados do Excel

Uso: python migrador.py  (ou duplo clique no .exe gerado)
"""

import os
import sys
import math
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# ── Cores ──────────────────────────────────────────────────────────────────────
BG     = "#0e1117"
BG2    = "#161b22"
BG3    = "#1c2128"
BORDA  = "#30363d"
AZUL   = "#388bfd"
TEXTO  = "#e0e0e0"
TEXTO2 = "#8b949e"
VERDE  = "#3fb950"
VERM   = "#f85149"
AMAR   = "#d29922"
FONTE  = ("Segoe UI", 10)
FONTE_SEC = ("Segoe UI", 10, "bold")

ABA_ANALISE = "Analise_tempo"
LINHA_DADOS = 13   # primeira linha de dados em L e M
COL_TEMPO   = 12   # coluna L
COL_SINAL   = 13   # coluna M


# ── Helpers ────────────────────────────────────────────────────────────────────
def _entry(parent, width=20, default=""):
    e = tk.Entry(parent, width=width, bg=BG3, fg=TEXTO, insertbackground=TEXTO,
                 relief="flat", font=FONTE,
                 highlightbackground=BORDA, highlightcolor=AZUL, highlightthickness=1)
    if default:
        e.insert(0, default)
    return e


def _lbl(parent, text, fg=TEXTO2, **kw):
    return tk.Label(parent, text=text, bg=BG2, fg=fg, font=FONTE, **kw)


def _btn(parent, text, cmd, cor=AZUL):
    return tk.Button(parent, text=text, command=cmd,
                     bg=cor, fg="#ffffff", font=FONTE,
                     relief="flat", padx=10, pady=4,
                     activebackground=cor, activeforeground="#ffffff",
                     cursor="hand2")


def _secao(parent, titulo):
    f = tk.LabelFrame(parent, text=f"  {titulo}  ", bg=BG2, fg=AZUL,
                      font=FONTE_SEC, relief="flat",
                      highlightbackground=BORDA, highlightthickness=1,
                      padx=8, pady=6)
    f.pack(fill="x", padx=10, pady=(6, 2))
    return f


def _sugerir_nome(caminho):
    """Extrai nome do experimento do nome do arquivo, removendo prefixo Dados_."""
    base = os.path.splitext(os.path.basename(caminho))[0]
    if base.lower().startswith("dados_"):
        base = base[6:]
    return base


def _calcular_beta_mb(T_gas, m_mol):
    return math.sqrt((4.04e5 * (300.0 / T_gas) * m_mol) / 2.0)


def _ler_xlsx(caminho):
    """
    Lê dados experimentais do .xlsx de dados brutos.
    Retorna lista de (tempo, sinal) filtrando pontos t=0, u/u0=1 (referências).
    """
    from openpyxl import load_workbook
    wb = load_workbook(caminho, data_only=True, read_only=True)
    if ABA_ANALISE not in wb.sheetnames:
        raise ValueError(
            f"Aba '{ABA_ANALISE}' nao encontrada no arquivo.\n"
            f"Abas disponiveis: {', '.join(wb.sheetnames)}"
        )
    ws = wb[ABA_ANALISE]
    pontos = []
    for row in ws.iter_rows(min_row=LINHA_DADOS, min_col=COL_TEMPO,
                             max_col=COL_SINAL, values_only=True):
        l, m = row[0], row[1]
        if l is None and m is None:
            break
        if l is None or m is None:
            continue
        try:
            t = float(l)
            s = float(m)
            # Filtrar pontos de referencia (t=0, u/u0=1)
            if t == 0.0 and abs(s - 1.0) < 1e-9:
                continue
            pontos.append((t, s))
        except (ValueError, TypeError):
            continue
    wb.close()
    return pontos


# ── App principal ──────────────────────────────────────────────────────────────
class MigradorApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Natalia Time - Criador de Experimento")
        self.configure(bg=BG)
        self.resizable(True, True)
        self.minsize(640, 680)
        self._build_ui()
        self._centralizar()

    def _centralizar(self):
        self.update_idletasks()
        w, h = 700, 760
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        self.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

    def _build_ui(self):
        tk.Label(self, text="Criador de Experimento",
                 bg=BG, fg=AZUL, font=("Segoe UI", 14, "bold")).pack(pady=(14, 2))
        tk.Label(self, text="Gera a pasta com dados.csv e parametros.toml para o Natalia Time",
                 bg=BG, fg=TEXTO2, font=FONTE).pack(pady=(0, 8))

        # ── Destino ───────────────────────────────────────────────────────────
        s0 = _secao(self, "Destino")
        r0 = tk.Frame(s0, bg=BG2); r0.pack(fill="x", pady=2)
        _lbl(r0, "Pasta raiz:").pack(side="left")
        self.e_destino = _entry(r0, width=44,
                                default=os.path.join(
                                    os.path.dirname(os.path.abspath(__file__)),
                                    "Dados e Parametros"))
        self.e_destino.pack(side="left", padx=(6, 4))
        _btn(r0, "...", self._procurar_destino, cor=BG3).pack(side="left")

        r0b = tk.Frame(s0, bg=BG2); r0b.pack(fill="x", pady=2)
        _lbl(r0b, "Nome do experimento:").pack(side="left")
        self.e_nome = _entry(r0b, width=30)
        self.e_nome.pack(side="left", padx=(6, 0))
        _lbl(r0b, "  (ex: O2_100eV)", fg=TEXTO2).pack(side="left")

        # ── Parametros fisicos ────────────────────────────────────────────────
        s1 = _secao(self, "Parametros fisicos")
        campos = [
            ("T_gas",    "T gas (K)",             "Temperatura do gas de expansao",                "300"),
            ("m_frag",   "m fragmento",            "Massa do fragmento ionico (ex: 16 para O+)",    "16.0"),
            ("m_mol",    "m molecula mae",         "Massa da molecula precursora (ex: 32 para O2)", "32.0"),
            ("DE",       "DE - deflexao (eV)",     "Energia de deflexao do espectrometro",          "350.0"),
            ("D",        "D - distancia (m)",      "Distancia de voo",                              "6.8"),
            ("LL",       "LL - comprimento (m)",   "Comprimento do tubo",                           "6.8"),
            ("I7_NORM",  "I7 - normaliz. (ns)",    "Tempo de normalizacao das amplitudes",          "500.0"),
            ("t_offset", "Offset de tempo (ns)",   "Correcao do tempo de aquisicao eletronica",     "600.0"),
        ]
        self.e_params = {}
        for chave, rotulo, tooltip, padrao in campos:
            r = tk.Frame(s1, bg=BG2); r.pack(fill="x", pady=3)
            lbl = _lbl(r, f"{rotulo}:"); lbl.config(width=20, anchor="w"); lbl.pack(side="left")
            e = _entry(r, width=12, default=padrao); e.pack(side="left", padx=(4, 10))
            _lbl(r, tooltip, fg=TEXTO2).pack(side="left")
            self.e_params[chave] = e

        tk.Label(s1, text="Beta_MB calculado automaticamente a partir de T e m molecula mae.",
                 bg=BG2, fg=TEXTO2, font=FONTE).pack(anchor="w", pady=(4, 0))
        tk.Label(s1, text="Energias (E_G1, E_G2, E_G3) sao definidas na interface do Natalia Time.",
                 bg=BG2, fg=AMAR, font=FONTE).pack(anchor="w", pady=(2, 0))
        tk.Label(s1,
                 text="ATENCAO: O offset de tempo (padrao 600 ns) corrige o atraso da aquisicao\n"
                      "eletronica. Ele e somado a todos os tempos dos dados brutos ao exportar.\n"
                      "Verifique sempre este valor antes de criar o experimento.",
                 bg=BG2, fg=VERM, font=FONTE, justify="left").pack(anchor="w", pady=(4, 0))

        # ── Dados - abas Automatico / Manual ─────────────────────────────────
        s2_frame = tk.Frame(self, bg=BG)
        s2_frame.pack(fill="both", expand=True, padx=10, pady=(6, 2))

        tk.Label(s2_frame, text="  Dados experimentais  ", bg=BG, fg=AZUL,
                 font=FONTE_SEC).pack(anchor="w")

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Mig.TNotebook", background=BG2, borderwidth=0)
        style.configure("Mig.TNotebook.Tab", background=BG3, foreground=TEXTO2,
                        padding=[10, 4], font=FONTE)
        style.map("Mig.TNotebook.Tab",
                  background=[("selected", BG2)],
                  foreground=[("selected", TEXTO)])

        nb = ttk.Notebook(s2_frame, style="Mig.TNotebook")
        nb.pack(fill="both", expand=True, pady=(4, 0))

        # ── Aba Automatico ────────────────────────────────────────────────────
        aba_auto = tk.Frame(nb, bg=BG2)
        nb.add(aba_auto, text="Automatico - ler do .xlsx")

        tk.Label(aba_auto,
                 text="Selecione o arquivo .xlsx de dados brutos (ex: Dados_O2_100eV.xlsx).\n"
                      "O programa le automaticamente a aba 'Analise_tempo', colunas L e M.",
                 bg=BG2, fg=TEXTO2, font=FONTE, justify="left").pack(anchor="w", padx=8, pady=(8, 6))

        r_auto = tk.Frame(aba_auto, bg=BG2); r_auto.pack(fill="x", padx=8, pady=4)
        _lbl(r_auto, "Arquivo .xlsx:").pack(side="left")
        self.e_xlsx = _entry(r_auto, width=42)
        self.e_xlsx.pack(side="left", padx=(6, 4))
        _btn(r_auto, "Procurar", self._procurar_xlsx, cor=BG3).pack(side="left")

        self.lbl_auto_status = tk.Label(aba_auto, text="", bg=BG2, fg=TEXTO2, font=FONTE)
        self.lbl_auto_status.pack(anchor="w", padx=8, pady=(2, 6))

        # ── Aba Manual ────────────────────────────────────────────────────────
        aba_manual = tk.Frame(nb, bg=BG2)
        nb.add(aba_manual, text="Manual - colar dados")

        tk.Label(aba_manual,
                 text="Cole os dados copiados do Excel (colunas Tempo e Sinal, separadas por Tab):",
                 bg=BG2, fg=TEXTO2, font=FONTE).pack(anchor="w", padx=8, pady=(8, 4))

        frame_txt = tk.Frame(aba_manual, bg=BG2)
        frame_txt.pack(fill="both", expand=True, padx=8)
        scroll_y = tk.Scrollbar(frame_txt, orient="vertical")
        scroll_y.pack(side="right", fill="y")
        scroll_x = tk.Scrollbar(frame_txt, orient="horizontal")
        scroll_x.pack(side="bottom", fill="x")
        self.txt_dados = tk.Text(frame_txt, height=8, bg=BG3, fg=TEXTO,
                                 insertbackground=TEXTO, font=("Consolas", 9),
                                 relief="flat", wrap="none",
                                 yscrollcommand=scroll_y.set,
                                 xscrollcommand=scroll_x.set,
                                 highlightbackground=BORDA, highlightthickness=1)
        self.txt_dados.pack(fill="both", expand=True)
        scroll_y.config(command=self.txt_dados.yview)
        scroll_x.config(command=self.txt_dados.xview)

        self._placeholder_ativo = True
        _ph = "100\t0.9057\n200\t0.8393\n300\t0.7706\n..."
        self.txt_dados.insert("1.0", _ph)
        self.txt_dados.config(fg=TEXTO2)
        self.txt_dados.bind("<FocusIn>",  self._limpar_placeholder)
        self.txt_dados.bind("<FocusOut>", self._repor_placeholder)
        self._placeholder_txt = _ph

        self.lbl_linhas = tk.Label(aba_manual, text="", bg=BG2, fg=TEXTO2, font=FONTE)
        self.lbl_linhas.pack(anchor="e", padx=8, pady=(2, 6))
        self.txt_dados.bind("<KeyRelease>", self._atualizar_contagem)
        self.txt_dados.bind("<<Paste>>",    lambda e: self.after(50, self._atualizar_contagem))

        self._nb_dados = nb

        # ── Botoes ────────────────────────────────────────────────────────────
        bf = tk.Frame(self, bg=BG); bf.pack(pady=10)
        _btn(bf, "Criar experimento", self._criar, cor=VERDE).pack(side="left", padx=8)
        _btn(bf, "Limpar",            self._limpar, cor=BG3).pack(side="left", padx=4)

        self.lbl_status = tk.Label(self, text="", bg=BG, fg=TEXTO2, font=FONTE)
        self.lbl_status.pack(pady=(0, 8))

    # ── Acoes ─────────────────────────────────────────────────────────────────

    def _procurar_destino(self):
        p = filedialog.askdirectory(title="Pasta raiz para os experimentos")
        if p:
            self.e_destino.delete(0, "end")
            self.e_destino.insert(0, p)

    def _procurar_xlsx(self):
        p = filedialog.askopenfilename(
            title="Selecionar arquivo de dados",
            filetypes=[("Excel", "*.xlsx *.xlsm"), ("Todos", "*.*")])
        if not p:
            return
        self.e_xlsx.delete(0, "end")
        self.e_xlsx.insert(0, p)

        # Sugerir nome
        nome_sug = _sugerir_nome(p)
        if not self.e_nome.get().strip():
            self.e_nome.delete(0, "end")
            self.e_nome.insert(0, nome_sug)

        # Contar pontos
        try:
            pontos = _ler_xlsx(p)
            self.lbl_auto_status.config(
                text=f"{len(pontos)} pontos encontrados na aba '{ABA_ANALISE}'",
                fg=VERDE)
        except Exception as ex:
            self.lbl_auto_status.config(text=f"Erro: {ex}", fg=VERM)

    def _limpar_placeholder(self, event=None):
        if self._placeholder_ativo:
            self.txt_dados.delete("1.0", "end")
            self.txt_dados.config(fg=TEXTO)
            self._placeholder_ativo = False

    def _repor_placeholder(self, event=None):
        if not self.txt_dados.get("1.0", "end").strip():
            self.txt_dados.insert("1.0", self._placeholder_txt)
            self.txt_dados.config(fg=TEXTO2)
            self._placeholder_ativo = True
            self.lbl_linhas.config(text="")

    def _atualizar_contagem(self, event=None):
        if self._placeholder_ativo:
            return
        n = len(self._parsear_dados_manual())
        self.lbl_linhas.config(
            text=f"{n} pontos validos" if n else "Nenhum ponto valido detectado",
            fg=VERDE if n >= 5 else VERM)

    def _parsear_dados_manual(self):
        texto = self.txt_dados.get("1.0", "end")
        pontos = []
        for linha in texto.splitlines():
            linha = linha.strip()
            if not linha:
                continue
            partes = linha.split("\t") if "\t" in linha else linha.replace(";", ",").split(",")
            if len(partes) < 2:
                continue
            try:
                t = float(partes[0].replace(",", "."))
                s = float(partes[1].replace(",", "."))
                pontos.append((t, s))
            except ValueError:
                continue
        return pontos

    def _limpar(self):
        for e in self.e_params.values():
            e.delete(0, "end")
        self.e_nome.delete(0, "end")
        self.e_xlsx.delete(0, "end")
        self.txt_dados.delete("1.0", "end")
        self._placeholder_ativo = False
        self._repor_placeholder()
        self.lbl_status.config(text="")
        self.lbl_auto_status.config(text="")
        self.lbl_linhas.config(text="")

    def _validar_params(self):
        valores = {}
        for chave, e in self.e_params.items():
            val_str = e.get().strip()
            if not val_str:
                messagebox.showwarning("Parametro vazio", f"O campo '{chave}' esta vazio.")
                e.focus_set(); return None
            try:
                valores[chave] = float(val_str.replace(",", "."))
            except ValueError:
                messagebox.showwarning("Valor invalido",
                    f"Valor '{val_str}' para '{chave}' nao e um numero valido.")
                e.focus_set(); return None
        return valores

    def _criar(self):
        self.lbl_status.config(text="")

        nome = self.e_nome.get().strip()
        if not nome:
            messagebox.showwarning("Nome vazio", "Informe o nome do experimento.")
            self.e_nome.focus_set(); return
        if any(c in set(r'\/:*?"<>|') for c in nome):
            messagebox.showwarning("Nome invalido",
                f"O nome '{nome}' contem caracteres invalidos para nome de pasta.")
            return

        valores = self._validar_params()
        if valores is None:
            return

        # Dados - automatico ou manual
        aba_ativa = self._nb_dados.index(self._nb_dados.select())
        if aba_ativa == 0:
            xlsx = self.e_xlsx.get().strip()
            if not xlsx:
                messagebox.showwarning("Arquivo nao selecionado",
                    "Selecione o arquivo .xlsx na aba Automatico.")
                return
            try:
                pontos = _ler_xlsx(xlsx)
            except Exception as ex:
                messagebox.showerror("Erro ao ler .xlsx", str(ex))
                return
        else:
            if self._placeholder_ativo:
                messagebox.showwarning("Dados vazios",
                    "Cole os dados experimentais na aba Manual.")
                return
            pontos = self._parsear_dados_manual()

        if len(pontos) < 5:
            messagebox.showwarning("Dados insuficientes",
                f"Apenas {len(pontos)} pontos validos detectados.\n"
                "Verifique se os dados estao no formato correto.")
            return

        destino_raiz = self.e_destino.get().strip()
        if not destino_raiz:
            messagebox.showwarning("Destino vazio", "Informe a pasta raiz de destino.")
            return

        pasta_exp = os.path.join(destino_raiz, nome)
        if os.path.exists(pasta_exp):
            if not messagebox.askyesno("Pasta existente",
                f"A pasta '{nome}' ja existe.\nDeseja sobrescrever os arquivos?"):
                return

        beta_mb  = _calcular_beta_mb(valores["T_gas"], valores["m_mol"])
        t_offset = valores["t_offset"]

        try:
            os.makedirs(pasta_exp, exist_ok=True)

            with open(os.path.join(pasta_exp, "dados.csv"), "w", encoding="utf-8") as f:
                f.write("tempo,sinal\n")
                for t, s in pontos:
                    f.write(f"{t + t_offset},{s}\n")

            with open(os.path.join(pasta_exp, "parametros.toml"), "w", encoding="utf-8") as f:
                f.write("# Parametros fisicos do experimento\n")
                f.write("# Gerado pelo Natalia Time - Criador de Experimento\n")
                f.write(f"# T_gas = {valores['T_gas']} K  ->  beta_mb = {beta_mb:.6f}\n")
                f.write(f"# Offset de tempo aplicado aos dados: {t_offset:.1f} ns (atraso da aquisicao eletronica)\n\n")
                f.write("# Parametros da molecula\n")
                f.write(f"beta_mb = {beta_mb:.6f}\n")
                f.write(f"m_frag  = {valores['m_frag']}\n")
                f.write(f"m_mol   = {valores['m_mol']}\n")
                f.write("\n# Parametros geometricos do espectrometro DETOF\n")
                f.write(f"DE      = {valores['DE']}\n")
                f.write(f"D       = {valores['D']}\n")
                f.write(f"LL      = {valores['LL']}\n")
                f.write(f"I7_NORM = {valores['I7_NORM']}\n")
                f.write("\n# Energias das gaussianas (eV)\n")
                f.write("# Definidas e gravadas pelo Natalia Time\n")
                f.write("# e_g3 = 0.7\n")
                f.write("# e_g4 = 3.0\n")
                f.write("# e_g5 = 2.0\n")

        except Exception as ex:
            messagebox.showerror("Erro ao criar", f"Nao foi possivel criar os arquivos:\n{ex}")
            return

        self.lbl_status.config(
            text=f"Experimento criado com {len(pontos)} pontos em: {pasta_exp}",
            fg=VERDE)

        if messagebox.askyesno("Pronto!",
            f"Experimento '{nome}' criado com sucesso!\n\n"
            f"  * {len(pontos)} pontos em dados.csv\n"
            f"  * Parametros em parametros.toml\n\n"
            f"Pasta: {pasta_exp}\n\nAbrir a pasta agora?"):
            try:
                os.startfile(pasta_exp)
            except Exception:
                pass


if __name__ == "__main__":
    app = MigradorApp()
    app.mainloop()
