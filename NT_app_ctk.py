"""
NT_app_ctk.py — Interface Natalia Time (customtkinter)
Uso: python NT_app_ctk.py
"""

import tkinter as tk
from tkinter import filedialog, messagebox
import threading
import subprocess
import sys
import os
import glob
import re
import tempfile
import customtkinter as ctk

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

# ── Detectar base dir ──────────────────────────────────────────────────────────
if getattr(sys, 'frozen', False):
    _base    = os.path.dirname(sys.executable)
    _meipass = getattr(sys, '_MEIPASS', _base)
else:
    _base    = os.path.dirname(os.path.abspath(__file__))
    _meipass = _base

SCRIPT_PATH    = os.path.join(_base, "analise_natalia_time.py")
if not os.path.isfile(SCRIPT_PATH):
    SCRIPT_PATH = os.path.join(_meipass, "analise_natalia_time.py")
RESULTADOS_DIR = os.path.join(_base, "Resultados")
LOGO_PATH      = os.path.join(_base, "logo_lms.png")

# ── Paleta ─────────────────────────────────────────────────────────────────────
BG      = "#0e1117"
BG2     = "#161b22"
BG3     = "#1c2128"
BORDA   = "#30363d"
AZUL    = "#388bfd"
AZUL2   = "#1f6feb"
TEXTO   = "#e6edf3"
TEXTO2  = "#7d8590"
VERDE   = "#238636"
VERDE2  = "#2ea043"
VERM    = "#b91c1c"
VERM2   = "#da3633"
AMAR    = "#d29922"
ROXO    = "#8957e5"

F_UI    = None
F_MONO  = None
F_MONO_S= None
F_TITLE = None
F_H2    = None
F_SEC   = None
F_SMALL = None

def _init_fonts():
    global F_UI, F_MONO, F_MONO_S, F_TITLE, F_H2, F_SEC, F_SMALL
    F_UI    = ctk.CTkFont("Segoe UI", 12)
    F_MONO  = ctk.CTkFont("Courier New", 11)
    F_MONO_S= ctk.CTkFont("Courier New", 10)
    F_TITLE = ctk.CTkFont("Segoe UI", 22, "bold")
    F_H2    = ctk.CTkFont("Segoe UI", 13, "bold")
    F_SEC   = ctk.CTkFont("Courier New", 11, "bold")
    F_SMALL = ctk.CTkFont("Segoe UI", 10)


def _fval(s):
    return float(str(s).strip().replace(",", "."))


# ── Helpers visuais ────────────────────────────────────────────────────────────

def secao_frame(parent, titulo, icon=""):
    """Card com borda esquerda colorida e título."""
    outer = ctk.CTkFrame(parent, fg_color=BG2, corner_radius=8,
                         border_width=1, border_color=BORDA)
    outer.pack(fill="x", padx=12, pady=5)

    header = ctk.CTkFrame(outer, fg_color="transparent", height=32)
    header.pack(fill="x", padx=14, pady=(10, 4))
    header.pack_propagate(False)

    accent = ctk.CTkFrame(header, fg_color=AZUL, width=3, corner_radius=2)
    accent.pack(side="left", fill="y", padx=(0, 8))

    ctk.CTkLabel(header, text=f"{icon}  {titulo}" if icon else titulo,
                 font=F_SEC, text_color=AZUL,
                 anchor="w").pack(side="left", fill="y")

    body = ctk.CTkFrame(outer, fg_color="transparent")
    body.pack(fill="x", padx=14, pady=(0, 12))
    return body, outer


def campo(parent, label, default="", width=110, tooltip=""):
    """Label + Entry empilhados verticalmente."""
    col = ctk.CTkFrame(parent, fg_color="transparent")
    ctk.CTkLabel(col, text=label, font=F_SMALL, text_color=TEXTO2,
                 anchor="w").pack(anchor="w")
    e = ctk.CTkEntry(col, width=width, fg_color=BG3,
                     border_color=BORDA, border_width=1,
                     text_color=TEXTO, font=F_MONO_S)
    e.insert(0, default)
    e.pack()
    return col, e


def btn_primary(parent, texto, cmd, width=130):
    return ctk.CTkButton(parent, text=texto, command=cmd, width=width,
                         fg_color=AZUL, hover_color=AZUL2,
                         font=F_UI, corner_radius=6)


def btn_success(parent, texto, cmd, width=150):
    return ctk.CTkButton(parent, text=texto, command=cmd, width=width,
                         fg_color=VERDE, hover_color=VERDE2,
                         font=ctk.CTkFont("Segoe UI", 13, "bold"),
                         corner_radius=6)


def btn_danger(parent, texto, cmd, width=120):
    return ctk.CTkButton(parent, text=texto, command=cmd, width=width,
                         fg_color=VERM, hover_color=VERM2,
                         font=F_UI, corner_radius=6)


def btn_ghost(parent, texto, cmd, width=110):
    return ctk.CTkButton(parent, text=texto, command=cmd, width=width,
                         fg_color=BG3, hover_color=BORDA,
                         border_width=1, border_color=BORDA,
                         text_color=TEXTO, font=F_MONO_S,
                         corner_radius=6)


def help_popup(titulo, texto):
    """Retorna função que abre popup de ajuda."""
    def _abrir():
        win = ctk.CTkToplevel()
        win.title(titulo)
        win.geometry("480x320")
        win.configure(fg_color=BG2)
        win.resizable(False, False)
        win.grab_set()
        ctk.CTkLabel(win, text=titulo, font=F_H2,
                     text_color=AZUL).pack(padx=20, pady=(16, 4))
        box = ctk.CTkTextbox(win, fg_color=BG3, font=F_SMALL,
                             border_width=0, wrap="word")
        box.pack(fill="both", expand=True, padx=20, pady=8)
        box.insert("end", texto)
        box.configure(state="disabled")
        ctk.CTkButton(win, text="Fechar", command=win.destroy,
                      fg_color=AZUL, hover_color=AZUL2,
                      width=100, corner_radius=6).pack(pady=(0, 16))
    return _abrir


# ── Gerador de config ──────────────────────────────────────────────────────────
def gerar_config_python(cfg: dict) -> str:
    def py(v):
        if v is None:           return "None"
        if isinstance(v, bool): return str(v)
        if isinstance(v, str):  return repr(v)
        return repr(v)

    seq = cfg.get("sequencia", [])
    if seq:
        seq_lines = "SEQUENCIA = [\n"
        for item in seq:
            ativas = item.get("curvas_ativas", {})
            seq_lines += "    {\n"
            seq_lines += f"        'planilha': {repr(item['planilha'])},\n"
            seq_lines += f"        'curvas_ativas': {ativas},\n"
            seq_lines += "    },\n"
        seq_lines += "]"
    else:
        seq_lines = "SEQUENCIA = []"

    return f"""
SCORE_MINIMO = {py(cfg['score_minimo'])}
DESVIO_ALVO  = {py(cfg['desvio_alvo'])}
ITERACOES_BETA  = {py(cfg['iteracoes_beta'])}
ITERACOES_PESOS = {py(cfg['iteracoes_pesos'])}
SEMENTE = {py(cfg['semente'])}
N_PROCESSOS = {py(cfg['n_processos'])}
ARQUIVO_PARAR = "PARAR.txt"

INTERVALOS_BETA = {{
    "beta_exp": ({py(cfg['beta_exp_lo'])}, {py(cfg['beta_exp_hi'])}),
    "beta_g3":  ({py(cfg['beta_g1_lo'])}, {py(cfg['beta_g1_hi'])}),
    "beta_g4":  ({py(cfg['beta_g2_lo'])}, {py(cfg['beta_g2_hi'])}),
    "beta_g5":  ({py(cfg['beta_g3_lo'])}, {py(cfg['beta_g3_hi'])}),
}}

CURVAS_PESO_ATIVAS = {{
    "p_mb":  {py(cfg['ativa_mb'])},
    "p_exp": {py(cfg['ativa_exp'])},
    "p_g3":  {py(cfg['ativa_g1'])},
    "p_g4":  {py(cfg['ativa_g2'])},
    "p_g5":  {py(cfg['ativa_g3'])},
}}

ENERGIAS_VALOR = {{
    "e_g3": {py(cfg['e_g1_valor'])},
    "e_g4": {py(cfg['e_g2_valor'])},
    "e_g5": {py(cfg['e_g3_valor'])},
}}

ENERGIA_NM_LIVRE = {{
    "e_g3": {py(cfg['e_g1_nm_livre'])},
    "e_g4": {py(cfg['e_g2_nm_livre'])},
    "e_g5": {py(cfg['e_g3_nm_livre'])},
}}

INTERVALOS_ENERGIA = {{
    "e_g3": (0.1, 6.0), "e_g4": (0.1, 6.0), "e_g5": (0.1, 6.0),
}}

PESO_MIN = {{
    "p_mb":  {py(cfg['peso_min_mb'])},
    "p_exp": None, "p_g3": None, "p_g4": None, "p_g5": None,
}}

SALVAR_GRAFICOS_TOP = {py(cfg['salvar_graficos_top'])}
VIEWER_TOP          = 50

USAR_NELDER_MEAD   = {py(cfg['usar_nm'])}
NM_TOP_CANDIDATOS  = {py(cfg['nm_top_candidatos'])}
NM_MAX_ITER        = {py(cfg['nm_max_iter'])}
NM_TOLERANCIA      = {py(cfg['nm_tolerancia'])}
NM_MAX_REINICIAR   = {py(cfg['nm_max_reiniciar'])}
NM_PERTURB_ESCALA  = {py(cfg['nm_perturb_escala'])}

NM_USAR_CLUSTERS   = {py(cfg['nm_usar_clusters'])}
NM_CLUSTER_LIMIAR  = {py(cfg['nm_cluster_limiar'])}
NM_CLUSTERS_MAX    = {py(cfg['nm_clusters_max'])}

BUSCA_DUAS_FASES   = False
ADAPTAR_INTERVALOS = False
MODO_VALIDACAO     = False
VALIDACAO_REFINAR_NM = False

{seq_lines}

EXCEL_PATH          = {repr(cfg['excel_path'])}
RESULTADOS_DIR      = {repr(cfg['resultados_dir'])}
"""


def criar_script_temporario(script_path, cfg):
    with open(script_path, "r", encoding="utf-8") as f:
        codigo = f.read()
    m2s = re.search(r'#\s*={10,}\s*\n#\s*SEÇÃO 2', codigo)
    m3s = re.search(r'#\s*={10,}\s*\n#\s*SEÇÃO 3', codigo)
    if not m2s or not m3s:
        raise ValueError("Marcadores SEÇÃO 2 / SEÇÃO 3 não encontrados no script.")
    novo_cfg  = gerar_config_python(cfg)
    novo_codigo = codigo[:m2s.start()] + novo_cfg + "\n" + codigo[m3s.start():]
    fd, tmp = tempfile.mkstemp(prefix="tof_run_", suffix=".py")
    with os.fdopen(fd, "w", encoding="utf-8") as f:
        f.write(novo_codigo)
    return tmp


# ══════════════════════════════════════════════════════════════════════════════
# APLICAÇÃO PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════
class NTApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        _init_fonts()
        self.title("Natalia Time")
        self.geometry("1000x720")
        self.minsize(820, 620)
        self.configure(fg_color=BG)

        self.rodando        = False
        self.proc_ref       = None
        self.ultima_pasta   = None
        self.sequencia      = []
        self._log_q         = None
        self._res_q         = None
        self._logo_pil      = self._carregar_logo()
        self._sel_historico = {}   # pasta → BooleanVar (seleção para comparação)

        self._build_header()
        self._build_topbar()
        self._build_tabs()

    # ── Logo ───────────────────────────────────────────────────────────────────
    @staticmethod
    def _carregar_logo():
        if not os.path.isfile(LOGO_PATH):
            return None
        try:
            from PIL import Image as _PImg
            return _PImg.open(LOGO_PATH)
        except Exception:
            return None

    # ── Header ─────────────────────────────────────────────────────────────────
    def _build_header(self):
        hdr = ctk.CTkFrame(self, fg_color=BG2, corner_radius=0, height=52)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)

        # Linha decorativa na base
        line = ctk.CTkFrame(hdr, fg_color=BORDA, height=1, corner_radius=0)
        line.pack(side="bottom", fill="x")

        # Logo LMS — lado direito do header
        if self._logo_pil is not None:
            h_logo = 34
            w_logo = int(h_logo * self._logo_pil.width / self._logo_pil.height)
            self._logo_img_hdr = ctk.CTkImage(self._logo_pil, size=(w_logo, h_logo))
            ctk.CTkLabel(hdr, image=self._logo_img_hdr, text="",
                         fg_color="transparent").pack(side="right", padx=16)

        inner = ctk.CTkFrame(hdr, fg_color="transparent")
        inner.pack(side="left", padx=18, fill="y")

        ctk.CTkLabel(inner, text="⚗  Natalia Time",
                     font=ctk.CTkFont("Courier New", 16, "bold"),
                     text_color=AZUL).pack(side="left", pady=14)
        ctk.CTkLabel(inner, text=" — Análise de Espectroscopia DETOF",
                     font=ctk.CTkFont("Segoe UI", 10),
                     text_color=TEXTO2).pack(side="left")

    # ── Barra de experimento ───────────────────────────────────────────────────
    def _build_topbar(self):
        bar = ctk.CTkFrame(self, fg_color=BG3, corner_radius=0, height=48)
        bar.pack(fill="x")
        bar.pack_propagate(False)

        inner = ctk.CTkFrame(bar, fg_color="transparent")
        inner.pack(side="left", padx=14, fill="y")

        ctk.CTkLabel(inner, text="Experimento:", font=F_SMALL,
                     text_color=TEXTO2).pack(side="left", padx=(0, 8))

        self.entry_planilha = ctk.CTkEntry(inner, width=500,
                                           placeholder_text="Selecione .xlsm, .xlsx ou dados.csv …",
                                           fg_color=BG2, border_color=BORDA, border_width=1,
                                           text_color=TEXTO, font=F_MONO_S)
        self.entry_planilha.pack(side="left", padx=(0, 8))

        btn_ghost(inner, "📂  Procurar", self._procurar_planilha, width=120).pack(side="left")

        self.lbl_formato = ctk.CTkLabel(inner, text="", font=F_SMALL, text_color=TEXTO2)
        self.lbl_formato.pack(side="left", padx=(12, 0))

        # Aviso OneDrive/Dropbox
        _sync = ["dropbox", "onedrive", "google drive", "googledrive", "icloud"]
        if any(p in _base.lower().replace("\\", "/") for p in _sync):
            ctk.CTkLabel(bar,
                         text="⚠  Pasta sincronizada detectada — pause a sincronização antes de rodar",
                         font=F_SMALL, text_color=AMAR).pack(side="right", padx=14)

    # ── Abas ───────────────────────────────────────────────────────────────────
    def _build_tabs(self):
        tabs = ctk.CTkTabview(self, fg_color=BG,
                              segmented_button_fg_color=BG2,
                              segmented_button_selected_color=AZUL,
                              segmented_button_unselected_color=BG2,
                              segmented_button_selected_hover_color=AZUL2,
                              segmented_button_unselected_hover_color=BG3,
                              text_color=TEXTO2,
                              text_color_disabled=TEXTO2)
        tabs.pack(fill="both", expand=True)

        nomes = ["🏠  Início", "⚙  Configuração", "📋  Sequência",
                 "▶  Executar", "📊  Resultados", "🕒  Histórico"]
        for n in nomes:
            tabs.add(n)

        self._aba_inicio(tabs.tab("🏠  Início"))
        self._aba_config(tabs.tab("⚙  Configuração"))
        self._aba_sequencia(tabs.tab("📋  Sequência"))
        self._aba_executar(tabs.tab("▶  Executar"))
        self._aba_resultados(tabs.tab("📊  Resultados"))
        self._aba_historico(tabs.tab("🕒  Histórico"))

    # ── Aba Início ─────────────────────────────────────────────────────────────
    def _aba_inicio(self, tab):
        tab.configure(fg_color=BG)
        scroll = ctk.CTkScrollableFrame(tab, fg_color=BG, corner_radius=0,
                                        scrollbar_button_color=BG3)
        scroll.pack(fill="both", expand=True)

        centro = ctk.CTkFrame(scroll, fg_color="transparent")
        centro.pack(pady=(30, 0), padx=80, fill="x")

        if self._logo_pil is not None:
            h_logo = 78
            w_logo = int(h_logo * self._logo_pil.width / self._logo_pil.height)
            self._logo_img_inicio = ctk.CTkImage(self._logo_pil, size=(w_logo, h_logo))
            ctk.CTkLabel(centro, image=self._logo_img_inicio, text="",
                         fg_color="transparent").pack(pady=(0, 14))

        ctk.CTkLabel(centro, text="Natalia Time", font=F_TITLE,
                     text_color=AZUL).pack()
        ctk.CTkLabel(centro, text="Ajuste automático de espectros DETOF de fragmentação por impacto de elétron",
                     font=ctk.CTkFont("Segoe UI", 11),
                     text_color=TEXTO2).pack(pady=(2, 28))

        # Passos
        ctk.CTkLabel(centro, text="Como começar", font=F_H2,
                     anchor="w").pack(anchor="w", pady=(0, 8))

        for icon, titulo, desc in [
            ("📂", "Selecione o experimento",
             "Clique em Procurar na barra superior e escolha sua planilha (.xlsm / .xlsx) "
             "ou o arquivo dados.csv da pasta do experimento (novo formato)."),
            ("⚙",  "Configure as curvas",
             "Na aba Configuração, verifique quais curvas estão ativas (MB, Exponencial, "
             "Gaussianas 1/2/3) e ajuste os intervalos de beta. "
             "No formato CSV, as energias precisam estar preenchidas ou gravadas."),
            ("▶",  "Execute a análise",
             "Na aba Executar, clique em Iniciar. O log mostrará o progresso em tempo real. "
             "Você pode parar a qualquer momento — tudo que foi encontrado é salvo."),
            ("📊", "Veja os resultados",
             "Ao terminar, os melhores ajustes aparecem na aba Resultados. "
             "Gráficos e o relatório Excel são salvos automaticamente na pasta Resultados/."),
        ]:
            card = ctk.CTkFrame(centro, fg_color=BG2, corner_radius=8,
                                border_width=1, border_color=BORDA)
            card.pack(fill="x", pady=4)
            row = ctk.CTkFrame(card, fg_color="transparent")
            row.pack(fill="x", padx=16, pady=12)

            ctk.CTkLabel(row, text=icon, font=ctk.CTkFont("Segoe UI", 16),
                         text_color=AZUL, width=32).pack(side="left", anchor="n", pady=2)
            col = ctk.CTkFrame(row, fg_color="transparent")
            col.pack(side="left", fill="x", expand=True, padx=(10, 0))
            ctk.CTkLabel(col, text=titulo, font=ctk.CTkFont("Segoe UI", 11, "bold"),
                         anchor="w").pack(anchor="w")
            ctk.CTkLabel(col, text=desc, font=F_SMALL, text_color=TEXTO2,
                         anchor="w", justify="left", wraplength=620).pack(anchor="w", pady=(2, 0))

        # Formatos
        ctk.CTkLabel(centro, text="Formatos de entrada aceitos", font=F_H2,
                     anchor="w").pack(anchor="w", pady=(22, 8))

        for icon, fmt, desc in [
            ("📊", "Excel (.xlsm / .xlsx)",
             "Formato original. A aba e a linha de início dos dados são detectadas "
             "automaticamente — basta selecionar o arquivo."),
            ("📄", "CSV + TOML",
             "Pasta com dados.csv e parametros.toml. Criada com o programa Migrador. "
             "Não requer o Excel instalado na máquina."),
        ]:
            card = ctk.CTkFrame(centro, fg_color=BG2, corner_radius=8,
                                border_width=1, border_color=BORDA)
            card.pack(fill="x", pady=4)
            inner = ctk.CTkFrame(card, fg_color="transparent")
            inner.pack(fill="x", padx=16, pady=10)
            ctk.CTkLabel(inner, text=f"{icon}  {fmt}",
                         font=ctk.CTkFont("Segoe UI", 10, "bold"),
                         text_color=AZUL).pack(anchor="w")
            ctk.CTkLabel(inner, text=desc, font=F_SMALL, text_color=TEXTO2,
                         wraplength=620, justify="left").pack(anchor="w", pady=(2, 0))

        ctk.CTkFrame(centro, fg_color="transparent", height=30).pack()

    # ── Aba Configuração ───────────────────────────────────────────────────────
    def _aba_config(self, tab):
        tab.configure(fg_color=BG)

        scroll = ctk.CTkScrollableFrame(tab, fg_color=BG, corner_radius=0,
                                        scrollbar_button_color=BG3)
        scroll.pack(fill="both", expand=True)

        cols = ctk.CTkFrame(scroll, fg_color="transparent")
        cols.pack(fill="x")

        left  = ctk.CTkFrame(cols, fg_color="transparent")
        left.pack(side="left", fill="both", expand=True)
        right = ctk.CTkFrame(cols, fg_color="transparent")
        right.pack(side="right", fill="both", expand=True)

        # ── Busca aleatória ──
        body, outer = secao_frame(left, "Busca aleatória", "🔍")
        ctk.CTkButton(outer, text="?", width=28, height=22,
                      fg_color=BG3, hover_color=AZUL, text_color=AZUL,
                      font=ctk.CTkFont("Courier New", 10, "bold"), corner_radius=4,
                      command=help_popup("Busca aleatória",
                          "O programa sorteia combinações aleatórias de parâmetros e avalia o ajuste.\n\n"
                          "• Iterações beta: quantas combinações de beta (largura das curvas) sortear.\n"
                          "• Iterações pesos: para cada beta, quantas combinações de pesos testar.\n"
                          "  Total de cálculos = Iterações beta × Iterações pesos.\n\n"
                          "• R² mínimo: limiar de qualidade — só resultados acima são considerados válidos.\n"
                          "  (R² = 1.0 seria ajuste perfeito; tente ≥ 0.99)\n\n"
                          "• RMSE alvo: erro quadrático médio desejado. A busca para ao atingi-lo.\n\n"
                          "• Semente: número fixo = resultados reproduzíveis. Vazio = aleatório.\n"
                          "• Núcleos CPU: quantos núcleos usar em paralelo.")
                      ).place(relx=1.0, x=-10, y=6, anchor="ne")

        row1 = ctk.CTkFrame(body, fg_color="transparent")
        row1.pack(fill="x", pady=(4, 0))

        c, self.e_iter_beta = campo(row1, "Iterações beta", "1500")
        c.pack(side="left", padx=(0, 16))
        c, self.e_iter_pesos = campo(row1, "Iterações pesos", "1500")
        c.pack(side="left", padx=(0, 16))
        c, self.e_score = campo(row1, "R² mínimo", "0.99", width=90)
        c.pack(side="left", padx=(0, 16))
        c, self.e_rmse = campo(row1, "RMSE alvo", "0.04", width=90)
        c.pack(side="left")

        row2 = ctk.CTkFrame(body, fg_color="transparent")
        row2.pack(fill="x", pady=(10, 0))

        c, self.e_semente = campo(row2, "Semente (vazio = aleatório)", "", width=130)
        c.pack(side="left", padx=(0, 24))

        cpu = os.cpu_count() or 1
        cpu_col = ctk.CTkFrame(row2, fg_color="transparent")
        cpu_col.pack(side="left")
        self.lbl_cpu_val = ctk.CTkLabel(cpu_col,
                                         text=f"Núcleos CPU: {cpu} / {cpu}",
                                         font=F_SMALL, text_color=TEXTO2, anchor="w")
        self.lbl_cpu_val.pack(anchor="w")
        self.slider_cpu = ctk.CTkSlider(cpu_col, from_=1, to=cpu, width=180,
                                         button_color=AZUL, button_hover_color=AZUL2,
                                         progress_color=AZUL, fg_color=BORDA)
        self.slider_cpu.set(cpu)
        self.slider_cpu.configure(command=lambda v: self.lbl_cpu_val.configure(
            text=f"Núcleos CPU: {int(v)} / {cpu}"))
        self.slider_cpu.pack()

        # ── Nelder-Mead ──
        body2, outer2 = secao_frame(left, "Nelder-Mead", "🎯")
        ctk.CTkButton(outer2, text="?", width=28, height=22,
                      fg_color=BG3, hover_color=AZUL, text_color=AZUL,
                      font=ctk.CTkFont("Courier New", 10, "bold"), corner_radius=4,
                      command=help_popup("Refinamento Nelder-Mead",
                          "Após a busca aleatória encontrar bons candidatos, o Nelder-Mead refina "
                          "cada um procurando o mínimo local mais preciso.\n\n"
                          "• Max iter/candidato: limite de passos por candidato.\n"
                          "• Máx reinícios: se travar, perturba e tenta novamente.\n"
                          "• Perturbação: intensidade da perturbação (0.10 = ±10%).\n"
                          "• Tolerância: critério de convergência (10⁻⁹ é muito preciso).\n\n"
                          "• Clusters: agrupa candidatos por similaridade antes de refinar.")
                      ).place(relx=1.0, x=-10, y=6, anchor="ne")

        self.v_usar_nm = tk.BooleanVar(value=True)
        cb_nm = ctk.CTkCheckBox(body2, text="Ativar refinamento Nelder-Mead",
                                 variable=self.v_usar_nm,
                                 checkbox_width=16, checkbox_height=16,
                                 checkmark_color="white", fg_color=AZUL,
                                 hover_color=AZUL2, font=F_UI)
        cb_nm.pack(anchor="w", pady=(2, 8))

        row3 = ctk.CTkFrame(body2, fg_color="transparent")
        row3.pack(fill="x")

        c, self.e_nm_iter   = campo(row3, "Max iter/candidato", "2000", width=110)
        c.pack(side="left", padx=(0, 16))
        c, self.e_nm_reinic = campo(row3, "Máx reinícios", "3", width=80)
        c.pack(side="left", padx=(0, 16))
        c, self.e_nm_perturb= campo(row3, "Perturbação reinício", "0.10", width=90)
        c.pack(side="left")

        row4 = ctk.CTkFrame(body2, fg_color="transparent")
        row4.pack(fill="x", pady=(10, 0))

        tol_col = ctk.CTkFrame(row4, fg_color="transparent")
        tol_col.pack(side="left", padx=(0, 24))
        self.lbl_nm_tol = ctk.CTkLabel(tol_col, text="Tolerância: 10⁻⁹",
                                        font=F_SMALL, text_color=TEXTO2, anchor="w")
        self.lbl_nm_tol.pack(anchor="w")
        self._nm_tol_val = -9
        self.slider_nm_tol = ctk.CTkSlider(tol_col, from_=-15, to=-3, width=170,
                                            button_color=ROXO, button_hover_color="#6e40c9",
                                            progress_color=ROXO, fg_color=BORDA)
        self.slider_nm_tol.set(-9)
        self.slider_nm_tol.configure(command=self._on_tol_change)
        self.slider_nm_tol.pack()

        cl_col = ctk.CTkFrame(row4, fg_color="transparent")
        cl_col.pack(side="left")
        self.v_nm_clusters = tk.BooleanVar(value=True)
        ctk.CTkCheckBox(cl_col, text="Usar clusters",
                         variable=self.v_nm_clusters,
                         checkbox_width=16, checkbox_height=16,
                         checkmark_color="white", fg_color=AZUL,
                         hover_color=AZUL2, font=F_UI).pack(anchor="w", pady=(14, 4))

        row5 = ctk.CTkFrame(body2, fg_color="transparent")
        row5.pack(fill="x", pady=(8, 0))
        c, self.e_nm_limiar = campo(row5, "Limiar cluster", "0.25", width=80)
        c.pack(side="left", padx=(0, 16))
        c, self.e_nm_max_cl = campo(row5, "Máx clusters", "10", width=80)
        c.pack(side="left", padx=(0, 16))
        c, self.e_nm_top    = campo(row5, "Top candidatos", "5", width=80)
        c.pack(side="left")

        # ── Curvas ativas ──
        body3, outer3 = secao_frame(right, "Curvas ativas", "📈")
        ctk.CTkButton(outer3, text="?", width=28, height=22,
                      fg_color=BG3, hover_color=AZUL, text_color=AZUL,
                      font=ctk.CTkFont("Courier New", 10, "bold"), corner_radius=4,
                      command=help_popup("Curvas ativas",
                          "O espectro é decomposto em soma de curvas:\n\n"
                          "• Maxwell-Boltzmann (MB): gás de fundo. Sempre ativa.\n"
                          "• Exponencial: fragmentos com distribuição exponencial de energia.\n"
                          "• Gaussiana 1/2/3: fragmentos com picos de energia específicos.\n\n"
                          "• Peso mínimo MB: garante que MB contribua com pelo menos este valor.")
                      ).place(relx=1.0, x=-10, y=6, anchor="ne")

        cb_row = ctk.CTkFrame(body3, fg_color="transparent")
        cb_row.pack(fill="x", pady=(2, 8))

        self.v_ativa_mb = tk.BooleanVar(value=True)
        ctk.CTkCheckBox(cb_row, text="Maxwell-Boltzmann",
                        variable=self.v_ativa_mb,
                        checkbox_width=16, checkbox_height=16,
                        checkmark_color="white", fg_color=AZUL,
                        hover_color=AZUL2, font=F_UI).pack(side="left", padx=(0, 16))

        self.v_ativa_exp = tk.BooleanVar(value=True)
        self.v_ativa_g1  = tk.BooleanVar(value=True)
        self.v_ativa_g2  = tk.BooleanVar(value=True)
        self.v_ativa_g3  = tk.BooleanVar(value=False)

        for texto, var in [("Exponencial", self.v_ativa_exp),
                            ("Gaussiana 1", self.v_ativa_g1),
                            ("Gaussiana 2", self.v_ativa_g2),
                            ("Gaussiana 3", self.v_ativa_g3)]:
            ctk.CTkCheckBox(cb_row, text=texto, variable=var,
                             checkbox_width=16, checkbox_height=16,
                             checkmark_color="white", fg_color=AZUL,
                             hover_color=AZUL2, font=F_UI).pack(side="left", padx=(0, 16))

        peso_row = ctk.CTkFrame(body3, fg_color="transparent")
        peso_row.pack(fill="x")
        c, self.e_peso_mb = campo(peso_row, "Peso mínimo MB", "0.01", width=90)
        c.pack(side="left")

        # ── Energias ──
        body4, outer4 = secao_frame(right, "Energias das gaussianas", "⚡")
        ctk.CTkButton(outer4, text="?", width=28, height=22,
                      fg_color=BG3, hover_color=AZUL, text_color=AZUL,
                      font=ctk.CTkFont("Courier New", 10, "bold"), corner_radius=4,
                      command=help_popup("Energias das gaussianas",
                          "Cada gaussiana é centrada numa energia de translação (eV).\n\n"
                          "• Vazio: usa o valor lido da planilha Excel ou do parametros.toml.\n"
                          "• Valor numérico: sobrescreve o valor para esta execução.\n\n"
                          "• NM livre: permite ao Nelder-Mead ajustar a energia durante o refinamento "
                          "(intervalo: 0.1 – 6.0 eV).\n\n"
                          "• Gravar (CSV): grava o valor no parametros.toml para uso futuro.")
                      ).place(relx=1.0, x=-10, y=6, anchor="ne")

        self.e_energias  = {}
        self.v_nm_livre  = {}
        self.v_gravar_e  = {}
        self.cb_gravar_e = {}

        for g, nome in [("g1", "Gaussiana 1"), ("g2", "Gaussiana 2"), ("g3", "Gaussiana 3")]:
            row = ctk.CTkFrame(body4, fg_color="transparent")
            row.pack(fill="x", pady=3)

            ctk.CTkLabel(row, text=f"{nome}:", font=F_SMALL,
                         text_color=TEXTO2, width=90, anchor="w").pack(side="left")

            e = ctk.CTkEntry(row, width=90, fg_color=BG3,
                              border_color=BORDA, border_width=1,
                              placeholder_text="eV",
                              text_color=TEXTO, font=F_MONO_S)
            e.pack(side="left", padx=(0, 12))
            self.e_energias[g] = e

            v_nm = tk.BooleanVar(value=False)
            ctk.CTkCheckBox(row, text="NM livre", variable=v_nm,
                             checkbox_width=14, checkbox_height=14,
                             checkmark_color="white", fg_color=ROXO,
                             hover_color="#6e40c9", font=F_SMALL).pack(side="left", padx=(0, 8))
            self.v_nm_livre[g] = v_nm

            v_gr = tk.BooleanVar(value=False)
            cb_gr = ctk.CTkCheckBox(row, text="Gravar", variable=v_gr,
                                     checkbox_width=14, checkbox_height=14,
                                     checkmark_color="white", fg_color=VERDE,
                                     hover_color=VERDE2, font=F_SMALL)
            cb_gr.pack_forget()
            self.v_gravar_e[g]  = v_gr
            self.cb_gravar_e[g] = cb_gr

        # ── Intervalos de beta ──
        body5, outer5 = secao_frame(right, "Intervalos de beta", "📐")
        ctk.CTkButton(outer5, text="?", width=28, height=22,
                      fg_color=BG3, hover_color=AZUL, text_color=AZUL,
                      font=ctk.CTkFont("Courier New", 10, "bold"), corner_radius=4,
                      command=help_popup("Intervalos de beta",
                          "Beta (β) controla a largura de cada curva no espectro DETOF.\n\n"
                          "Estes intervalos definem os limites do sorteio aleatório:\n"
                          "• Min: valor mínimo de β a sortear\n"
                          "• Max: valor máximo de β a sortear\n\n"
                          "Os padrões cobrem bem a maioria dos experimentos.")
                      ).place(relx=1.0, x=-10, y=6, anchor="ne")

        self.e_beta_lo = {}
        self.e_beta_hi = {}
        labels_beta   = {"exp": "Exponencial", "g1": "Gaussiana 1",
                          "g2": "Gaussiana 2",  "g3": "Gaussiana 3"}
        defaults_beta = {"exp": (50, 600), "g1": (100, 1200),
                          "g2": (100, 1200), "g3": (100, 1200)}

        hdr_row = ctk.CTkFrame(body5, fg_color="transparent")
        hdr_row.pack(fill="x", pady=(0, 4))
        for txt, w in [("Curva", 110), ("Min", 80), ("Max", 80)]:
            ctk.CTkLabel(hdr_row, text=txt, font=F_SMALL,
                         text_color=TEXTO2, width=w, anchor="w").pack(side="left", padx=(0, 8))

        for k, (lo, hi) in defaults_beta.items():
            row = ctk.CTkFrame(body5, fg_color="transparent")
            row.pack(fill="x", pady=2)
            ctk.CTkLabel(row, text=labels_beta[k], font=F_SMALL,
                         text_color=TEXTO, width=110, anchor="w").pack(side="left", padx=(0, 8))
            e_lo = ctk.CTkEntry(row, width=80, fg_color=BG3,
                                 border_color=BORDA, border_width=1,
                                 text_color=TEXTO, font=F_MONO_S)
            e_lo.insert(0, str(lo))
            e_lo.pack(side="left", padx=(0, 8))
            e_hi = ctk.CTkEntry(row, width=80, fg_color=BG3,
                                 border_color=BORDA, border_width=1,
                                 text_color=TEXTO, font=F_MONO_S)
            e_hi.insert(0, str(hi))
            e_hi.pack(side="left")
            self.e_beta_lo[k] = e_lo
            self.e_beta_hi[k] = e_hi

        # ── Saídas ──
        body6, _ = secao_frame(right, "Saídas", "💾")
        row6 = ctk.CTkFrame(body6, fg_color="transparent")
        row6.pack(fill="x")
        c, self.e_salvar_top = campo(row6, "Gráficos PNG a salvar (vazio = todos)", "20", width=90)
        c.pack(side="left")

    # ── Aba Sequência ──────────────────────────────────────────────────────────
    def _aba_sequencia(self, tab):
        tab.configure(fg_color=BG)
        scroll = ctk.CTkScrollableFrame(tab, fg_color=BG, corner_radius=0,
                                        scrollbar_button_color=BG3)
        scroll.pack(fill="both", expand=True)

        body, _ = secao_frame(scroll, "Modo sequência", "📋")
        ctk.CTkLabel(_.winfo_children()[0].winfo_children()[-1] if False else _,
                     text="", fg_color="transparent")

        self.v_usar_seq = tk.BooleanVar(value=False)
        ctk.CTkCheckBox(body, text="Ativar modo sequência",
                         variable=self.v_usar_seq,
                         checkbox_width=16, checkbox_height=16,
                         checkmark_color="white", fg_color=AZUL,
                         hover_color=AZUL2, font=F_UI).pack(anchor="w", pady=(0, 4))
        ctk.CTkLabel(body,
                     text="Processa múltiplas planilhas automaticamente. "
                          "Quando ativo, a planilha da barra superior é ignorada.",
                     font=F_SMALL, text_color=TEXTO2, wraplength=500,
                     justify="left").pack(anchor="w")

        # Adicionar
        body2, _ = secao_frame(scroll, "Adicionar planilha", "➕")
        add_row = ctk.CTkFrame(body2, fg_color="transparent")
        add_row.pack(fill="x", pady=(0, 8))
        ctk.CTkLabel(add_row, text="Caminho:", font=F_SMALL,
                     text_color=TEXTO2).pack(side="left", padx=(0, 8))
        self.entry_sq_path = ctk.CTkEntry(add_row, width=380, fg_color=BG3,
                                           border_color=BORDA, border_width=1,
                                           text_color=TEXTO, font=F_MONO_S)
        self.entry_sq_path.pack(side="left", padx=(0, 8))
        btn_ghost(add_row, "Procurar", self._procurar_seq, width=100).pack(side="left")

        cb_row = ctk.CTkFrame(body2, fg_color="transparent")
        cb_row.pack(fill="x", pady=(0, 8))
        ctk.CTkLabel(cb_row, text="Curvas:", font=F_SMALL,
                     text_color=TEXTO2).pack(side="left", padx=(0, 8))
        self.v_sq_exp = tk.BooleanVar(value=True)
        self.v_sq_g1  = tk.BooleanVar(value=True)
        self.v_sq_g2  = tk.BooleanVar(value=True)
        self.v_sq_g3  = tk.BooleanVar(value=False)
        for txt, var in [("EXP", self.v_sq_exp), ("G1", self.v_sq_g1),
                          ("G2", self.v_sq_g2),  ("G3", self.v_sq_g3)]:
            ctk.CTkCheckBox(cb_row, text=txt, variable=var,
                             checkbox_width=14, checkbox_height=14,
                             checkmark_color="white", fg_color=AZUL,
                             hover_color=AZUL2, font=F_UI).pack(side="left", padx=(0, 12))

        btn_primary(body2, "➕  Adicionar", self._adicionar_seq, width=140).pack(anchor="w")

        # Fila
        _, outer_fila = secao_frame(scroll, "Fila", "📄")
        self.frame_seq_body = ctk.CTkFrame(outer_fila, fg_color="transparent")
        self.frame_seq_body.pack(fill="x", padx=14, pady=(0, 12))
        self._atualizar_lista_seq()

    # ── Aba Executar ───────────────────────────────────────────────────────────
    def _aba_executar(self, tab):
        tab.configure(fg_color=BG)

        # Botões + status
        top = ctk.CTkFrame(tab, fg_color="transparent")
        top.pack(fill="x", padx=16, pady=(14, 6))

        self.btn_iniciar = btn_success(top, "▶  Iniciar análise", self._iniciar, width=170)
        self.btn_iniciar.pack(side="left", padx=(0, 10))

        self.btn_parar = btn_danger(top, "⏹  Parar", self._parar, width=120)
        self.btn_parar.pack(side="left", padx=(0, 18))
        self.btn_parar.configure(state="disabled")

        self.lbl_status = ctk.CTkLabel(top, text="", font=F_UI, text_color=TEXTO2)
        self.lbl_status.pack(side="left")

        # Progresso
        prog_frame = ctk.CTkFrame(tab, fg_color="transparent")
        prog_frame.pack(fill="x", padx=16, pady=(0, 8))

        self.progressbar = ctk.CTkProgressBar(prog_frame, height=10, corner_radius=4,
                                               fg_color=BG3, progress_color=AZUL)
        self.progressbar.set(0)
        self.progressbar.pack(fill="x", side="left", expand=True, padx=(0, 10))

        self.lbl_pct = ctk.CTkLabel(prog_frame, text="0%", width=42,
                                     font=F_MONO_S, text_color=TEXTO2)
        self.lbl_pct.pack(side="left")

        # Terminal
        term_header = ctk.CTkFrame(tab, fg_color=BG2, corner_radius=0, height=30)
        term_header.pack(fill="x", padx=16)
        term_header.pack_propagate(False)
        ctk.CTkLabel(term_header, text="  Output",
                     font=F_SEC, text_color=AZUL).pack(side="left", padx=8, fill="y")

        self.terminal = ctk.CTkTextbox(tab, fg_color="#0d1117",
                                        text_color="#c9d1d9",
                                        font=ctk.CTkFont("Courier New", 10),
                                        corner_radius=0, border_width=0,
                                        wrap="none", activate_scrollbars=True)
        self.terminal.pack(fill="both", expand=True, padx=16, pady=(0, 14))
        self.terminal.configure(state="disabled")

    # ── Aba Resultados ─────────────────────────────────────────────────────────
    def _aba_resultados(self, tab):
        tab.configure(fg_color=BG)

        top = ctk.CTkFrame(tab, fg_color="transparent")
        top.pack(fill="x", padx=16, pady=(12, 8))
        btn_ghost(top, "🔄  Atualizar", self._atualizar_resultados, width=130).pack(side="left")

        self.frame_res = ctk.CTkScrollableFrame(tab, fg_color=BG, corner_radius=0,
                                                 scrollbar_button_color=BG3)
        self.frame_res.pack(fill="both", expand=True, padx=16)

        self.lbl_res_info = ctk.CTkLabel(self.frame_res,
                                          text="Rode uma análise para ver os resultados.",
                                          font=F_UI, text_color=TEXTO2)
        self.lbl_res_info.pack(pady=40)

    # ── Aba Histórico ──────────────────────────────────────────────────────────
    def _aba_historico(self, tab):
        tab.configure(fg_color=BG)

        top = ctk.CTkFrame(tab, fg_color="transparent")
        top.pack(fill="x", padx=16, pady=(12, 8))
        btn_ghost(top, "🔄  Atualizar", self._atualizar_historico, width=120).pack(side="left", padx=(0, 8))
        self.btn_comparar = btn_primary(top, "⚖  Comparar selecionados",
                                        self._abrir_comparacao, width=200)
        self.btn_comparar.pack(side="left")
        self.btn_comparar.configure(state="disabled")

        self.frame_hist = ctk.CTkScrollableFrame(tab, fg_color=BG, corner_radius=0,
                                                  scrollbar_button_color=BG3)
        self.frame_hist.pack(fill="both", expand=True, padx=16)
        self._atualizar_historico()

    # ── Callbacks ──────────────────────────────────────────────────────────────
    def _on_tol_change(self, val):
        v = int(round(val))
        self._nm_tol_val = v
        self.lbl_nm_tol.configure(text=f"Tolerância: 10{self._sup(v)}")

    @staticmethod
    def _sup(n):
        sup_map = str.maketrans("-0123456789", "⁻⁰¹²³⁴⁵⁶⁷⁸⁹")
        return str(n).translate(sup_map)

    def _procurar_planilha(self):
        p = filedialog.askopenfilename(
            filetypes=[("Experimento", "*.xlsm *.xlsx *.csv"),
                       ("Excel", "*.xlsm *.xlsx"),
                       ("CSV", "*.csv"),
                       ("Todos", "*.*")])
        if p:
            self.entry_planilha.delete(0, "end")
            self.entry_planilha.insert(0, p)
            self._atualizar_lbl_formato(p)

    def _atualizar_lbl_formato(self, caminho):
        ext = os.path.splitext(caminho)[1].lower()
        eh_csv = (ext == ".csv")
        if eh_csv:
            self.lbl_formato.configure(text="📄 CSV + TOML", text_color=AZUL)
        elif ext in (".xlsm", ".xlsx"):
            self.lbl_formato.configure(text="📊 Excel", text_color=TEXTO2)
        else:
            self.lbl_formato.configure(text="", text_color=TEXTO2)
        for g in ("g1", "g2", "g3"):
            cb = self.cb_gravar_e.get(g)
            if cb:
                if eh_csv:
                    cb.pack(side="left", padx=(8, 0))
                else:
                    cb.pack_forget()
                    self.v_gravar_e[g].set(False)

    def _gravar_energias_toml(self, caminho_csv):
        pasta     = os.path.dirname(os.path.abspath(caminho_csv))
        toml_path = os.path.join(pasta, "parametros.toml")
        mapa_g  = {"g1": "e_g3", "g2": "e_g4", "g3": "e_g5"}
        nomes   = {"g1": "Gaussiana 1", "g2": "Gaussiana 2", "g3": "Gaussiana 3"}
        gravou  = []
        erros   = []
        para_gravar = {}
        for g, chave in mapa_g.items():
            if not self.v_gravar_e.get(g, tk.BooleanVar()).get():
                continue
            val_str = self.e_energias[g].get().strip()
            if not val_str:
                erros.append(f"{nomes[g]}: campo vazio")
                continue
            try:
                para_gravar[chave] = _fval(val_str)
                gravou.append(f"{nomes[g]} = {_fval(val_str)} eV")
            except ValueError:
                erros.append(f"{nomes[g]}: valor inválido '{val_str}'")
        if erros:
            messagebox.showwarning("Gravar energias",
                                   "Não foi possível gravar:\n" + "\n".join(erros))
        if not gravou:
            return
        try:
            try:
                with open(toml_path, "r", encoding="utf-8") as f:
                    linhas = f.readlines()
            except FileNotFoundError:
                linhas = []
            pendentes = set(para_gravar.keys())
            novas = []
            for linha in linhas:
                stripped = linha.strip().lstrip("# ").split("=")[0].strip()
                if stripped in para_gravar:
                    novas.append(f"{stripped} = {para_gravar[stripped]}\n")
                    pendentes.discard(stripped)
                else:
                    novas.append(linha)
            for chave in pendentes:
                novas.append(f"{chave} = {para_gravar[chave]}\n")
            with open(toml_path, "w", encoding="utf-8") as f:
                f.writelines(novas)
        except Exception as e:
            messagebox.showerror("Erro ao gravar",
                                  f"Não foi possível salvar parametros.toml:\n{e}")

    def _procurar_seq(self):
        p = filedialog.askopenfilename(
            filetypes=[("Excel", "*.xlsm *.xlsx"), ("Todos", "*.*")])
        if p:
            self.entry_sq_path.delete(0, "end")
            self.entry_sq_path.insert(0, p)

    def _adicionar_seq(self):
        p = self.entry_sq_path.get().strip()
        if not p:
            messagebox.showwarning("Aviso", "Informe o caminho da planilha.")
            return
        self.sequencia.append({
            "planilha": p,
            "curvas_ativas": {
                "p_mb":  True,
                "p_exp": self.v_sq_exp.get(),
                "p_g3":  self.v_sq_g1.get(),
                "p_g4":  self.v_sq_g2.get(),
                "p_g5":  self.v_sq_g3.get(),
            }
        })
        self.entry_sq_path.delete(0, "end")
        self._atualizar_lista_seq()

    def _atualizar_lista_seq(self):
        for w in self.frame_seq_body.winfo_children():
            w.destroy()
        if not self.sequencia:
            ctk.CTkLabel(self.frame_seq_body, text="Nenhuma planilha na fila.",
                          font=F_SMALL, text_color=TEXTO2).pack(padx=4, pady=4)
            return
        for i, item in enumerate(self.sequencia):
            row = ctk.CTkFrame(self.frame_seq_body, fg_color=BG3, corner_radius=6)
            row.pack(fill="x", pady=3)
            ctk.CTkLabel(row, text=f"  {i+1}.  {os.path.basename(item['planilha'])}",
                          font=F_MONO_S, text_color=TEXTO,
                          anchor="w").pack(side="left", padx=4, pady=6, fill="x", expand=True)
            ctk.CTkButton(row, text="✕", width=32, height=26,
                           fg_color=VERM, hover_color=VERM2,
                           font=ctk.CTkFont("Segoe UI", 11, "bold"), corner_radius=4,
                           command=lambda idx=i: self._remover_seq(idx)).pack(side="right", padx=6, pady=4)

    def _remover_seq(self, i):
        self.sequencia.pop(i)
        self._atualizar_lista_seq()

    def _coletar_cfg(self):
        def f(e):  return _fval(e.get()) if e.get().strip() else None
        def i_v(e): return int(e.get()) if e.get().strip().isdigit() else None

        cpu_val = int(self.slider_cpu.get())
        n_proc  = None if cpu_val == (os.cpu_count() or 1) else cpu_val

        return dict(
            excel_path          = self.entry_planilha.get().strip(),
            resultados_dir      = RESULTADOS_DIR,
            score_minimo        = _fval(self.e_score.get() or "0.99"),
            desvio_alvo         = _fval(self.e_rmse.get() or "0.04"),
            iteracoes_beta      = int(self.e_iter_beta.get() or 1500),
            iteracoes_pesos     = int(self.e_iter_pesos.get() or 1500),
            semente             = i_v(self.e_semente),
            n_processos         = n_proc,
            beta_exp_lo         = _fval(self.e_beta_lo["exp"].get() or "50"),
            beta_exp_hi         = _fval(self.e_beta_hi["exp"].get() or "600"),
            beta_g1_lo          = _fval(self.e_beta_lo["g1"].get() or "100"),
            beta_g1_hi          = _fval(self.e_beta_hi["g1"].get() or "1200"),
            beta_g2_lo          = _fval(self.e_beta_lo["g2"].get() or "100"),
            beta_g2_hi          = _fval(self.e_beta_hi["g2"].get() or "1200"),
            beta_g3_lo          = _fval(self.e_beta_lo["g3"].get() or "100"),
            beta_g3_hi          = _fval(self.e_beta_hi["g3"].get() or "1200"),
            ativa_mb            = self.v_ativa_mb.get(),
            ativa_exp           = self.v_ativa_exp.get(),
            ativa_g1            = self.v_ativa_g1.get(),
            ativa_g2            = self.v_ativa_g2.get(),
            ativa_g3            = self.v_ativa_g3.get(),
            e_g1_valor          = f(self.e_energias["g1"]),
            e_g2_valor          = f(self.e_energias["g2"]),
            e_g3_valor          = f(self.e_energias["g3"]),
            e_g1_nm_livre       = self.v_nm_livre["g1"].get(),
            e_g2_nm_livre       = self.v_nm_livre["g2"].get(),
            e_g3_nm_livre       = self.v_nm_livre["g3"].get(),
            peso_min_mb         = _fval(self.e_peso_mb.get() or "0.01"),
            salvar_graficos_top = int(self.e_salvar_top.get() or 20),
            usar_nm             = self.v_usar_nm.get(),
            nm_top_candidatos   = int(self.e_nm_top.get() or 5),
            nm_max_iter         = int(self.e_nm_iter.get() or 2000),
            nm_tolerancia       = 10 ** int(round(self.slider_nm_tol.get())),
            nm_max_reiniciar    = int(self.e_nm_reinic.get() or 3),
            nm_perturb_escala   = float(self.e_nm_perturb.get() or 0.10),
            nm_usar_clusters    = self.v_nm_clusters.get(),
            nm_cluster_limiar   = float(self.e_nm_limiar.get() or 0.25),
            nm_clusters_max     = int(self.e_nm_max_cl.get() or 10),
            sequencia           = self.sequencia if self.v_usar_seq.get() else [],
        )

    def _iniciar(self):
        if not os.path.isfile(SCRIPT_PATH):
            messagebox.showerror("Erro", f"Script não encontrado:\n{SCRIPT_PATH}")
            return
        planilha = self.entry_planilha.get().strip()
        if not self.v_usar_seq.get() and not planilha:
            messagebox.showwarning("Aviso", "Selecione um experimento antes de rodar.")
            return

        _eh_csv = planilha.lower().endswith(".csv")
        if _eh_csv and not self.v_usar_seq.get():
            self._gravar_energias_toml(planilha)
            import tomllib as _tl
            _pasta = os.path.dirname(os.path.abspath(planilha))
            _toml  = os.path.join(_pasta, "parametros.toml")
            _toml_dados = {}
            try:
                with open(_toml, "rb") as _f:
                    _toml_dados = _tl.load(_f)
            except Exception:
                pass
            _mapa   = {"g1": "e_g3", "g2": "e_g4", "g3": "e_g5"}
            _nomes  = {"g1": "Gaussiana 1", "g2": "Gaussiana 2", "g3": "Gaussiana 3"}
            _ativa  = {"g1": self.v_ativa_g1.get(),
                       "g2": self.v_ativa_g2.get(),
                       "g3": self.v_ativa_g3.get()}
            _faltando = []
            for g, chave in _mapa.items():
                if not _ativa[g]:
                    continue
                if not self.e_energias[g].get().strip() and not _toml_dados.get(chave):
                    _faltando.append(_nomes[g])
            if _faltando:
                messagebox.showwarning("Energias não definidas",
                    f"As seguintes energias não estão definidas:\n{', '.join(_faltando)}\n\n"
                    "Preencha os campos de energia ou grave valores no parametros.toml.")
                return

        try:
            cfg = self._coletar_cfg()
            tmp = criar_script_temporario(SCRIPT_PATH, cfg)
        except Exception as ex:
            messagebox.showerror("Erro", f"Erro ao preparar script:\n{ex}")
            return

        self.rodando = True
        self.btn_iniciar.configure(state="disabled")
        self.btn_parar.configure(state="normal")
        self.lbl_status.configure(text="● Rodando...", text_color=AMAR)
        self.progressbar.set(0)
        self.lbl_pct.configure(text="0%")
        self._log_clear()

        import queue as _queue
        self._log_q = _queue.Queue()
        self._res_q = _queue.Queue()

        def _runner():
            env = os.environ.copy()
            env["PYTHONIOENCODING"] = "utf-8"
            env["PYTHONUTF8"]       = "1"
            env["PYTHONUNBUFFERED"] = "1"
            try:
                if getattr(sys, "frozen", False):
                    cmd = [sys.executable, "--analyze", tmp]
                else:
                    cmd = [sys.executable, "-u", tmp]
                proc = subprocess.Popen(
                    cmd,
                    stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                    text=True, encoding="utf-8", errors="replace",
                    bufsize=1, env=env,
                )
                self.proc_ref = proc
                ultima_pasta = None
                for line in proc.stdout:
                    self._log_q.put(line.rstrip())
                    for token in line.split():
                        token = token.strip().rstrip(")")
                        if os.path.isdir(token) and "Resultados" in token:
                            ultima_pasta = token
                proc.wait()
                self._res_q.put((proc.returncode, ultima_pasta))
            except Exception as ex:
                self._log_q.put(f"ERRO: {ex}")
                self._res_q.put((1, None))
            finally:
                self.proc_ref = None
                try: os.unlink(tmp)
                except: pass

        threading.Thread(target=_runner, daemon=True).start()
        self.after(300, self._poll)

    def _poll(self):
        import queue as _queue
        while True:
            try:
                linha = self._log_q.get_nowait()
                self._log(linha + "\n")
                m = re.search(r'\(\s*([\d.]+)%\)', linha)
                if m:
                    pct = float(m.group(1))
                    self.progressbar.set(pct / 100)
                    self.lbl_pct.configure(text=f"{pct:.1f}%")
                for token in linha.split():
                    token = token.strip().rstrip(")")
                    if os.path.isdir(token) and "Resultados" in token:
                        self.ultima_pasta = token
            except _queue.Empty:
                break

        try:
            rc, pasta = self._res_q.get_nowait()
            if pasta:
                self.ultima_pasta = pasta
            self._finalizar(rc)
            return
        except _queue.Empty:
            pass

        if self.rodando:
            self.after(300, self._poll)

    def _parar(self):
        self.btn_parar.configure(state="disabled")
        self.lbl_status.configure(text="⏳ Encerrando...", text_color=AMAR)

        pasta = self.ultima_pasta
        if not pasta:
            candidatas = sorted([p for p in glob.glob(
                os.path.join(RESULTADOS_DIR, "*")) if os.path.isdir(p)], reverse=True)
            if candidatas:
                pasta = candidatas[0]
        if pasta and os.path.isdir(pasta):
            open(os.path.join(pasta, "PARAR.txt"), "w").close()

        def _aguardar():
            import time
            for _ in range(20):
                if self.proc_ref is None or self.proc_ref.poll() is not None:
                    return
                time.sleep(0.5)
            if self.proc_ref:
                try: self.proc_ref.terminate()
                except: pass

        threading.Thread(target=_aguardar, daemon=True).start()

    def _finalizar(self, rc):
        self.rodando = False
        self.btn_iniciar.configure(state="normal")
        self.btn_parar.configure(state="disabled")
        if rc == 0:
            self.progressbar.set(1.0)
            self.lbl_pct.configure(text="100%")
            self.lbl_status.configure(text="✓ Concluído", text_color=VERDE2)
            self._atualizar_resultados()
        elif rc == -1:
            self.lbl_status.configure(text="⏹ Interrompido", text_color=AMAR)
        else:
            self.lbl_status.configure(text=f"✗ Erro (código {rc})", text_color=VERM2)

    def _log(self, texto):
        self.terminal.configure(state="normal")
        self.terminal.insert("end", texto)
        self.terminal.see("end")
        self.terminal.configure(state="disabled")

    def _log_clear(self):
        self.terminal.configure(state="normal")
        self.terminal.delete("1.0", "end")
        self.terminal.configure(state="disabled")

    def _atualizar_resultados(self):
        for w in self.frame_res.winfo_children():
            w.destroy()

        pasta = self.ultima_pasta
        if not pasta:
            candidatas = sorted([p for p in glob.glob(
                os.path.join(RESULTADOS_DIR, "*")) if os.path.isdir(p)], reverse=True)
            if candidatas:
                pasta = candidatas[0]

        if not pasta or not os.path.isdir(pasta):
            ctk.CTkLabel(self.frame_res, text="Nenhum resultado encontrado.",
                          font=F_UI, text_color=TEXTO2).pack(pady=40)
            return

        ctk.CTkLabel(self.frame_res, text=os.path.basename(pasta),
                      font=ctk.CTkFont("Courier New", 13, "bold"),
                      text_color=AZUL, anchor="w").pack(anchor="w", pady=(8, 4))

        txts = glob.glob(os.path.join(pasta, "*.txt"))
        if txts:
            with open(txts[0], encoding="utf-8", errors="replace") as f:
                conteudo = f.read()
            r2   = re.search(r'R²\s*=\s*([\d.]+)', conteudo)
            rmse = re.search(r'RMSE\s*=\s*([\d.]+)', conteudo)
            mf = ctk.CTkFrame(self.frame_res, fg_color="transparent")
            mf.pack(anchor="w", pady=4)
            if r2:
                ctk.CTkLabel(mf, text=f"R² = {r2.group(1)}",
                              font=ctk.CTkFont("Courier New", 13, "bold"),
                              text_color=VERDE2).pack(side="left", padx=(0, 20))
            if rmse:
                ctk.CTkLabel(mf, text=f"RMSE = {rmse.group(1)}",
                              font=ctk.CTkFont("Courier New", 13),
                              text_color=TEXTO).pack(side="left")

        bf = ctk.CTkFrame(self.frame_res, fg_color="transparent")
        bf.pack(anchor="w", pady=6)
        btn_ghost(bf, "📁  Abrir pasta",
                  lambda: os.startfile(pasta), width=130).pack(side="left", padx=(0, 8))
        xlsx = glob.glob(os.path.join(pasta, "*.xlsx"))
        if xlsx:
            btn_ghost(bf, "📊  Abrir Excel",
                      lambda: os.startfile(xlsx[0]), width=130).pack(side="left")

    def _atualizar_historico(self):
        for w in self.frame_hist.winfo_children():
            w.destroy()
        self._sel_historico.clear()

        if not os.path.isdir(RESULTADOS_DIR):
            ctk.CTkLabel(self.frame_hist, text="Nenhum resultado ainda.",
                          font=F_UI, text_color=TEXTO2).pack(pady=20)
            return

        pastas = sorted([p for p in glob.glob(os.path.join(RESULTADOS_DIR, "*"))
                          if os.path.isdir(p)], reverse=True)
        if not pastas:
            ctk.CTkLabel(self.frame_hist, text="Nenhuma análise anterior encontrada.",
                          font=F_UI, text_color=TEXTO2).pack(pady=20)
            return

        def _on_sel_change(*_):
            n = sum(1 for item in self._sel_historico.values() if item["var"].get())
            self.btn_comparar.configure(state="normal" if n >= 2 else "disabled")

        for pasta in pastas:
            xlsx = glob.glob(os.path.join(pasta, "*.xlsx"))

            outer = ctk.CTkFrame(self.frame_hist, fg_color=BG2, corner_radius=6,
                                  border_width=1, border_color=BORDA)
            outer.pack(fill="x", pady=3)

            # ── Linha principal ──
            hdr = ctk.CTkFrame(outer, fg_color="transparent")
            hdr.pack(fill="x")

            v = tk.BooleanVar(value=False)
            v.trace_add("write", _on_sel_change)
            key = ("best", pasta)
            if xlsx:
                self._sel_historico[key] = {
                    "var":    v,
                    "getter": lambda p=pasta: self._ler_resultado_xlsx(
                        glob.glob(os.path.join(p, "*.xlsx"))[0], os.path.basename(p)),
                }
            ctk.CTkCheckBox(hdr, text="", variable=v, width=28,
                             checkbox_width=16, checkbox_height=16,
                             checkmark_color="white", fg_color=AZUL, hover_color=AZUL2,
                             state="normal" if xlsx else "disabled"
                             ).pack(side="left", padx=(8, 0), pady=8)

            ctk.CTkLabel(hdr, text=os.path.basename(pasta),
                          font=F_MONO_S, text_color=AZUL,
                          anchor="w").pack(side="left", padx=6, pady=8, fill="x", expand=True)

            bf = ctk.CTkFrame(hdr, fg_color="transparent")
            bf.pack(side="right", padx=8, pady=4)

            # ── Botão expandir runs ──
            runs_frame = ctk.CTkFrame(outer, fg_color="transparent")
            if xlsx:
                btn_r = btn_ghost(bf, "▶ runs", None, width=80)
                btn_r.pack(side="left", padx=(0, 6))

                def _toggle(p=pasta, xp=xlsx[0], rf=runs_frame, b=[btn_r]):
                    if rf.winfo_ismapped():
                        rf.pack_forget()
                        b[0].configure(text="▶ runs")
                    else:
                        self._carregar_runs_frame(p, xp, rf, _on_sel_change)
                        rf.pack(fill="x", padx=(40, 8), pady=(0, 6))
                        b[0].configure(text="▼ runs")

                btn_r.configure(command=_toggle)

            btn_ghost(bf, "📁 Abrir",
                      lambda p=pasta: os.startfile(p), width=90).pack(side="left", padx=(0, 6))
            if xlsx:
                btn_ghost(bf, "📊 Excel",
                          lambda x=xlsx[0]: os.startfile(x), width=80).pack(side="left")


    # ── Comparação ─────────────────────────────────────────────────────────────

    def _abrir_comparacao(self):
        selecionados = [(k, item) for k, item in self._sel_historico.items()
                        if item["var"].get()]
        if len(selecionados) < 2:
            return

        dados  = []
        falhas = []
        for key, item in selecionados:
            d = item["getter"]()
            if d:
                dados.append(d)
            else:
                falhas.append(str(key[1] if len(key) > 1 else key))

        if falhas:
            messagebox.showwarning("Comparação",
                "Não foi possível ler resultado de:\n" + "\n".join(falhas))
        if len(dados) < 2:
            return

        self._janela_comparacao(dados)

    def _carregar_runs_frame(self, pasta, xlsx_path, frame, on_sel_change):
        """Popula o sub-frame com as rodadas válidas do experimento."""
        for w in frame.winfo_children():
            w.destroy()

        runs = self._ler_runs_xlsx(xlsx_path, os.path.basename(pasta))
        if not runs:
            ctk.CTkLabel(frame, text="Nenhuma rodada válida encontrada.",
                          font=F_SMALL, text_color=TEXTO2).pack(anchor="w", padx=4, pady=4)
            return

        for run in runs:
            row = ctk.CTkFrame(frame, fg_color=BG3, corner_radius=4)
            row.pack(fill="x", pady=1)

            v = tk.BooleanVar(value=False)
            v.trace_add("write", on_sel_change)
            key = ("run", pasta, run.get("rodada", ""))
            self._sel_historico[key] = {"var": v, "getter": lambda r=run, p=pasta: {**r, "pasta": p}}

            ctk.CTkCheckBox(row, text="", variable=v, width=24,
                             checkbox_width=14, checkbox_height=14,
                             checkmark_color="white", fg_color=ROXO,
                             hover_color="#6e40c9").pack(side="left", padx=(6, 0), pady=3)

            def _fmt(val, fmt):
                return format(val, fmt) if val else "—"

            r2_str   = f"R²={run.get('r2', 0):.4f}"
            rmse_str = f"RMSE={run.get('rmse', 0):.4f}"
            b_exp    = f"βEXP={_fmt(run.get('beta_exp'), '.0f')}"
            b_g1     = f"βG1={_fmt(run.get('beta_g1'),  '.0f')}"
            b_g2     = f"βG2={_fmt(run.get('beta_g2'),  '.0f')}"
            b_g3     = f"βG3={_fmt(run.get('beta_g3'),  '.0f')}"
            ctk.CTkLabel(row,
                          text=f"  {run.get('rodada', '?')}    {r2_str}    {rmse_str}"
                               f"    {b_exp}    {b_g1}    {b_g2}    {b_g3}",
                          font=F_MONO_S, text_color=TEXTO,
                          anchor="w").pack(side="left", padx=4, pady=3, fill="x", expand=True)

    @staticmethod
    def _ler_runs_xlsx(path, pasta_nome, max_runs=15):
        """Lê as rodadas válidas da tabela completa do melhor_resultado.xlsx."""
        try:
            from openpyxl import load_workbook
            wb  = load_workbook(path, data_only=True, read_only=True)
            ws  = wb.active

            col_map = {
                "Rodada":          "rodada",
                "R²":              "r2",
                "RMSE":            "rmse",
                "Peso MB":         "p_mb",
                "Peso Exp":        "p_exp",
                "Peso G1":         "p_g1",
                "Peso G2":         "p_g2",
                "Peso G3":         "p_g3",
                "Beta Exp":        "beta_exp",
                "Beta G1":         "beta_g1",
                "Beta G2":         "beta_g2",
                "Beta G3":         "beta_g3",
                "Energia G1 (eV)": "e_g1",
                "Energia G2 (eV)": "e_g2",
                "Energia G3 (eV)": "e_g3",
            }

            hmap       = {}
            idx_status = None
            runs       = []
            header_found = False

            # Passagem única — encontra header e lê dados no mesmo loop
            for row in ws.iter_rows(values_only=True):
                if not row:
                    continue
                if not header_found:
                    # O cabeçalho real da tabela tem "Status" na coluna seguinte;
                    # a linha par("Rodada", valor) no resumo tem só col A preenchida.
                    if row[0] == "Rodada" and len(row) > 1 and row[1] == "Status":
                        header_found = True
                        for j, cell in enumerate(row):
                            if cell is not None:
                                hmap[str(cell)] = j
                        idx_status = hmap.get("Status")
                    continue

                if row[0] is None:
                    continue
                if idx_status is not None and row[idx_status] != "VÁLIDO":
                    continue

                d = {"nome": f"{pasta_nome} / {row[hmap['Rodada']]}"}
                for col_name, key in col_map.items():
                    idx = hmap.get(col_name)
                    if idx is not None and idx < len(row) and row[idx] is not None:
                        try:
                            if key == "rodada":
                            raw = row[idx]
                            try:
                                d[key] = int(float(raw))
                            except (ValueError, TypeError):
                                d[key] = str(raw)  # NM1, NM2, etc.
                        else:
                            d[key] = float(row[idx])
                        except (ValueError, TypeError):
                            d[key] = row[idx]
                runs.append(d)
                if len(runs) >= max_runs:
                    break

            wb.close()
            return runs
        except Exception:
            return []

    @staticmethod
    def _ler_resultado_xlsx(path, nome):
        """Lê parâmetros do melhor resultado a partir do melhor_resultado.xlsx."""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, data_only=True, read_only=True)
            ws = wb.active
            mapa = {
                "R²":                      "r2",
                "RMSE":                    "rmse",
                "Peso Maxwell-Boltzmann":  "p_mb",
                "Peso Exponencial":        "p_exp",
                "Peso Gaussiana 1":        "p_g1",
                "Peso Gaussiana 2":        "p_g2",
                "Peso Gaussiana 3":        "p_g3",
                "Beta Exponencial":        "beta_exp",
                "Beta Gaussiana 1":        "beta_g1",
                "Beta Gaussiana 2":        "beta_g2",
                "Beta Gaussiana 3":        "beta_g3",
                "Energia Gaussiana 1 (eV)":"e_g1",
                "Energia Gaussiana 2 (eV)":"e_g2",
                "Energia Gaussiana 3 (eV)":"e_g3",
            }
            d = {"nome": nome, "pasta": os.path.dirname(path), "rodada": None}
            for row in ws.iter_rows(values_only=True):
                if row and row[0] in mapa:
                    d[mapa[row[0]]] = float(row[1]) if row[1] is not None else 0.0
            wb.close()
            return d if len(d) > 3 else None
        except Exception:
            return None

    @staticmethod
    def _encontrar_png(d):
        """Localiza o PNG correspondente a um resultado."""
        pasta  = d.get("pasta", "")
        rodada = d.get("rodada")
        if not pasta or not os.path.isdir(pasta):
            return None
        if rodada is None:
            # Melhor resultado global
            pngs = glob.glob(os.path.join(pasta, "*_melhor_global.png"))
            return pngs[0] if pngs else None
        else:
            rod_str = str(int(float(rodada)))
            rod_fmt = rod_str.zfill(4)
            pngs = glob.glob(os.path.join(pasta, "graficos_rodadas", f"rod{rod_fmt}_*.png"))
            return pngs[0] if pngs else None

    def _janela_comparacao(self, dados):
        import numpy as np
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

        win = ctk.CTkToplevel(self)
        win.title("Comparação de resultados")
        win.geometry("1050x900")
        win.resizable(True, True)
        win.configure(fg_color=BG)
        win.grab_set()

        # ── Botão exportar (bottom, packed before expand widget) ──────────────
        bf = ctk.CTkFrame(win, fg_color="transparent")
        bf.pack(side="bottom", pady=(0, 14))
        btn_ghost(bf, "💾  Exportar Excel", lambda: self._exportar_comparacao(dados), width=160).pack(side="left", padx=(0, 10))

        # ── Container scrollável principal ────────────────────────────────────
        scroll_main = ctk.CTkScrollableFrame(win, fg_color=BG, corner_radius=0,
                                             scrollbar_button_color=BG3)
        scroll_main.pack(fill="both", expand=True, padx=0, pady=0)

        ctk.CTkLabel(scroll_main, text="Comparação de resultados", font=F_H2,
                     text_color=AZUL).pack(anchor="w", padx=20, pady=(14, 4))

        # ── Tabela de parâmetros ──────────────────────────────────────────────
        tbl_frame = ctk.CTkScrollableFrame(scroll_main, fg_color=BG2, corner_radius=8,
                                            height=210, scrollbar_button_color=BG3)
        tbl_frame.pack(fill="x", padx=20, pady=(0, 10))

        colunas = ["Parâmetro"] + [d["nome"] for d in dados]
        linhas = [
            ("R²",              "r2"),
            ("RMSE",            "rmse"),
            ("p_MB (%)",        "p_mb",  True),
            ("p_EXP (%)",       "p_exp", True),
            ("p_G1 (%)",        "p_g1",  True),
            ("p_G2 (%)",        "p_g2",  True),
            ("p_G3 (%)",        "p_g3",  True),
            ("β_EXP",           "beta_exp"),
            ("β_G1",            "beta_g1"),
            ("β_G2",            "beta_g2"),
            ("β_G3",            "beta_g3"),
            ("E_G1 (eV)",       "e_g1"),
            ("E_G2 (eV)",       "e_g2"),
            ("E_G3 (eV)",       "e_g3"),
        ]

        # Cabeçalho
        for j, col in enumerate(colunas):
            ctk.CTkLabel(tbl_frame, text=col, font=ctk.CTkFont("Segoe UI", 10, "bold"),
                         text_color=AZUL, width=max(160, 120),
                         anchor="w" if j == 0 else "center").grid(
                row=0, column=j, padx=(4, 8), pady=(6, 2), sticky="w")

        for i, spec in enumerate(linhas, start=1):
            nome_lbl = spec[0]
            chave    = spec[1]
            pct      = len(spec) > 2 and spec[2]
            bg_row   = BG3 if i % 2 == 0 else BG2

            ctk.CTkLabel(tbl_frame, text=nome_lbl, font=F_SMALL,
                         text_color=TEXTO2, width=160, anchor="w",
                         fg_color=bg_row).grid(row=i, column=0, padx=(4, 8), pady=1, sticky="w")

            for j, d in enumerate(dados, start=1):
                val = d.get(chave, 0) or 0
                if pct:
                    txt = f"{val*100:.1f}%"
                elif chave in ("r2", "rmse"):
                    txt = f"{val:.6f}"
                elif chave.startswith("beta"):
                    txt = f"{val:.1f}"
                elif chave.startswith("e_g"):
                    txt = f"{val:.3f}" if val else "—"
                else:
                    txt = f"{val}"

                ctk.CTkLabel(tbl_frame, text=txt, font=F_MONO_S,
                             text_color=TEXTO, width=120, anchor="center",
                             fg_color=bg_row).grid(row=i, column=j, padx=(0, 8), pady=1)

        # ── Gráfico de barras empilhadas ──────────────────────────────────────
        nomes_exp = [d["nome"] for d in dados]
        chaves_p  = ["p_mb", "p_exp", "p_g1", "p_g2", "p_g3"]
        labels_p  = ["MB", "EXP", "G1", "G2", "G3"]
        cores_p   = ["#9b59b6", "#2980b9", "#27ae60", "#e67e22", "#e74c3c"]

        fig, ax = plt.subplots(figsize=(9, max(1.8, len(dados) * 0.7 + 0.6)))
        fig.patch.set_facecolor(BG2)
        ax.set_facecolor(BG2)
        ax.tick_params(colors=TEXTO2, labelsize=8)
        for spine in ax.spines.values():
            spine.set_edgecolor(BORDA)

        lefts = np.zeros(len(dados))
        for chave, label, cor in zip(chaves_p, labels_p, cores_p):
            vals = np.array([d.get(chave, 0) or 0 for d in dados])
            bars = ax.barh(nomes_exp, vals * 100, left=lefts * 100,
                           color=cor, label=label, height=0.55)
            for bar, v in zip(bars, vals):
                if v > 0.04:
                    ax.text(bar.get_x() + bar.get_width() / 2,
                            bar.get_y() + bar.get_height() / 2,
                            f"{v*100:.1f}%", ha="center", va="center",
                            fontsize=8, color="white", fontweight="bold")
            lefts += vals

        ax.set_xlabel("Contribuição (%)", color=TEXTO2, fontsize=9)
        ax.set_xlim(0, 100)
        ax.xaxis.label.set_color(TEXTO2)
        ax.tick_params(axis="y", labelcolor=TEXTO)
        lgd = ax.legend(loc="lower right", fontsize=8,
                        facecolor=BG3, edgecolor=BORDA, labelcolor=TEXTO)
        fig.tight_layout(pad=0.8)

        fig_h_px = int(fig.get_figheight() * fig.dpi) + 20
        canvas = FigureCanvasTkAgg(fig, master=scroll_main)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="x", padx=20, pady=(0, 8))
        canvas.get_tk_widget().configure(height=fig_h_px)

        # ── Gráficos ──────────────────────────────────────────────────────────
        pngs_map = [(d, self._encontrar_png(d)) for d in dados]

        ctk.CTkLabel(scroll_main, text="Gráficos", font=F_H2,
                     text_color=AZUL).pack(anchor="w", padx=20, pady=(8, 4))

        png_scroll = ctk.CTkScrollableFrame(scroll_main, fg_color=BG, corner_radius=0,
                                            height=400, scrollbar_button_color=BG3)
        png_scroll.pack(fill="x", padx=20, pady=(0, 8))

        try:
            from PIL import Image as _PImg
            IMG_W = 460
            cols  = 2
            grid_i = 0
            for d, png_path in pngs_map:
                col_idx = grid_i % cols
                row_idx = grid_i // cols
                card = ctk.CTkFrame(png_scroll, fg_color=BG2, corner_radius=6)
                card.grid(row=row_idx, column=col_idx, padx=6, pady=6, sticky="nsew")
                png_scroll.grid_columnconfigure(col_idx, weight=1)

                ctk.CTkLabel(card, text=d["nome"], font=F_SMALL,
                             text_color=AZUL, anchor="w").pack(anchor="w", padx=8, pady=(6, 2))

                if png_path and os.path.isfile(png_path):
                    img     = _PImg.open(png_path)
                    w, h    = img.size
                    new_h   = int(h * IMG_W / w)
                    ctk_img = ctk.CTkImage(img, size=(IMG_W, new_h))
                    ctk.CTkLabel(card, image=ctk_img, text="",
                                 fg_color="transparent").pack(padx=6, pady=(0, 6))
                else:
                    rod = d.get("rodada")
                    label = f"Gráfico {rod} não encontrado" if rod is not None else "Gráfico não encontrado"
                    ctk.CTkLabel(card, text=label, font=F_SMALL,
                                 text_color=TEXTO2).pack(padx=8, pady=(4, 8))

                grid_i += 1
        except Exception as ex:
            ctk.CTkLabel(png_scroll, text=f"Erro ao carregar imagens: {ex}",
                         font=F_SMALL, text_color=VERM2).pack(padx=8, pady=8)

        btn_ghost(bf, "Fechar", win.destroy, width=100).pack(side="left")

    def _exportar_comparacao(self, dados):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="comparacao_resultados.xlsx",
            title="Salvar comparação")
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Comparação"

        cab_font = Font(bold=True, color="FFFFFF")
        cab_fill = PatternFill("solid", fgColor="2C3E50")

        linhas = [
            ("R²",              "r2",       False),
            ("RMSE",            "rmse",     False),
            ("p_MB",            "p_mb",     True),
            ("p_EXP",           "p_exp",    True),
            ("p_G1",            "p_g1",     True),
            ("p_G2",            "p_g2",     True),
            ("p_G3",            "p_g3",     True),
            ("β_EXP",           "beta_exp", False),
            ("β_G1",            "beta_g1",  False),
            ("β_G2",            "beta_g2",  False),
            ("β_G3",            "beta_g3",  False),
            ("E_G1 (eV)",       "e_g1",     False),
            ("E_G2 (eV)",       "e_g2",     False),
            ("E_G3 (eV)",       "e_g3",     False),
        ]

        # Cabeçalho
        ws.cell(1, 1, "Parâmetro").font = cab_font
        ws.cell(1, 1).fill = cab_fill
        for j, d in enumerate(dados, start=2):
            c = ws.cell(1, j, d["nome"])
            c.font = cab_font
            c.fill = cab_fill
            c.alignment = Alignment(horizontal="center")

        for i, (nome_lbl, chave, pct) in enumerate(linhas, start=2):
            ws.cell(i, 1, nome_lbl)
            for j, d in enumerate(dados, start=2):
                val = d.get(chave, 0) or 0
                ws.cell(i, j, round(val * 100, 2) if pct else val)

        wb.save(path)
        messagebox.showinfo("Exportar", f"Comparação salva em:\n{path}")


if __name__ == "__main__":
    app = NTApp()
    app.mainloop()
