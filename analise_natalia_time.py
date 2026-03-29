"""
TOF Spectroscopy Analysis
=========================
Replica e otimiza os cálculos da aba "N+" do Excel (.xlsm) em Python.

Resultados auditáveis:
  - melhor resultado de CADA rodada (beta) salvo em CSV
  - gráfico individual (log + resíduo absoluto) por rodada válida
  - viewer interativo para navegar e comparar

ESTRUTURA:
  Seção  1 — Importações e utilitários
  Seção  2 — Configuração (caminhos, critérios, intervalos)
  Seção  3 — Constantes físicas e geométricas
  Seção  4 — Leitura do Excel
  Seção  5 — Kernels de Gauss-Hermite
  Seção  6 — Filtros de aceitação do detector
  Seção  7 — Normalização das amplitudes
  Seção  8 — Integração do sinal teórico
  Seção  9 — Métricas (R² e RMSE)
  Seção 10 — Busca aleatória global
  Seção 11 — Gráficos e viewer interativo
  Seção 12 — Execução principal
"""

# =============================================================================
# SEÇÃO 1 — IMPORTAÇÕES E UTILITÁRIOS
# =============================================================================

import os
import datetime
import multiprocessing as mp
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec


def br(valor, decimais=6):
    """Formata número no padrão brasileiro (vírgula decimal)."""
    if isinstance(valor, (float, np.floating)):
        return f"{valor:.{decimais}f}".replace(".", ",")
    return str(valor)



# Nomes completos das curvas (usados em relatórios e gráficos)
NOME_CURVA = {
    "p_mb":  "MB",
    "p_exp": "Exponencial",
    "p_g3":  "Gaussiana 1",
    "p_g4":  "Gaussiana 2",
    "p_g5":  "Gaussiana 3",
}
NOME_CURVA_CURTO = {
    "p_mb":  "MB",
    "p_exp": "EXP",
    "p_g3":  "G1",
    "p_g4":  "G2",
    "p_g5":  "G3",
}

# =============================================================================
# SEÇÃO 2 — CONFIGURAÇÃO
# =============================================================================

# Caminho da planilha Excel
EXCEL_PATH = (
    r"C:\Users\theki\OneDrive\Área de Trabalho\Analises"
    r"\Natalia_time_O2 - 100eV - vMFinal teste 2.xlsm"
)

# Pasta raiz onde serão criadas subpastas por execução
RESULTADOS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Resultados")

# ── Configuração da planilha ──────────────────────────────────────────────────
# A aba e a linha de início são detectadas automaticamente:
#   .xlsm → usa sempre a primeira aba; linha de início = linha abaixo de "tempo(ns)"
#   .csv  → usa dados.csv + parametros.toml na mesma pasta

# ── Critérios de aceitação ────────────────────────────────────────────────────
SCORE_MINIMO = 0.99    # R² mínimo para resultado "válido"
DESVIO_ALVO  = 0.04    # RMSE máximo para "alta precisão"

# ── Configuração da busca ─────────────────────────────────────────────────────
ITERACOES_BETA  = 10   # nº de combinações de beta
ITERACOES_PESOS = 10    # nº de pesos por beta  →  100 × 50 = 5000 cálculos
SEMENTE         = None  # None = aleatório; inteiro = reproduzível

# ── Paralelismo ───────────────────────────────────────────────────────
# None = usa todos os núcleos disponíveis
# int  = limita ao número especificado (ex: 8 para deixar núcleos livres)
N_PROCESSOS = None

# ── Parada suave ──────────────────────────────────────────────────────
# Crie um arquivo com esse nome na pasta de resultados para encerrar a
# busca ao fim da rodada atual, salvando tudo que já foi coletado.
# (Ctrl+C também funciona e faz o mesmo)
ARQUIVO_PARAR = "PARAR.txt"

# ── Intervalos de busca para os betas variáveis ───────────────────────────────
INTERVALOS_BETA = {
    "beta_exp": (50,   5000),
    "beta_g3":  (50, 5000),
    "beta_g4":  (100, 5000),
    "beta_g5":  (100, 5000),
}

# ── Quais curvas participam do sorteio de pesos ───────────────────────────────
CURVAS_PESO_ATIVAS = {
    "p_mb":  True,
    "p_exp": True,
    "p_g3":  True,
    "p_g4":  True,
    "p_g5":  False,   # Gaussiana 3 desativada (peso=0)
}

# ── Energias das gaussianas ───────────────────────────────────────────────────
# ENERGIAS_VALOR:
#   None  = usa o valor lido da planilha (padrão para usuários)
#   float = sobrescreve o valor da planilha (útil sem abrir o Excel)
ENERGIAS_VALOR = {
    "e_g3": None,   # eV — None = planilha
    "e_g4": None,   # eV — None = planilha
    "e_g5": None,   # eV — None = planilha
}


# ENERGIA_NM_LIVRE:
#   False = energia fixada no valor acima (ou da planilha) — NM não toca
#   True  = NM pode refinar a energia a partir do valor de partida
ENERGIA_NM_LIVRE = {
    "e_g3": False,
    "e_g4": False,
    "e_g5": False,
}

INTERVALOS_ENERGIA = {
    "e_g3": (0.1, 6.0),   # eV — limites para o NM quando ENERGIA_NM_LIVRE = True
    "e_g4": (0.1, 6.0),
    "e_g5": (0.1, 6.0),
}

# ── Limites mínimos de peso ───────────────────────────────────────────
# None = sem restrição  |  float = peso mínimo após normalização
# Aplicado tanto na busca aleatória quanto no NM.
PESO_MIN = {
    "p_mb":  0.01,
    "p_exp": None,
    "p_g3":  None,
    "p_g4":  None,
    "p_g5":  None,
}

# Limites de saída — evita gerar centenas de arquivos em buscas grandes
# None = sem limite (gera tudo)
SALVAR_GRAFICOS_TOP = 20   # salva PNG dos N melhores por RMSE (além do melhor global)
VIEWER_TOP          = 50   # mostra apenas os N melhores no viewer interativo

# ── Refinamento Nelder-Mead (roda após a busca aleatória) ────────────────────
USAR_NELDER_MEAD   = True   # True = ativa refinamento local
NM_TOP_CANDIDATOS  = 5      # quantos melhores candidatos refinar
NM_MAX_ITER        = 2000   # iterações máximas por candidato (por tentativa)
NM_TOLERANCIA      = 1e-9   # critério de convergência (xatol e fatol)
NM_MAX_REINICIAR   = 3      # máx de reinícios por candidato (1 = sem reinício)
NM_PERTURB_ESCALA  = 0.10   # perturbação no reinício: ±10% dos intervalos

# ── Seleção de candidatos por clusters ───────────────────────────────
# Agrupa os resultados válidos no espaço dos betas (normalizado pelos
# intervalos de busca) e escolhe o melhor de cada cluster como candidato
# para o NM — garantindo diversidade em vez de pegar só os top N por RMSE.
#
# NM_CLUSTER_LIMIAR : distância normalizada para separar clusters (0–1)
#                     valores menores → mais clusters (mais diversidade)
#                     valores maiores → menos clusters (mais parecido com top-N)
# NM_CLUSTERS_MAX   : limite máximo de clusters (= candidatos para o NM)
NM_USAR_CLUSTERS   = True
NM_CLUSTER_LIMIAR  = 0.25
NM_CLUSTERS_MAX    = 10

# ── Busca em duas fases ───────────────────────────────────────────────────────
# Fase 1: varredura grossa sobre o espaço completo
# Fase 2: busca densa concentrada nos melhores intervalos encontrados na fase 1
# Se False: usa busca simples com ITERACOES_BETA × ITERACOES_PESOS normal
BUSCA_DUAS_FASES    = False
FASE1_ITER_BETA     = 200    # betas na fase grossa
FASE1_ITER_PESOS    = 30     # pesos por beta na fase grossa
FASE2_ITER_BETA     = 300    # betas na fase fina
FASE2_ITER_PESOS    = 200    # pesos por beta na fase fina
FASE2_TOP_REGIOES   = 10     # quantos melhores da fase 1 definem os novos intervalos
FASE2_MARGEM        = 0.20   # ±20% ao redor dos melhores para definir intervalos finos

# ── Adaptação automática de intervalos ───────────────────────────────────────
# Após a busca, estreita INTERVALOS_BETA com base nos válidos encontrados
# e roda uma segunda busca automaticamente.
# Se False: roda apenas uma busca com os intervalos definidos acima.
ADAPTAR_INTERVALOS  = False
ADAPT_MARGEM        = 0.15   # ±15% ao redor dos limites dos válidos

# ── Sequência de planilhas ────────────────────────────────────────────────────
# Deixe vazio [] para processar apenas EXCEL_PATH (comportamento padrão).
# Preencha para processar várias planilhas em sequência automática.
# Cada entrada pode sobrescrever CURVAS_PESO_ATIVAS para aquela planilha.
# Exemplo:
# SEQUENCIA = [
#     {
#         "planilha": r"C:\...\experimento_100eV.xlsm",
#         "curvas_ativas": {"p_mb": True, "p_exp": True, "p_g3": True, "p_g4": True, "p_g5": False},
#     },
#     {
#         "planilha": r"C:\...\experimento_150eV.xlsm",
#         "curvas_ativas": {"p_mb": True, "p_exp": True, "p_g3": True, "p_g4": False, "p_g5": False},
#     },
# ]
SEQUENCIA = []

# ── Validação manual ──────────────────────────────────────────────────────────
# MODO_VALIDACAO = True  → calcula o espectro com VALIDACAO_PARAMS e exibe
# VALIDACAO_REFINAR_NM = True → após calcular, roda o NM a partir desses params
#   como ponto de partida (útil para explorar regiões específicas do espaço)
MODO_VALIDACAO     = False
VALIDACAO_REFINAR_NM = False   # só tem efeito quando MODO_VALIDACAO = True

VALIDACAO_PARAMS = {
    # Larguras (beta) — beta_mb é fixo e lido da planilha
    "beta_exp": 529.63,
    "beta_g3":  382.08,
    "beta_g4":  241.07,
    "beta_g5":  913.98,
    # Energias (eV) — usadas quando ENERGIA_NM_LIVRE = True como ponto de partida
    "e_g3": 0.70,
    "e_g4": 2.59,
    "e_g5": 2.00,
    # Pesos — devem somar 1.0 (são normalizados automaticamente se não somarem)
    "p_mb":  0.030252,
    "p_exp": 0.240645,
    "p_g3":  0.482782,
    "p_g4":  0.246321,
    "p_g5":  0.000000,
}

# =============================================================================
# SEÇÃO 3 — CONSTANTES FÍSICAS E GEOMÉTRICAS
# =============================================================================

D    = 6.8
DD   = 0.8
LL   = 6.8
DE   = 350.0
FMB  = 1.7

FATOR_ENERGIA    = 7837.5
FATOR_VELOCIDADE = 16.8

H6    = 120000
N_VEL = 1492

# Nós e pesos de Gauss-Hermite (4 pontos, fixos)
X_GH = np.array([0.381186990207322, 1.15719371244678,
                  1.98165675669584,  2.93063742025724])
W_GH = np.array([0.7645441286517,   0.7928900483864,
                  0.8667526065634,   1.071930144248])

H6_INV = 1.0 / H6
N0     = 1.0 / (np.pi * D * (DD / 2) ** 2)
H9_VEC = np.arange(1, N_VEL + 1) * H6_INV

L9_VEC = None   # preenchido por inicializar_grade()
M9_VEC = None

# Grade de tempos — gerada dinamicamente em inicializar_grade()
# para garantir que cobre todo o range experimental.
# Os segmentos replicam a resolução da planilha Excel original:
#   510–2960:  passo 50 ns
#   3060–9960: passo 100 ns
#   10160+:    passo 200 ns
GRADE_BM = None   # preenchido por inicializar_grade()

I7_NORM = 500.0   # tempo fixo para normalização das amplitudes

# =============================================================================
# SEÇÃO 4 — LEITURA DO EXCEL
# =============================================================================

CELULAS_FIXAS = {
    "beta_mb": (3, 26),
    "e_g3":    (5, 28),
    "e_g4":    (5, 29),
    "e_g5":    (5, 30),
    "m_frag":  (13, 1),
    "m_mol":   (12, 1),
    # Parâmetros geométricos/físicos do espectrômetro
    "DE":      (4, 1),    # energia de deflexão (eV) — célula B5
    "D":       (8, 1),    # distância de voo (m)    — célula B9
    "LL":      (10, 1),   # comprimento do tubo (m) — célula B11
    "I7_NORM": (6, 8),   # tempo de normalização   — célula I7
}

FIXOS_FALLBACK = {
    "beta_mb": 2572.6309556813,
    "e_g3":    0.7,
    "e_g4":    3.0,
    "e_g5":    2.0,
    "m_frag":  16.0,
    "m_mol":   32.0,
    # Fallback = valores originais da planilha de referência (O2 300eV)
    "DE":      350.0,
    "D":       6.8,
    "LL":      6.8,
    "I7_NORM": 500.0,
}


def _detectar_linha_dados(ws) -> int:
    """
    Procura 'tempo' (case-insensitive) na coluna A e retorna a linha
    seguinte como início dos dados. Lança ValueError se não encontrar.
    NOTA: não usar em modo read_only — consome o iterador. Usar apenas
    para verificação; a leitura real deve ser feita em passagem única.
    """
    for i, row in enumerate(ws.iter_rows(min_col=1, max_col=1, values_only=True), start=1):
        val = row[0]
        if val is not None and isinstance(val, str) and "tempo" in val.lower():
            return i + 1  # linha seguinte = início dos dados
    raise ValueError(
        "Cabeçalho 'tempo' não encontrado na coluna A da planilha. "
        "Verifique se a planilha está no formato esperado."
    )


def ler_parametros_fixos(caminho: str) -> dict:
    """
    Lê parâmetros fixos do .xlsm.
    Usa sempre a primeira aba da planilha.
    """
    from openpyxl import load_workbook
    wb = load_workbook(caminho, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    params = {}
    for nome, (r, c) in CELULAS_FIXAS.items():
        cel = ws.cell(row=r + 1, column=c + 1)
        try:
            val = cel.value
            if val is None or (isinstance(val, str) and not val.strip()):
                params[nome] = None   # célula vazia — usará valor padrão
            else:
                params[nome] = float(val)
        except Exception as e:
            raise ValueError(f"Erro lendo '{nome}' ({cel.coordinate}): {e}")
    wb.close()
    return params


def ler_dados_experimentais(caminho: str) -> tuple:
    """
    Lê dados experimentais do .xlsm em passagem única (necessário para read_only).
    Detecta o início dos dados pela célula 'tempo' na coluna A,
    depois coleta pares (tempo, sinal) até a primeira linha inválida.
    """
    from openpyxl import load_workbook
    wb = load_workbook(caminho, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    tempos, sinal = [], []
    coletando = False
    header_encontrado = False
    for i, row in enumerate(ws.iter_rows(min_col=1, max_col=2, values_only=True), start=1):
        a, b = row[0], row[1]
        if not coletando:
            # Procura cabeçalho 'tempo' na coluna A
            if a is not None and isinstance(a, str) and "tempo" in a.lower():
                coletando = True
                header_encontrado = True
            continue
        # Já passamos do cabeçalho — coletar dados
        if a is None or b is None:
            break
        try:
            tempos.append(float(a))
            sinal.append(float(b))
        except (TypeError, ValueError):
            break
    wb.close()
    if not header_encontrado:
        raise ValueError(
            "Cabeçalho 'tempo' não encontrado na coluna A da planilha. "
            "Verifique se a planilha está no formato esperado."
        )
    return np.array(tempos), np.array(sinal)


def ler_parametros_toml(caminho_csv: str) -> dict:
    """
    Lê parametros.toml na mesma pasta do .csv.
    Chaves obrigatórias: beta_mb, m_frag, m_mol
    Chaves opcionais:    e_g3, e_g4, e_g5  (podem ser gravadas pelo NT)
    Retorna dict com as chaves lidas; energias ausentes ficam como None.
    """
    import tomllib
    pasta = os.path.dirname(os.path.abspath(caminho_csv))
    toml_path = os.path.join(pasta, "parametros.toml")
    if not os.path.isfile(toml_path):
        raise FileNotFoundError(
            f"Arquivo 'parametros.toml' não encontrado em:\n{pasta}\n\n"
            "Ele deve estar na mesma pasta que o dados.csv."
        )
    with open(toml_path, "rb") as f:
        dados = tomllib.load(f)
    # Chaves obrigatórias
    obrigatorias = ["beta_mb", "m_frag", "m_mol"]
    params = {}
    for chave in obrigatorias:
        if chave not in dados:
            raise ValueError(
                f"Chave obrigatória '{chave}' ausente no parametros.toml. "
                f"Chaves obrigatórias: {obrigatorias}"
            )
        params[chave] = float(dados[chave])
    # Chaves opcionais — energias
    for chave in ("e_g3", "e_g4", "e_g5"):
        params[chave] = float(dados[chave]) if chave in dados else None
    # Chaves opcionais — geometria do espectrômetro
    for chave in ("DE", "D", "LL", "I7_NORM"):
        params[chave] = float(dados[chave]) if chave in dados else None
    return params


def ler_dados_csv(caminho_csv: str) -> tuple:
    """
    Lê dados.csv com colunas: tempo, sinal
    (primeira linha = cabeçalho, ignorada automaticamente se não numérica)
    """
    tempos, sinal = [], []
    with open(caminho_csv, "r", encoding="utf-8") as f:
        for linha in f:
            linha = linha.strip()
            if not linha:
                continue
            partes = linha.replace(";", ",").split(",")
            if len(partes) < 2:
                continue
            try:
                t = float(partes[0])
                s = float(partes[1])
                tempos.append(t)
                sinal.append(s)
            except ValueError:
                continue   # pula cabeçalho ou linhas inválidas
    if len(tempos) < 5:
        raise ValueError(
            f"Poucos pontos lidos do dados.csv ({len(tempos)}). "
            "Verifique se o arquivo tem colunas 'tempo' e 'sinal' separadas por vírgula."
        )
    return np.array(tempos), np.array(sinal)


def inicializar_grade(fixos: dict, t_min_exp: float = 510.0, t_max_exp: float = 19960.0,
                      verbose: bool = True):
    global L9_VEC, M9_VEC, P9_GRADE, U9_GRADE, P9_NORM, U9_NORM, GRADE_BM, N0, D, DD, LL, DE, I7_NORM
    L9_VEC = FATOR_ENERGIA * fixos["m_frag"] * H9_VEC ** 2
    M9_VEC = FATOR_ENERGIA * fixos["m_mol"]  * H9_VEC ** 2
    # Atualizar globais geométricas/físicas com valores da planilha (ou fallback)
    D       = fixos.get("D",       6.8)
    DD      = fixos.get("DD",      0.8)
    LL      = fixos.get("LL",      6.8)
    DE      = fixos.get("DE",    350.0)
    I7_NORM = fixos.get("I7_NORM", 500.0)
    N0 = 1.0 / (np.pi * D * (DD / 2) ** 2)

    # Construir GRADE_BM garantindo cobertura do range experimental
    t_inicio = max(10.0,    min(t_min_exp, 510.0)  - 50.0)
    t_fim    = max(19960.0, t_max_exp + 200.0)   # margem de 200 ns após o último ponto

    seg1 = np.arange(max(t_inicio, 50), min(2961,   t_fim + 1),   50)
    seg2 = np.arange(3060,              min(9961,   t_fim + 1),   100)
    seg3 = np.arange(10160,             min(t_fim + 201, 99999),  200)
    GRADE_BM = np.unique(np.concatenate([seg1, seg2, seg3])).astype(float)

    # Pré-calcular filtros para todos os pontos da grade e para I7_NORM
    _p = []; _u = []
    for t in GRADE_BM:
        p9, u9 = calcular_filtros(t)
        _p.append(p9); _u.append(u9)
    P9_GRADE = np.array(_p)
    U9_GRADE = np.array(_u)

    P9_NORM, U9_NORM = calcular_filtros(I7_NORM)
    if verbose:
        print(f"  Grade: {len(GRADE_BM)} pontos  ({GRADE_BM[0]:.0f}–{GRADE_BM[-1]:.0f} ns)")


def energia_para_v0(e_ev, m_frag: float) -> float:
    if e_ev is None or e_ev <= 0:
        return 0.0
    return np.sqrt(e_ev / m_frag) / 72.284


# =============================================================================
# SEÇÃO 5 — KERNELS DE GAUSS-HERMITE
# =============================================================================

def gh_exp(beta2: float) -> np.ndarray:
    x, w = X_GH[:, None], W_GH[:, None]
    return (w * np.exp(-x**2) * x
            / np.sqrt(x**4 + x**2 * beta2 * H9_VEC**2)).sum(axis=0)


def gh_gauss(beta2: float, v0: float) -> np.ndarray:
    x, w = X_GH[:, None], W_GH[:, None]
    num = w * np.exp(-x**2) * np.exp(-(x**2 * beta2 * (H9_VEC**2 - v0**2)**2))
    den = np.sqrt(x**2 + x * beta2 * H9_VEC**2)
    return (num / den).sum(axis=0)


# =============================================================================
# SEÇÃO 6 — FILTROS DE ACEITAÇÃO DO DETECTOR
# =============================================================================

def calcular_filtros(t: float, _D=None, _DD=None, _LL=None, _DE=None) -> tuple:
    """Retorna (p9, u9) para o tempo t (ns).
    Usa as globais D, DD, LL, DE, N0 por padrão (já atualizadas por inicializar_grade).
    """
    _D  = _D  if _D  is not None else D
    _DD = _DD if _DD is not None else DD
    _LL = _LL if _LL is not None else LL
    _DE = _DE if _DE is not None else DE

    k9 = _D / (2 * H9_VEC * t)
    j9 = 1.0 / (1.0 / k9 + _DD / _D)
    ks = np.clip(k9, -1 + 1e-10, 1 - 1e-10)
    js = np.clip(j9, -1 + 1e-10, 1 - 1e-10)

    o9 = (2 * N0 * _D * _LL *
          ((_D/2) * (  np.arcsin(js) / (2*js**2) - np.arcsin(ks) / (2*ks**2)
                     + np.sqrt(1 - js**2) / (2*js) - np.sqrt(1 - ks**2) / (2*ks))
           - H9_VEC * t * (  np.arcsin(js) / js - np.arcsin(ks) / ks
                             + np.log((1 + np.sqrt(1 - js**2)) / js)
                             - np.log((1 + np.sqrt(1 - ks**2)) / ks))))

    geo    = np.pi * N0 * _LL * (_DD / 2)**2
    filtro = np.where(k9 < 1, o9, geo)

    q9 = _D / (2 * (H9_VEC * t + np.sqrt(L9_VEC / _DE) * FATOR_VELOCIDADE))
    s9 = _D / (2 * (H9_VEC * t + np.sqrt(M9_VEC / (2 * _DE)) * FATOR_VELOCIDADE))

    cf = np.where(q9 > 1, 1.0, (2/np.pi) * np.arcsin(np.clip(q9, -1+1e-10, 1-1e-10)))
    cm = np.where(s9 > 1, 1.0, (2/np.pi) * np.arcsin(np.clip(s9, -1+1e-10, 1-1e-10)))

    return filtro * cf, filtro * cm


# =============================================================================
# SEÇÃO 7 — NORMALIZAÇÃO DAS AMPLITUDES
# =============================================================================

def normalizar_amplitudes(fixos, beta_exp, beta_g3, beta_g4, beta_g5) -> dict:
    """
    Calcula amplitudes replicando exatamente o VBA NormalizarProcessoCientifico.

    O VBA faz (por curva, isoladamente):
      Nv=1, peso_curva=1, outros=0 → calcula → dv = AK7 → Nv = 1/dv

    Com Nv=1 e peso=1, a distribuição bruta é (ex. EXP):
      AH9 = H9 * exp(-beta²*H9²) * GH_kernel   (sem AK3, sem p_exp/soma)
      AI9 = P9 * AH9
      AK9 = AI2 * H9 * AH9 * dv
          = m_frag * H9 * AH9 * (1/H6)          (dv = I3 = 1/H6)

    Portanto o peso de integração é: (m_frag / H6) * H9
    E: int_e = SUM(p9 * y_e_pura * (m_frag/H6) * H9)

    O fator (m_frag/H6) é constante e cancela em BS=sinal/u0 para curvas
    que usam o mesmo filtro (EXP, G3, G4). Mas a MB usa filtro u9 diferente
    de p9, então o fator relativo entre MB e os demais importa e deve ser
    calculado com o mesmo peso.

    Curvas inativas (CURVAS_PESO_ATIVAS=False) não são calculadas.
    """
    p9_fix, u9_fix = P9_NORM, U9_NORM
    mf  = fixos["m_frag"]
    mm  = fixos["m_mol"]

    # Peso de integração: dv = 1/H6
    # MB  usa AA9 = X5*H9*X9*dv  onde X5 = m_mol = 32
    # EXP/G3/G4 usam AK9 = AI2*H9*AH9*dv  onde AI2 = m_frag = 16
    peso_int_mf = mf * H6_INV * H9_VEC   # EXP, G3, G4
    peso_int_mm = mm * H6_INV * H9_VEC   # MB

    # Normalização SEM filtro detector:
    # AA9 = X5*H9*X9*dv   (X9 = dist MB pura, sem U9)
    # AK9 = AI2*H9*AH9*dv (AH9 = dist EXP pura, sem P9)
    # O filtro aparece apenas no sinal (AS9=P9*AQ9*H9/2)

    # MB
    y_mb_pura = H9_VEC * np.exp(-((fixos["beta_mb"] * H9_VEC)**2))
    int_mb    = np.sum(y_mb_pura * peso_int_mm)
    x1        = 1.0 / int_mb if int_mb > 0 else 1.0

    # EXP
    y_e = None
    if CURVAS_PESO_ATIVAS.get("p_exp", True):
        b2e   = beta_exp**2
        y_e   = H9_VEC * np.exp(-(b2e * H9_VEC**2)) * gh_exp(b2e)
        int_e = np.sum(y_e * peso_int_mf)
        ak3   = 1.0 / int_e if int_e > 0 else 1.0
    else:
        ak3 = 1.0

    # G3
    y_g3 = None
    if CURVAS_PESO_ATIVAS.get("p_g3", True):
        v0_g3  = energia_para_v0(fixos["e_g3"], mf)
        b2g3   = beta_g3**2
        y_g3   = (1/2)*H9_VEC*np.exp(-(b2g3**2*(H9_VEC**2-v0_g3**2)**2))*gh_gauss(b2g3, v0_g3)
        int_g3 = np.sum(y_g3 * peso_int_mf)
        au3    = 1.0 / int_g3 if int_g3 > 0 else 1.0
    else:
        au3 = 1.0

    # G4
    y_g4 = None
    if CURVAS_PESO_ATIVAS.get("p_g4", True):
        v0_g4  = energia_para_v0(fixos["e_g4"], mf)
        b2g4   = beta_g4**2
        y_g4   = (1/2)*H9_VEC*np.exp(-(b2g4**2*(H9_VEC**2-v0_g4**2)**2))*gh_gauss(b2g4, v0_g4)
        int_g4 = np.sum(y_g4 * peso_int_mf)
        bd3    = 1.0 / int_g4 if int_g4 > 0 else 1.0
    else:
        bd3 = 1.0

    # G5
    y_g5 = None
    if CURVAS_PESO_ATIVAS.get("p_g5", False):
        v0_g5  = energia_para_v0(fixos["e_g5"], mf)
        b2g5   = beta_g5**2
        y_g5   = (1/2)*H9_VEC*np.exp(-(b2g5**2*(H9_VEC**2-v0_g5**2)**2))*gh_gauss(b2g5, v0_g5)
        int_g5 = np.sum(y_g5 * peso_int_mf)
        bm3    = 1.0 / int_g5 if int_g5 > 0 else 1.0
    else:
        bm3 = 1.0

    # y_mb sempre calculado (usado também em calcular_espectro)
    y_mb = H9_VEC * np.exp(-((fixos["beta_mb"] * H9_VEC)**2))

    return {
        "x1":  x1,
        "ak3": ak3,
        "au3": au3,
        "bd3": bd3,
        "bm3": bm3,
        # vetores y pré-calculados — reutilizados em calcular_espectro
        "y_mb": y_mb,
        "y_e":  y_e   if CURVAS_PESO_ATIVAS.get("p_exp", True) else None,
        "y_g3": y_g3  if CURVAS_PESO_ATIVAS.get("p_g3",  True) else None,
        "y_g4": y_g4  if CURVAS_PESO_ATIVAS.get("p_g4",  True) else None,
        "y_g5": y_g5  if CURVAS_PESO_ATIVAS.get("p_g5", False) else None,
    }


# =============================================================================
# SEÇÃO 8 — INTEGRAÇÃO DO SINAL TEÓRICO
# =============================================================================

def _sinal_um_tempo(t, beta_exp, beta_g3, beta_g4, beta_g5,
                    v0_g3, v0_g4, v0_g5,
                    p_mb, p_exp, p_g3, p_g4, p_g5,
                    amps, fixos) -> float:
    """
    Sinal bruto para um único tempo t.
    Os pesos p_* chegam aqui já normalizados pela soma (feito em calcular_espectro).
    Curvas inativas têm p_*=0, portanto não contribuem — e não são calculadas.
    """
    p9, u9  = calcular_filtros(t)
    beta_mb = fixos["beta_mb"]

    # MB — sempre presente
    y_mb = H9_VEC * np.exp(-((beta_mb * H9_VEC)**2))
    ac   = np.sum(u9 * p_mb * amps["x1"] * y_mb * H9_VEC / 2)

    # EXP
    am = 0.0
    if p_exp > 0:
        b2e = beta_exp**2
        y_e = H9_VEC * np.exp(-(b2e * H9_VEC**2)) * gh_exp(b2e)
        am  = np.sum(p9 * p_exp * amps["ak3"] * y_e * H9_VEC / 2)

    # G3
    av = 0.0
    if p_g3 > 0:
        b2g3 = beta_g3**2
        y_g3 = (1/2)*H9_VEC*np.exp(-(b2g3**2*(H9_VEC**2-v0_g3**2)**2))*gh_gauss(b2g3, v0_g3)
        av   = np.sum(p9 * p_g3 * amps["au3"] * y_g3 * H9_VEC / 2)

    # G4
    be = 0.0
    if p_g4 > 0:
        b2g4 = beta_g4**2
        y_g4 = (1/2)*H9_VEC*np.exp(-(b2g4**2*(H9_VEC**2-v0_g4**2)**2))*gh_gauss(b2g4, v0_g4)
        be   = np.sum(p9 * p_g4 * amps["bd3"] * y_g4 * H9_VEC / 2)

    # G5
    bn = 0.0
    if p_g5 > 0:
        b2g5 = beta_g5**2
        y_g5 = (1/2)*H9_VEC*np.exp(-(b2g5**2*(H9_VEC**2-v0_g5**2)**2))*gh_gauss(b2g5, v0_g5)
        bn   = np.sum(p9 * p_g5 * amps["bm3"] * y_g5 * H9_VEC / 2)

    return ac * FMB + am + av + be + bn


def calcular_espectro(tempos_exp, beta_exp, beta_g3, beta_g4, beta_g5,
                      p_mb, p_exp, p_g3, p_g4, p_g5, fixos, amps=None):
    """
    Calcula BS(t) interpolado nos tempos experimentais.
    Usa grade BM densa + interpolação linear (np.interp), igual ao Excel.

    Estrutura confirmada pela planilha:
      BS = ((AC*fmb) + AM + AV + BE + BN/2) / u0

    Onde:
      AC = Z7 = CONSTANTE — integral MB calculada uma vez com u9(I7_NORM).
           Hardcoded em todas as linhas da grade no Excel.
      AM/AV/BE = integrais EXP/G3/G4 com p9(t) variável — Data Table.
      Peso de integração do sinal: H9/2  (colunas Z, AJ, AS, BB)
      Peso de integração da normalização: m_frag/H6*H9  (colunas AK, AT, BC)
      São DIFERENTES — ambos corretos conforme planilha.

    Pesos normalizados por soma(p) antes do cálculo.
    Vetores y calculados uma vez fora do loop de tempo.
    """
    if amps is None:
        amps = normalizar_amplitudes(fixos, beta_exp, beta_g3, beta_g4, beta_g5)

    # Normalizar pesos (replica Excel: p / SUM(AA5:AE5))
    soma_p = p_mb + p_exp + p_g3 + p_g4 + p_g5
    if soma_p <= 0:
        return np.zeros(len(tempos_exp))
    pn_mb  = p_mb  / soma_p
    pn_exp = p_exp / soma_p
    pn_g3  = p_g3  / soma_p
    pn_g4  = p_g4  / soma_p
    pn_g5  = p_g5  / soma_p

    mf    = fixos["m_frag"]
    v0_g3 = energia_para_v0(fixos["e_g3"], mf)
    v0_g4 = energia_para_v0(fixos["e_g4"], mf)
    v0_g5 = energia_para_v0(fixos["e_g5"], mf)

    # ── Vetores y pré-calculados (independem de t) ────────────────────────────
    # Reutiliza os vetores de amps se disponíveis (calculados em normalizar_amplitudes)
    # evitando recalcular gh_exp/gh_gauss que são as operações mais custosas.
    y_mb = amps.get("y_mb") if amps.get("y_mb") is not None else            H9_VEC * np.exp(-((fixos["beta_mb"] * H9_VEC)**2))
    y_e  = (amps.get("y_e") if amps.get("y_e") is not None else
            (H9_VEC * np.exp(-(beta_exp**2 * H9_VEC**2)) * gh_exp(beta_exp**2)
             if pn_exp > 0 else None))
    y_g3 = (amps.get("y_g3") if amps.get("y_g3") is not None else
            ((1/2)*H9_VEC*np.exp(-(beta_g3**4*(H9_VEC**2-v0_g3**2)**2))
             * gh_gauss(beta_g3**2, v0_g3) if pn_g3 > 0 else None))
    y_g4 = (amps.get("y_g4") if amps.get("y_g4") is not None else
            ((1/2)*H9_VEC*np.exp(-(beta_g4**4*(H9_VEC**2-v0_g4**2)**2))
             * gh_gauss(beta_g4**2, v0_g4) if pn_g4 > 0 else None))
    y_g5 = (amps.get("y_g5") if amps.get("y_g5") is not None else
            ((1/2)*H9_VEC*np.exp(-(beta_g5**4*(H9_VEC**2-v0_g5**2)**2))
             * gh_gauss(beta_g5**2, v0_g5) if pn_g5 > 0 else None))

    # ── Integrar vetorialmente usando filtros pré-calculados ─────────────────
    # P9_GRADE shape (175, 1492), U9_GRADE shape (175, 1492)
    # y_* shape (1492,) — einsum substitui o loop Python de 175 iterações
    dv = H9_VEC / 2   # shape (1492,)
    y_mb_dv = y_mb * dv

    # AC: usa U9_GRADE (filtro MB)  — shape (175,)
    ac_vec = pn_mb * amps["x1"] * np.einsum('ij,j->i', U9_GRADE, y_mb_dv)

    sinal_grade = ac_vec * FMB

    if pn_exp > 0:
        sinal_grade += pn_exp * amps["ak3"] * np.einsum('ij,j->i', P9_GRADE, y_e  * dv)
    if pn_g3 > 0:
        sinal_grade += pn_g3  * amps["au3"] * np.einsum('ij,j->i', P9_GRADE, y_g3 * dv)
    if pn_g4 > 0:
        sinal_grade += pn_g4  * amps["bd3"] * np.einsum('ij,j->i', P9_GRADE, y_g4 * dv)
    if pn_g5 > 0:
        sinal_grade += pn_g5  * amps["bm3"] * np.einsum('ij,j->i', P9_GRADE, y_g5 * dv)

    u0 = sinal_grade[0]
    if u0 == 0:
        return np.zeros(len(tempos_exp))

    bs = sinal_grade / u0
    return np.interp(tempos_exp, GRADE_BM, bs)


# =============================================================================
# SEÇÃO 9 — MÉTRICAS
# =============================================================================

def calcular_af4(sinal_exp: np.ndarray, sinal_teo: np.ndarray) -> float:
    """
    R² = 1 - SUMSQ(teo-exp) / SUMSQ(exp-mean)
    Replica exata da fórmula Excel: =1-SOMAQUAD(D)/SOMAQUAD(F)
    onde D=teo-exp, F=exp-mean(exp). SEM raiz quadrada.
    """
    ym     = np.mean(sinal_exp)
    ss_res = np.sum((sinal_teo - sinal_exp)**2)
    ss_tot = np.sum((sinal_exp - ym)**2)
    if ss_tot == 0:
        return 0.0
    return float(1.0 - ss_res / ss_tot)


def verificar_normalizacao(historico: list, fixos: dict):
    """
    Diagnóstico de normalização por rodada.

    Calcula a contribuição real de cada curva para o sinal em t=510 (u0),
    equivalente às células AE7+AN7+AW7+BF7+BO7 do Excel.

    Cada contribuição = pn_i * Nv_i * SUM(filtro_i * y_i * H9/2)
    A soma deve ser igual a u0 (sinal total no primeiro ponto da grade).
    Reporta cada componente em valor absoluto e como fração de u0.
    """
    p9_norm, u9_norm = calcular_filtros(I7_NORM)
    mf = fixos["m_frag"]

    print(f"\n{'─'*105}")
    print(f"CONTRIBUIÇÕES DO SINAL em t=I7_NORM=500 (AE7+AN7+AW7+BF7+BO7)")
    print(f"{'─'*105}")
    print(f"{'Rod':>5}  {'AC(MB)':>10}  {'AM(EXP)':>10}  {'AV(G3)':>10}  "
          f"{'BE(G4)':>10}  {'BN(G5)':>10}  {'u0':>12}  {'SOMA/u0':>8}")
    print(f"{'─'*105}")

    for reg in historico:
        be  = reg["beta_exp"]
        bg3 = reg["beta_g3"]
        bg4 = reg["beta_g4"]
        bg5 = reg["beta_g5"]

        amps = normalizar_amplitudes(fixos, be, bg3, bg4, bg5)

        soma_p = (reg["p_mb"] + reg["p_exp"] + reg["p_g3"]
                  + reg["p_g4"] + reg["p_g5"])
        if soma_p <= 0:
            continue
        pn_mb  = reg["p_mb"]  / soma_p
        pn_exp = reg["p_exp"] / soma_p
        pn_g3  = reg["p_g3"]  / soma_p
        pn_g4  = reg["p_g4"]  / soma_p
        pn_g5  = reg["p_g5"]  / soma_p

        dv = H9_VEC / 2   # peso do sinal

        # AC — MB usa u9
        y_mb = H9_VEC * np.exp(-((fixos["beta_mb"] * H9_VEC)**2))
        ac   = pn_mb * amps["x1"] * np.sum(u9_norm * y_mb * dv)

        # AM — EXP usa p9
        b2e  = be**2
        y_e  = H9_VEC * np.exp(-(b2e * H9_VEC**2)) * gh_exp(b2e)
        am   = pn_exp * amps["ak3"] * np.sum(p9_norm * y_e * dv)

        # AV — G3 usa p9
        v0_g3 = energia_para_v0(fixos["e_g3"], mf)
        b2g3  = bg3**2
        y_g3  = (1/2)*H9_VEC*np.exp(-(b2g3**2*(H9_VEC**2-v0_g3**2)**2))*gh_gauss(b2g3, v0_g3)
        av    = pn_g3 * amps["au3"] * np.sum(p9_norm * y_g3 * dv)

        # BE — G4 usa p9
        v0_g4 = energia_para_v0(fixos["e_g4"], mf)
        b2g4  = bg4**2
        y_g4  = (1/2)*H9_VEC*np.exp(-(b2g4**2*(H9_VEC**2-v0_g4**2)**2))*gh_gauss(b2g4, v0_g4)
        be_   = pn_g4 * amps["bd3"] * np.sum(p9_norm * y_g4 * dv)

        # BN — G5 usa p9
        v0_g5 = energia_para_v0(fixos["e_g5"], mf)
        b2g5  = bg5**2
        y_g5  = (1/2)*H9_VEC*np.exp(-(b2g5**2*(H9_VEC**2-v0_g5**2)**2))*gh_gauss(b2g5, v0_g5)
        bn    = pn_g5 * amps["bm3"] * np.sum(p9_norm * y_g5 * dv)

        u0   = ac * FMB + am + av + be_ + bn
        soma = (ac * FMB + am + av + be_ + bn) / u0 if u0 > 0 else 0.0

        print(f"{reg['rodada']:>5}  {ac*FMB:>10.4f}  {am:>10.4f}  "
              f"{av:>10.4f}  {be_:>10.4f}  {bn:>10.4f}  "
              f"{u0:>12.4f}  {soma:>8.6f}")

    print(f"{'─'*105}\n")


def calcular_rmse(tempos_exp, sinal_exp, sinal_teo) -> tuple:
    """
    Replica exata do VBA CalcularDesvioResiduos.

    Passo 1 — ajusta reta sobre resíduos ABSOLUTOS D = teo - exp:
        m = SLOPE(D, t),  b = INTERCEPT(D, t)

    Passo 2 — mede RMSE dos resíduos RELATIVOS E = (teo-exp)/exp
              em relação a essa reta:
        RMSE = sqrt( mean( (E_i - (m*t_i + b))² ) )

    Retorna (rmse, m, b) para uso no gráfico e no relatório.
    """
    mask = np.abs(sinal_exp) > 1e-12
    if not np.any(mask):
        return 99999.0, 0.0, 0.0

    t = tempos_exp[mask]
    D = (sinal_teo[mask] - sinal_exp[mask])           # resíduo absoluto
    E = D / sinal_exp[mask]                            # resíduo relativo

    # Regressão linear de D sobre t  (replica SLOPE/INTERCEPT do Excel)
    t_mean = np.mean(t)
    D_mean = np.mean(D)
    denom  = np.sum((t - t_mean)**2)
    if denom == 0:
        m, b = 0.0, D_mean
    else:
        m = np.sum((t - t_mean) * (D - D_mean)) / denom
        b = D_mean - m * t_mean

    # RMSE de E em relação à reta
    y_tend = m * t + b
    rmse   = float(np.sqrt(np.mean((E - y_tend)**2)))
    return rmse, float(m), float(b)


# =============================================================================
# SEÇÃO 10 — BUSCA ALEATÓRIA GLOBAL
# =============================================================================

# ── Worker: estado compartilhado entre rodadas do mesmo processo ──────

def _worker_init(fixos_arg, tempos_arg, sinal_arg):
    """
    Inicializador de cada processo worker.
    Reconstrói P9_GRADE/U9_GRADE uma única vez por worker.
    """
    global _w_fixos, _w_tempos, _w_sinal
    _w_fixos  = fixos_arg
    _w_tempos = tempos_arg
    _w_sinal  = sinal_arg
    # Reinicializa a grade neste processo (globals P9_GRADE etc.)
    inicializar_grade(fixos_arg,
                      t_min_exp=float(tempos_arg.min()),
                      t_max_exp=float(tempos_arg.max()),
                      verbose=False)


def _processar_rodada(args):
    """
    Processa UMA rodada completa (1 beta × ITERACOES_PESOS pesos).
    Retorna o dicionário do melhor resultado desta rodada.
    Chamada pelo Pool — usa globals do worker inicializado por _worker_init.
    """
    i_b, semente_rodada = args
    rng           = np.random.default_rng(semente_rodada)
    curvas_ativas = [k for k, v in CURVAS_PESO_ATIVAS.items() if v]
    n_curvas      = len(curvas_ativas)
    nomes_pesos   = ["p_mb", "p_exp", "p_g3", "p_g4", "p_g5"]
    fixos         = _w_fixos
    tempos_exp    = _w_tempos
    sinal_exp     = _w_sinal

    # Sorteia betas
    betas = {nome: rng.uniform(lo, hi)
             for nome, (lo, hi) in INTERVALOS_BETA.items()}

    # Energias: ENERGIAS_VALOR sobrescreve planilha se não for None
    energias_rodada = {}
    for k in ("e_g3", "e_g4", "e_g5"):
        val = ENERGIAS_VALOR.get(k)
        if val is not None:
            energias_rodada[k] = val        # sobrescreve planilha
        else:
            energias_rodada[k] = fixos[k]   # sempre salva o valor usado (da planilha)

    fixos_rodada = {**fixos, **energias_rodada}
    amps_beta    = normalizar_amplitudes(
        fixos_rodada,
        betas["beta_exp"], betas["beta_g3"],
        betas["beta_g4"],  betas["beta_g5"]
    )

    melhor_local = {"af4": -9999, "rmse": 99999,
                    "m_reta": 0.0, "b_reta": 0.0,
                    "valido": False, "pesos": None}
    validos = alta_precisao = 0

    for _ in range(ITERACOES_PESOS):
        dir_s = rng.dirichlet(np.ones(n_curvas))
        pesos = {k: 0.0 for k in nomes_pesos}
        for idx, nome in enumerate(curvas_ativas):
            pesos[nome] = float(dir_s[idx])
        for k, ativo in CURVAS_PESO_ATIVAS.items():
            if not ativo:
                pesos[k] = 0.0
        # Aplicar piso de peso e renormalizar
        ajustado = False
        for k, pmin in PESO_MIN.items():
            if pmin is not None and pesos.get(k, 0) < pmin:
                pesos[k] = pmin
                ajustado = True
        if ajustado:
            soma = sum(pesos[k] for k in curvas_ativas)
            if soma > 0:
                for k in curvas_ativas:
                    pesos[k] /= soma

        sinal_teo        = calcular_espectro(
            tempos_exp,
            betas["beta_exp"], betas["beta_g3"],
            betas["beta_g4"],  betas["beta_g5"],
            pesos["p_mb"], pesos["p_exp"],
            pesos["p_g3"], pesos["p_g4"], pesos["p_g5"],
            fixos_rodada, amps_beta
        )
        af4    = calcular_af4(sinal_exp, sinal_teo)
        valido = af4 >= SCORE_MINIMO

        ml = melhor_local
        if valido:
            # RMSE só calculado quando R² passa — evita custo desnecessário
            rmse, m_r, b_r = calcular_rmse(tempos_exp, sinal_exp, sinal_teo)
            validos += 1
            if rmse <= DESVIO_ALVO:
                alta_precisao += 1
            if not ml["valido"] or rmse < ml["rmse"] or \
                    (rmse == ml["rmse"] and af4 > ml["af4"]):
                melhor_local = {"af4": af4, "rmse": rmse,
                                "m_reta": m_r, "b_reta": b_r,
                                "valido": True, "pesos": dict(pesos)}
        else:
            if not ml["valido"] and af4 > ml["af4"]:
                melhor_local = {"af4": af4, "rmse": 99999,
                                "m_reta": 0.0, "b_reta": 0.0,
                                "valido": False, "pesos": dict(pesos)}

    return {
        "rodada":   i_b + 1,
        **betas,
        **energias_rodada,
        **melhor_local["pesos"],
        "af4":      melhor_local["af4"],
        "rmse":     melhor_local["rmse"],
        "m_reta":   melhor_local["m_reta"],
        "b_reta":   melhor_local["b_reta"],
        "status":   "VÁLIDO" if melhor_local["valido"] else "Abaixo do Corte",
        # contadores para o consolidador
        "_validos":        validos,
        "_alta_precisao":  alta_precisao,
        "_total":          ITERACOES_PESOS,
    }


def busca_aleatoria(tempos_exp: np.ndarray,
                    sinal_exp:  np.ndarray,
                    fixos:      dict,
                    pasta_parar: str = ".") -> dict:
    """
    Para cada rodada de beta (ITERACOES_BETA):
      - Sorteia beta_exp, beta_g3, beta_g4, beta_g5
      - Testa ITERACOES_PESOS combinações de pesos (Dirichlet)
      - Guarda o MELHOR resultado desta rodada no histórico

    Com N_PROCESSOS > 1 (ou None), divide as rodadas entre workers via
    multiprocessing.Pool — cada rodada é independente, ganho quase linear
    com o número de núcleos disponíveis.
    """
    n_proc = N_PROCESSOS if N_PROCESSOS is not None else mp.cpu_count()
    n_proc = max(1, min(n_proc, ITERACOES_BETA))

    # Gera uma semente independente por rodada (reprodutível se SEMENTE definida)
    rng_master = np.random.default_rng(SEMENTE)
    sementes   = rng_master.integers(0, 2**31, size=ITERACOES_BETA).tolist()
    args_list  = list(enumerate(sementes))

    t0 = datetime.datetime.now()
    total_calc = ITERACOES_BETA * ITERACOES_PESOS
    print(f"\nIniciando busca: {ITERACOES_BETA} × {ITERACOES_PESOS} = "
          f"{total_calc:,} combinações  |  {n_proc} processo(s)")
    print(f"Critério: R² ≥ {SCORE_MINIMO}  |  RMSE ≤ {DESVIO_ALVO}\n")

    interrompido  = False
    arquivo_parar = os.path.join(pasta_parar, ARQUIVO_PARAR)

    if n_proc == 1:
        # ── Modo single-process ───────────────────────────────────────────
        historico     = []
        melhor_global = {"af4": -9999, "rmse": 99999, "valido": False}
        validos = alta_precisao = 0
        _worker_init(fixos, tempos_exp, sinal_exp)

        try:
            for i_b, semente in enumerate(sementes):
                if os.path.exists(arquivo_parar):
                    print(f"\n  [{ARQUIVO_PARAR} detectado] Encerrando após {i_b} rodadas...")
                    interrompido = True
                    break
                reg = _processar_rodada((i_b, semente))
                historico.append({k: v for k, v in reg.items()
                                  if not k.startswith("_")})
                validos       += reg["_validos"]
                alta_precisao += reg["_alta_precisao"]
                _atualizar_melhor_global(melhor_global, reg)
                if (i_b + 1) % 50 == 0 or i_b == ITERACOES_BETA - 1:
                    _imprimir_progresso(i_b + 1, ITERACOES_BETA, validos,
                                        melhor_global, (i_b + 1) * ITERACOES_PESOS, t0)
        except KeyboardInterrupt:
            print(f"\n  [Ctrl+C] Interrompido após {len(historico)} rodadas — salvando...")
            interrompido = True

    else:
        # ── Modo multi-process ────────────────────────────────────────────
        chunk         = max(1, ITERACOES_BETA // (n_proc * 4))
        historico     = []
        melhor_global = {"af4": -9999, "rmse": 99999, "valido": False}
        validos = alta_precisao = concluidas = 0
        pool = mp.Pool(processes=n_proc,
                       initializer=_worker_init,
                       initargs=(fixos, tempos_exp, sinal_exp))
        try:
            for reg in pool.imap_unordered(_processar_rodada,
                                           args_list, chunksize=chunk):
                historico.append({k: v for k, v in reg.items()
                                  if not k.startswith("_")})
                validos       += reg["_validos"]
                alta_precisao += reg["_alta_precisao"]
                concluidas    += 1
                _atualizar_melhor_global(melhor_global, reg)
                if concluidas % 50 == 0 or concluidas == ITERACOES_BETA:
                    calc_feitos = concluidas * ITERACOES_PESOS
                    _imprimir_progresso(concluidas, ITERACOES_BETA, validos,
                                        melhor_global, calc_feitos, t0)
                if os.path.exists(arquivo_parar):
                    print(f"\n  [{ARQUIVO_PARAR} detectado] Encerrando após {concluidas} rodadas...")
                    interrompido = True
                    break
        except KeyboardInterrupt:
            print(f"\n  [Ctrl+C] Interrompido após {len(historico)} rodadas — salvando...")
            interrompido = True
        finally:
            pool.terminate()
            pool.join()

        historico.sort(key=lambda r: r["rodada"])

    duracao        = (datetime.datetime.now() - t0).total_seconds()
    total_calculos = len(historico) * ITERACOES_PESOS
    status_str     = " (INTERROMPIDO)" if interrompido else ""
    print(f"\nBusca concluída{status_str} em {duracao:.1f}s  "
          f"({total_calculos / max(duracao, 1):.0f} cálculos/s)  "
          f"[{len(historico)}/{ITERACOES_BETA} rodadas]")

    return {
        "historico":     historico,
        "melhor":        melhor_global,
        "validos":       validos,
        "alta_precisao": alta_precisao,
        "total":         total_calculos,
        "duracao_s":     duracao,
    }


def _atualizar_melhor_global(mg: dict, reg: dict):
    """Atualiza mg in-place com reg se reg for melhor."""
    valido = reg["status"] == "VÁLIDO"
    if valido:
        if not mg["valido"] or reg["rmse"] < mg["rmse"] or            (reg["rmse"] == mg["rmse"] and reg["af4"] > mg["af4"]):
            mg.update({
                "af4": reg["af4"], "rmse": reg["rmse"],
                "m_reta": reg["m_reta"], "b_reta": reg["b_reta"],
                "valido": True,
                **{k: reg[k] for k in reg if not k.startswith("_")
                   and k not in ("rodada","status")},
            })
    else:
        if not mg["valido"] and reg["af4"] > mg.get("af4", -9999):
            mg.update({
                "af4": reg["af4"], "rmse": reg["rmse"],
                "m_reta": reg["m_reta"], "b_reta": reg["b_reta"],
                "valido": False,
                **{k: reg[k] for k in reg if not k.startswith("_")
                   and k not in ("rodada","status")},
            })


def _imprimir_progresso(concluidas, total, validos, melhor_global, calc_feitos, t0):
    pct       = concluidas / total * 100
    decorrido = (datetime.datetime.now() - t0).total_seconds()
    calc_s    = calc_feitos / max(decorrido, 1)
    if pct > 0:
        restante_s = max(0, decorrido / (pct / 100) - decorrido)
        h, rem = divmod(int(restante_s), 3600)
        m, s   = divmod(rem, 60)
        eta_str = f"{h:02d}:{m:02d}:{s:02d}"
    else:
        eta_str = "--:--:--"
    print(f"  {concluidas:>5}/{total}  ({pct:5.1f}%)  "
          f"válidos={validos:>5}  "
          f"melhor R²={melhor_global['af4']:.4f}  "
          f"RMSE={melhor_global['rmse']:.4f}  "
          f"ETA {eta_str}  ({calc_s:.0f} calc/s)")



def busca_duas_fases(tempos_exp, sinal_exp, fixos, pasta_parar="."):
    """
    Busca em duas fases:
      Fase 1 — varredura grossa sobre INTERVALOS_BETA completos
      Fase 2 — busca densa concentrada nos melhores intervalos da fase 1
    """
    global ITERACOES_BETA, ITERACOES_PESOS, INTERVALOS_BETA
    print("\n" + "─" * 70)
    print(f"BUSCA EM DUAS FASES")
    print(f"  Fase 1: {FASE1_ITER_BETA} × {FASE1_ITER_PESOS} = "
          f"{FASE1_ITER_BETA * FASE1_ITER_PESOS:,} combinações (varredura grossa)")
    print(f"  Fase 2: {FASE2_ITER_BETA} × {FASE2_ITER_PESOS} = "
          f"{FASE2_ITER_BETA * FASE2_ITER_PESOS:,} combinações (busca fina)")
    print("─" * 70)

    # ── Fase 1 ────────────────────────────────────────────────────────────────
    _ib_orig, _ip_orig = ITERACOES_BETA, ITERACOES_PESOS
    ITERACOES_BETA, ITERACOES_PESOS = FASE1_ITER_BETA, FASE1_ITER_PESOS

    print("\nFASE 1 — varredura grossa")
    res1 = busca_aleatoria(tempos_exp, sinal_exp, fixos, pasta_parar=pasta_parar)

    # Pegar os melhores da fase 1 para definir novos intervalos
    import pandas as pd
    df1 = pd.DataFrame(res1["historico"])
    validos1 = df1[df1["status"] == "VÁLIDO"].sort_values("rmse")
    if len(validos1) >= FASE2_TOP_REGIOES:
        top = validos1.head(FASE2_TOP_REGIOES)
    elif len(validos1) > 0:
        top = validos1
    else:
        top = df1.sort_values("af4", ascending=False).head(FASE2_TOP_REGIOES)

    # Calcular novos intervalos ± margem ao redor dos melhores
    novos_intervalos = {}
    for k, (lo_orig, hi_orig) in INTERVALOS_BETA.items():
        vals = top[k].values if k in top.columns else []
        if len(vals) > 0:
            span   = max(vals) - min(vals)
            margem = max(span * FASE2_MARGEM, (hi_orig - lo_orig) * 0.05)
            lo_new = max(lo_orig, min(vals) - margem)
            hi_new = min(hi_orig, max(vals) + margem)
            novos_intervalos[k] = (lo_new, hi_new)
        else:
            novos_intervalos[k] = (lo_orig, hi_orig)

    print(f"\n  Novos intervalos para fase 2:")
    for k, (lo, hi) in novos_intervalos.items():
        lo_o, hi_o = INTERVALOS_BETA[k]
        print(f"    {k:<12}: [{lo:.1f}, {hi:.1f}]  "
              f"(era [{lo_o:.1f}, {hi_o:.1f}])")

    # ── Fase 2 ────────────────────────────────────────────────────────────────
    ITERACOES_BETA, ITERACOES_PESOS = FASE2_ITER_BETA, FASE2_ITER_PESOS
    _intervalos_orig = dict(INTERVALOS_BETA)

    INTERVALOS_BETA = novos_intervalos

    print(f"\nFASE 2 — busca fina nos melhores intervalos")
    res2 = busca_aleatoria(tempos_exp, sinal_exp, fixos, pasta_parar=pasta_parar)

    # Restaurar configurações originais
    ITERACOES_BETA, ITERACOES_PESOS = _ib_orig, _ip_orig
    INTERVALOS_BETA = _intervalos_orig

    # Combinar históricos
    for r in res2["historico"]:
        r["rodada"] = f"F2-{r['rodada']}"
    historico_total = res1["historico"] + res2["historico"]

    # Melhor global entre as duas fases
    melhor = res1["melhor"]
    m2 = res2["melhor"]
    if (m2["valido"] and not melhor.get("valido", False)) or \
       (m2["valido"] == melhor.get("valido", False) and
            m2.get("rmse", 99999) < melhor.get("rmse", 99999)):
        melhor = m2
        print("  → Melhor global veio da Fase 2")
    else:
        print("  → Melhor global veio da Fase 1")

    return {
        "historico":     historico_total,
        "melhor":        melhor,
        "validos":       res1["validos"] + res2["validos"],
        "alta_precisao": res1["alta_precisao"] + res2["alta_precisao"],
        "total":         res1["total"] + res2["total"],
        "duracao_s":     res1["duracao_s"] + res2["duracao_s"],
    }


def adaptar_e_rebuscar(tempos_exp, sinal_exp, fixos, resultado_anterior, pasta_parar="."):
    """
    Usa os válidos da busca anterior para estreitar INTERVALOS_BETA
    e roda uma segunda busca mais focada.
    """
    global INTERVALOS_BETA
    import pandas as pd
    df = pd.DataFrame(resultado_anterior["historico"])
    validos = df[df["status"] == "VÁLIDO"]

    if len(validos) == 0:
        print("\n  Adaptação: sem válidos para adaptar intervalos — pulando.")
        return resultado_anterior

    print("\n" + "─" * 70)
    print("ADAPTAÇÃO AUTOMÁTICA DE INTERVALOS")
    print("─" * 70)

    novos = {}
    for k, (lo_orig, hi_orig) in INTERVALOS_BETA.items():
        if k not in validos.columns:
            novos[k] = (lo_orig, hi_orig)
            continue
        vals   = validos[k].values
        span   = max(vals) - min(vals)
        margem = max(span * ADAPT_MARGEM, (hi_orig - lo_orig) * 0.03)
        lo_new = max(lo_orig, min(vals) - margem)
        hi_new = min(hi_orig, max(vals) + margem)
        novos[k] = (lo_new, hi_new)
        print(f"  {k:<12}: [{lo_new:.1f}, {hi_new:.1f}]  "
              f"(era [{lo_orig:.1f}, {hi_orig:.1f}])")

    _orig = dict(INTERVALOS_BETA)
    INTERVALOS_BETA = novos

    print(f"\n  Rodando busca adaptada: {ITERACOES_BETA} × {ITERACOES_PESOS} combinações")
    res2 = busca_aleatoria(tempos_exp, sinal_exp, fixos, pasta_parar=pasta_parar)

    INTERVALOS_BETA = _orig

    for r in res2["historico"]:
        r["rodada"] = f"A-{r['rodada']}"

    historico_total = resultado_anterior["historico"] + res2["historico"]
    melhor = resultado_anterior["melhor"]
    m2 = res2["melhor"]
    if (m2["valido"] and not melhor.get("valido", False)) or \
       (m2["valido"] == melhor.get("valido", False) and
            m2.get("rmse", 99999) < melhor.get("rmse", 99999)):
        melhor = m2
        print("  → Melhor global atualizado pela busca adaptada")
    else:
        print("  → Busca adaptada não melhorou o melhor global")

    return {
        "historico":     historico_total,
        "melhor":        melhor,
        "validos":       resultado_anterior["validos"] + res2["validos"],
        "alta_precisao": resultado_anterior["alta_precisao"] + res2["alta_precisao"],
        "total":         resultado_anterior["total"] + res2["total"],
        "duracao_s":     resultado_anterior["duracao_s"] + res2["duracao_s"],
    }


def selecionar_candidatos_por_clusters(validos_df):
    """
    Agrupa os resultados válidos no espaço normalizado dos betas e
    retorna o melhor (menor RMSE) de cada cluster como candidato para o NM.

    O espaço de features é formado pelos betas normalizados pelos seus
    intervalos de busca — cada dimensão fica em [0, 1].
    Usa clustering hierárquico aglomerativo (linkage completo) com o
    limiar NM_CLUSTER_LIMIAR para definir o corte.

    Retorna lista de dicts (records) com até NM_CLUSTERS_MAX candidatos.
    """
    from scipy.cluster.hierarchy import linkage, fcluster
    from scipy.spatial.distance  import pdist

    if len(validos_df) == 0:
        return []

    # Features: betas normalizados
    features = []
    nomes_beta = ["beta_exp", "beta_g3", "beta_g4", "beta_g5"]
    for nb in nomes_beta:
        lo, hi = INTERVALOS_BETA.get(nb, (0, 1))
        span   = hi - lo if hi > lo else 1.0
        col    = validos_df[nb].values if nb in validos_df.columns else np.zeros(len(validos_df))
        features.append((col - lo) / span)
    X = np.column_stack(features)   # shape (n_validos, 4)

    if len(X) == 1:
        return validos_df.to_dict("records")

    # Clustering hierárquico
    dist   = pdist(X, metric="euclidean")
    Z      = linkage(dist, method="complete")
    labels = fcluster(Z, t=NM_CLUSTER_LIMIAR * np.sqrt(X.shape[1]),
                      criterion="distance")

    n_clusters = int(labels.max())
    print(f"\n  Clusters encontrados: {n_clusters}  "
          f"(limiar={NM_CLUSTER_LIMIAR:.2f}  max={NM_CLUSTERS_MAX})")

    # Melhor de cada cluster (menor RMSE)
    df = validos_df.copy()
    df["_cluster"] = labels
    candidatos = []
    for cid in sorted(df["_cluster"].unique()):
        grupo = df[df["_cluster"] == cid].sort_values("rmse")
        melhor_c = grupo.iloc[0].to_dict()
        melhor_c.pop("_cluster", None)
        candidatos.append(melhor_c)
        print(f"    Cluster {cid:>2}: {len(grupo):>4} resultados  "
              f"melhor RMSE={melhor_c['rmse']:.6f}  "
              f"β_exp={melhor_c.get('beta_exp',0):.1f}  "
              f"β_G3={melhor_c.get('beta_g3',0):.1f}  "
              f"β_G4={melhor_c.get('beta_g4',0):.1f}")

    # Ordenar por RMSE e limitar ao máximo
    candidatos.sort(key=lambda r: r.get("rmse", 99999))
    if len(candidatos) > NM_CLUSTERS_MAX:
        print(f"  → Limitado a {NM_CLUSTERS_MAX} clusters (NM_CLUSTERS_MAX)")
        candidatos = candidatos[:NM_CLUSTERS_MAX]

    return candidatos


def refinar_nelder_mead(tempos_exp, sinal_exp, fixos, candidatos):
    """
    Refinamento local com Nelder-Mead a partir dos melhores candidatos
    da busca aleatória.

    Otimiza simultaneamente os 4 betas e os 4 pesos ativos (os inativos
    ficam em 0). Os pesos são parametrizados em espaço livre (softmax)
    para garantir que somem 1 e sejam ≥ 0 sem precisar de constraints.

    Parâmetros otimizados (8 ou menos, dependendo de CURVAS_PESO_ATIVAS):
      [beta_exp, beta_g3, beta_g4, beta_g5,  ← betas livres
       w0, w1, w2, w3]                        ← logits dos pesos (softmax)

    A função objetivo é o RMSE (não R²) porque RMSE mede a qualidade
    do ajuste de forma mais sensível que R² perto do ótimo.
    """
    from scipy.optimize import minimize

    curvas_ativas  = [k for k, v in CURVAS_PESO_ATIVAS.items() if v]
    nomes_beta     = ["beta_exp", "beta_g3", "beta_g4", "beta_g5"]
    nomes_pesos    = ["p_mb", "p_exp", "p_g3", "p_g4", "p_g5"]
    energias_livres = [k for k in ("e_g3","e_g4","e_g5")
                       if ENERGIA_NM_LIVRE.get(k, False)]

    def softmax_com_piso(w):
        """Softmax que respeita PESO_MIN: aplica piso e renormaliza."""
        e = np.exp(w - np.max(w))
        probs = e / e.sum()
        # Aplicar piso
        ajustado = False
        for idx, nome in enumerate(curvas_ativas):
            pmin = PESO_MIN.get(nome)
            if pmin is not None and probs[idx] < pmin:
                probs[idx] = pmin
                ajustado = True
        if ajustado:
            probs = probs / probs.sum()
        return probs

    def params_de_x(x):
        """Converte vetor x do otimizador em betas, energias e pesos."""
        betas = {
            "beta_exp": float(np.clip(x[0], 1, 10000)),
            "beta_g3":  float(np.clip(x[1], 1, 10000)),
            "beta_g4":  float(np.clip(x[2], 1, 10000)),
            "beta_g5":  float(np.clip(x[3], 1, 10000)),
        }
        # Energias: fixas (ENERGIAS_VALOR ou planilha) ou livres (NM)
        energias = {}
        for k in ("e_g3", "e_g4", "e_g5"):
            if not ENERGIA_NM_LIVRE.get(k, False):
                val = ENERGIAS_VALOR.get(k)
                if val is not None:
                    energias[k] = val      # sobrescreve planilha
                # else: usa fixos (planilha) — não injeta nada
        for idx, k in enumerate(energias_livres):
            lo, hi = INTERVALOS_ENERGIA[k]
            sig = 1.0 / (1.0 + np.exp(-x[4 + idx]))   # sigmoid ∈ (0,1)
            energias[k] = lo + sig * (hi - lo)
        n_e    = len(energias_livres)
        logits = x[4 + n_e:]
        probs  = softmax_com_piso(logits)
        pesos  = {k: 0.0 for k in nomes_pesos}
        for idx, nome in enumerate(curvas_ativas):
            pesos[nome] = float(probs[idx])
        return betas, energias, pesos

    def objetivo(x):
        betas, energias, pesos = params_de_x(x)
        fixos_nm = {**fixos, **energias}
        amps = normalizar_amplitudes(
            fixos_nm, betas["beta_exp"], betas["beta_g3"],
            betas["beta_g4"], betas["beta_g5"]
        )
        sinal_teo = calcular_espectro(
            tempos_exp,
            betas["beta_exp"], betas["beta_g3"],
            betas["beta_g4"], betas["beta_g5"],
            pesos["p_mb"], pesos["p_exp"],
            pesos["p_g3"], pesos["p_g4"], pesos["p_g5"],
            fixos_nm, amps
        )
        rmse, _, _ = calcular_rmse(tempos_exp, sinal_exp, sinal_teo)
        return rmse

    resultados_nm = []
    print(f"\n{'─'*70}")
    print(f"REFINAMENTO NELDER-MEAD  ({len(candidatos)} candidato(s))")
    print(f"{'─'*70}")

    # Escalas de cada dimensão para o simplex inicial e perturbação
    # betas: escala = amplitude do intervalo / 4
    # energias: escala baseada em INTERVALOS_ENERGIA
    # logits (pesos): escala fixa de 1.0 (espaço livre)
    escala_betas = [(INTERVALOS_BETA.get(k, (0,1000))[1] -
                     INTERVALOS_BETA.get(k, (0,1000))[0]) / 4
                    for k in ["beta_exp","beta_g3","beta_g4","beta_g5"]]
    escala_energ = [(INTERVALOS_ENERGIA[k][1] - INTERVALOS_ENERGIA[k][0]) / 4
                    for k in energias_livres]
    n_pesos      = len(curvas_ativas)
    escala_total = np.array(escala_betas + escala_energ + [1.0] * n_pesos)

    rng_nm = np.random.default_rng(42)   # seed fixa para reinícios reproduzíveis

    for i, cand in enumerate(candidatos):
        af4_ini  = cand.get("af4",  0)
        rmse_ini = cand.get("rmse", 99)
        print(f"  Candidato {i+1}: R²={af4_ini:.6f}  RMSE={rmse_ini:.6f}  "
              f"→ otimizando...", end="", flush=True)

        def _x0_de_cand(c):
            """Constrói vetor x0 a partir de um candidato."""
            x_betas = [c.get("beta_exp", 400), c.get("beta_g3", 500),
                       c.get("beta_g4",  300), c.get("beta_g5", 500)]
            x_energ = []
            for k in energias_livres:
                lo, hi  = INTERVALOS_ENERGIA[k]
                val     = float(np.clip(c.get(k, (lo+hi)/2), lo+1e-6, hi-1e-6))
                sig_val = np.clip((val - lo) / (hi - lo), 1e-6, 1-1e-6)
                x_energ.append(float(np.log(sig_val / (1 - sig_val))))
            probs0  = np.array([max(c.get(k, 0.01), 1e-6) for k in curvas_ativas])
            probs0  = probs0 / probs0.sum()
            logits0 = np.log(probs0)
            return np.array(x_betas + x_energ + list(logits0))

        def _simplex_inicial(x0, escala):
            """Constrói simplex N+1 pontos com passo proporcional à escala."""
            n   = len(x0)
            sim = np.tile(x0, (n + 1, 1))
            for j in range(n):
                sim[j + 1, j] += escala[j]
            return sim

        def _rodar_nm(x0_start):
            sim0 = _simplex_inicial(x0_start, escala_total)
            return minimize(
                objetivo, x0_start,
                method="Nelder-Mead",
                options={
                    "maxiter":        NM_MAX_ITER,
                    "xatol":          NM_TOLERANCIA,
                    "fatol":          NM_TOLERANCIA,
                    "adaptive":       True,
                    "initial_simplex": sim0,
                }
            )

        # ── Primeira tentativa ────────────────────────────────────────────────
        x0  = _x0_de_cand(cand)
        res = _rodar_nm(x0)
        iter_total = res.nit

        # ── Reinícios se não convergiu ou se vale tentar melhorar ─────────────
        for reinicio in range(1, NM_MAX_REINICIAR):
            # Perturba o melhor ponto encontrado até agora
            perturb  = rng_nm.uniform(-1, 1, size=len(res.x)) * escala_total * NM_PERTURB_ESCALA
            x0_new   = res.x + perturb
            res_new  = _rodar_nm(x0_new)
            iter_total += res_new.nit
            if res_new.fun < res.fun:
                res = res_new   # melhorou — guarda

        betas_nm, energias_nm, pesos_nm = params_de_x(res.x)
        fixos_nm = {**fixos, **energias_nm}
        amps_nm  = normalizar_amplitudes(
            fixos_nm, betas_nm["beta_exp"], betas_nm["beta_g3"],
            betas_nm["beta_g4"], betas_nm["beta_g5"]
        )
        sinal_nm = calcular_espectro(
            tempos_exp,
            betas_nm["beta_exp"], betas_nm["beta_g3"],
            betas_nm["beta_g4"],  betas_nm["beta_g5"],
            pesos_nm["p_mb"], pesos_nm["p_exp"],
            pesos_nm["p_g3"], pesos_nm["p_g4"], pesos_nm["p_g5"],
            fixos_nm, amps_nm
        )
        af4_nm            = calcular_af4(sinal_exp, sinal_nm)
        rmse_nm, m_r, b_r = calcular_rmse(tempos_exp, sinal_exp, sinal_nm)
        valido_nm         = af4_nm >= SCORE_MINIMO

        delta_af4  = af4_nm  - af4_ini
        delta_rmse = rmse_nm - rmse_ini
        conv_str   = "✓" if res.success else f"~{NM_MAX_REINICIAR} tent."
        print(f"  R²={af4_nm:.6f} ({delta_af4:+.6f})  "
              f"RMSE={rmse_nm:.6f} ({delta_rmse:+.6f})  "
              f"iter={iter_total} {conv_str}")

        reg = {
            "rodada":       f"NM{i+1}",
            **betas_nm,
            **energias_nm,
            **pesos_nm,
            "af4":          af4_nm,
            "rmse":         rmse_nm,
            "m_reta":       m_r,
            "b_reta":       b_r,
            "valido":       valido_nm,
            "status":       "VÁLIDO" if valido_nm else "Abaixo do Corte",
            "nm_iter":      iter_total,
            "nm_convergiu": res.success,
            "beta_mb":      fixos.get("beta_mb", 0),
        }
        resultados_nm.append(reg)

    # Melhor entre os refinados
    validos_nm  = [r for r in resultados_nm if r["valido"]]
    pool        = validos_nm if validos_nm else resultados_nm
    melhor_nm   = min(pool, key=lambda r: r["rmse"])

    print(f"{'─'*70}")
    print(f"  Melhor NM: AF4={melhor_nm['af4']:.6f}  RMSE={melhor_nm['rmse']:.6f}  "
          f"Status={'VÁLIDO ✓' if melhor_nm['valido'] else 'Abaixo do critério'}")
    print(f"{'─'*70}\n")

    return resultados_nm, melhor_nm

# =============================================================================
# SEÇÃO 11 — GRÁFICOS E VIEWER INTERATIVO
# =============================================================================

def _componentes_grade(params: dict, fixos: dict) -> dict:
    """Calcula cada curva na grade BM (170 pts), normalizado por u0."""
    amps    = normalizar_amplitudes(fixos, params["beta_exp"], params["beta_g3"],
                                    params["beta_g4"], params["beta_g5"])
    mf      = fixos["m_frag"]
    beta_mb = fixos["beta_mb"]
    v0_g3   = energia_para_v0(params.get("e_g3") or fixos["e_g3"], mf)
    v0_g4   = energia_para_v0(params.get("e_g4") or fixos["e_g4"], mf)
    v0_g5   = energia_para_v0(params.get("e_g5") or fixos["e_g5"], mf)
    b2e  = params["beta_exp"]**2
    b2g3 = params["beta_g3"]**2
    b2g4 = params["beta_g4"]**2
    b2g5 = params["beta_g5"]**2
    gh_e_v  = gh_exp(b2e)
    gh_g3_v = gh_gauss(b2g3, v0_g3)
    gh_g4_v = gh_gauss(b2g4, v0_g4)
    gh_g5_v = gh_gauss(b2g5, v0_g5)

    n  = len(GRADE_BM)
    ac = np.zeros(n); am = np.zeros(n)
    av = np.zeros(n); be = np.zeros(n); bn = np.zeros(n)

    for i, t in enumerate(GRADE_BM):
        p9, u9 = calcular_filtros(t)
        ac[i] = np.sum(u9 * params["p_mb"]  * amps["x1"]
                       * H9_VEC * np.exp(-((beta_mb * H9_VEC)**2)) * H9_VEC / 2)
        am[i] = np.sum(p9 * params["p_exp"] * amps["ak3"]
                       * H9_VEC * np.exp(-(b2e * H9_VEC**2)) * gh_e_v * H9_VEC / 2)
        y_g3  = (1/2)*H9_VEC*np.exp(-(b2g3**2*(H9_VEC**2-v0_g3**2)**2))*gh_g3_v
        av[i] = np.sum(p9 * params["p_g3"]  * amps["au3"] * y_g3 * H9_VEC / 2)
        y_g4  = (1/2)*H9_VEC*np.exp(-(b2g4**2*(H9_VEC**2-v0_g4**2)**2))*gh_g4_v
        be[i] = np.sum(p9 * params["p_g4"]  * amps["bd3"] * y_g4 * H9_VEC / 2)
        y_g5  = (1/2)*H9_VEC*np.exp(-(b2g5**2*(H9_VEC**2-v0_g5**2)**2))*gh_g5_v
        bn[i] = np.sum(p9 * params["p_g5"]  * amps["bm3"] * y_g5 * H9_VEC / 2)

    total = ac * FMB + am + av + be + bn
    u0    = total[0]
    return {"ac": ac*FMB/u0, "am": am/u0, "av": av/u0,
            "be": be/u0, "bn": bn/u0, "total": total/u0}


def _desenhar_nos_eixos(ax1, ax2, tempos_exp, sinal_exp, params, fixos, titulo):
    """Preenche ax1 (log) e ax2 (resíduo) para um dado conjunto de parâmetros."""
    comp    = _componentes_grade(params, fixos)
    t       = GRADE_BM
    tot     = comp["total"]
    teo_exp = np.interp(tempos_exp, t, tot)

    af4              = params.get("af4", calcular_af4(sinal_exp, teo_exp))
    rmse_val         = calcular_rmse(tempos_exp, sinal_exp, teo_exp)
    if isinstance(rmse_val, tuple):
        rmse, m_reta, b_reta = rmse_val
    else:
        rmse, m_reta, b_reta = rmse_val, params.get("m_reta", 0.0), params.get("b_reta", 0.0)
    if "rmse" in params:
        rmse   = params["rmse"]
        m_reta = params.get("m_reta", m_reta)
        b_reta = params.get("b_reta", b_reta)

    # Resíduo relativo E = (teo-exp)/exp  para o painel inferior
    mask   = np.abs(sinal_exp) > 1e-12
    E      = np.where(mask, (teo_exp - sinal_exp) / sinal_exp, 0.0)
    t_tend = np.array([tempos_exp[0], tempos_exp[-1]])
    y_tend = m_reta * t_tend + b_reta
    eq_str = (f"y = {m_reta:.2e}x "
              f"{'+ ' if b_reta >= 0 else '- '}{abs(b_reta):.4f}")

    # ── Painel log ────────────────────────────────────────────────────────────
    ax1.cla()
    ax1.plot(tempos_exp, sinal_exp, 'ko', ms=5, zorder=6, label="Experimental")
    ax1.plot(t, tot, color="#e31a1c", lw=2, label="Total")
    ax1.plot(t, comp["ac"], color="#9b59b6", ls="--", lw=1.5,
             label=f"MB   {params['p_mb']:.6f}")
    ax1.plot(t, comp["am"], color="#2980b9", ls="--", lw=1.5,
             label=f"EXP  {params['p_exp']:.6f}")
    ax1.plot(t, comp["av"], color="#27ae60", ls="--", lw=1.5,
             label=f"G1   {params['p_g3']:.6f}")
    ax1.plot(t, comp["be"], color="#f39c12", ls="--", lw=1.5,
             label=f"G2   {params['p_g4']:.6f}")
    if np.any(comp["bn"] > 1e-10):
        ax1.plot(t, comp["bn"], color="#c0392b", ls="--", lw=1.5,
                 label=f"G3   {params['p_g5']:.6f}")

    ax1.set_xscale("log"); ax1.set_yscale("log")
    tmin = min(np.min(tempos_exp), t[0])
    tmax = max(np.max(tempos_exp), t[-1])
    ax1.set_xlim(tmin * 0.9, tmax * 1.1)
    spos = sinal_exp[sinal_exp > 0]
    ax1.set_ylim(np.min(spos) * 0.3 if len(spos) else 1e-4, 3.0)
    ax1.set_title(
        f"{titulo}\n"
        f"R²={af4:.6f}   RMSE={rmse:.6f}   "
        f"β_exp={params['beta_exp']:.1f}  β_G1={params['beta_g3']:.1f}  "
        f"β_G2={params['beta_g4']:.1f}  β_G3={params['beta_g5']:.1f}  "
        f"Reta: {eq_str}",
        fontsize=9
    )
    ax1.set_ylabel("Sinal normalizado")
    ax1.legend(fontsize=8, ncol=3, loc="lower left")
    plt.setp(ax1.get_xticklabels(), visible=False)

    # ── Painel resíduo relativo E=(teo-exp)/exp com reta de tendência ──────────
    ax2.cla()
    cores = ["tomato" if e > 0 else "steelblue" for e in E]
    ax2.axhline(0, color="k", lw=0.8, zorder=1)
    # Reta de tendência (ajustada sobre D, exibida sobre E)
    ax2.plot(t_tend, y_tend, 'k--', lw=1.2, zorder=4,
             label=f"Tendência: {eq_str}")
    for xi, ei, ci in zip(tempos_exp, E, cores):
        ax2.plot([xi, xi], [0, ei], color=ci, lw=1.5, zorder=2)
    ax2.scatter(tempos_exp, E, c=cores, s=25, zorder=3)
    lim = max(np.abs(E).max() * 1.5, 1e-4) if len(E) else 0.05
    ax2.set_ylim(-lim, lim)
    ax2.set_xlim(tmin * 0.9, tmax * 1.1)
    ax2.set_xscale("log")
    ax2.set_xlabel("Tempo (ns)")
    ax2.set_ylabel("Resíduo relativo\n(teo-exp)/exp")
    ax2.legend(fontsize=7, loc="upper right")
    ax2.yaxis.set_major_formatter(plt.FormatStrFormatter("%.4f"))


# Cores das curvas — usadas no gráfico e no painel lateral
COR_CURVA = {
    "p_mb":  "#9b59b6",   # roxo
    "p_exp": "#2980b9",   # azul
    "p_g3":  "#27ae60",   # verde
    "p_g4":  "#f39c12",   # laranja/amarelo
    "p_g5":  "#c0392b",   # vermelho
}


def _painel_parametros(ax, params, fixos):
    """Painel lateral com tabela de parâmetros usando matplotlib.table."""
    ax.axis("off")
    ax.set_facecolor("#f5f5f5")
    for spine in ax.spines.values():
        spine.set_visible(False)

    soma_p = sum(params.get(k, 0) for k in ["p_mb","p_exp","p_g3","p_g4","p_g5"])
    af4    = params.get("af4",  0)
    rmse   = params.get("rmse", 0)
    rod    = params.get("rodada", "—")
    status = params.get("status", "—")
    is_nm  = str(rod).startswith("NM")

    # Texto de cabeçalho
    ax.text(0.5, 0.99, f"Rodada {rod}", fontsize=9, fontweight="bold",
            ha="center", va="top", transform=ax.transAxes,
            color="#c0392b" if is_nm else "#2c3e50")
    ax.text(0.5, 0.93, status, fontsize=8,
            ha="center", va="top", transform=ax.transAxes,
            color="#27ae60" if "VÁLIDO" in str(status) else "#e74c3c")

    # Linha R² / RMSE
    ax.text(0.05, 0.87, f"R²   = {af4:.6f}", fontsize=8, va="top",
            transform=ax.transAxes, family="monospace")
    ax.text(0.05, 0.81, f"RMSE = {rmse:.6f}", fontsize=8, va="top",
            transform=ax.transAxes, family="monospace")

    # Separador
    ax.plot([0, 1], [0.78, 0.78], color="#aaaaaa", lw=0.8, transform=ax.transAxes, clip_on=False)

    # Cabeçalho da tabela de curvas
    ax.text(0.05, 0.74, "Curva",    fontsize=7.5, va="top", fontweight="bold",
            transform=ax.transAxes, family="monospace")
    ax.text(0.55, 0.74, "Larg.",    fontsize=7.5, va="top", fontweight="bold",
            ha="right", transform=ax.transAxes, family="monospace")
    ax.text(0.98, 0.74, "Peso",     fontsize=7.5, va="top", fontweight="bold",
            ha="right", transform=ax.transAxes, family="monospace")
    ax.plot([0, 1], [0.71, 0.71], color="#aaaaaa", lw=0.5, transform=ax.transAxes, clip_on=False)

    # Linhas de curvas
    pares = [("beta_mb","p_mb"),("beta_exp","p_exp"),
             ("beta_g3","p_g3"),("beta_g4","p_g4"),("beta_g5","p_g5")]
    y = 0.67
    dy = 0.10
    for kb, kp in pares:
        nome  = NOME_CURVA.get(kp, kp)
        cor   = COR_CURVA.get(kp, "black")
        b_val = fixos.get(kb) or params.get(kb, 0)
        p_val = params.get(kp, 0)
        ax.text(0.05, y, nome,             fontsize=7.5, va="top",
                transform=ax.transAxes, family="monospace", color=cor)
        ax.text(0.55, y, f"{b_val:.1f}",   fontsize=7.5, va="top", ha="right",
                transform=ax.transAxes, family="monospace")
        ax.text(0.98, y, f"{p_val:.6f}",   fontsize=7.5, va="top", ha="right",
                transform=ax.transAxes, family="monospace", color=cor)
        y -= dy

    ax.plot([0, 1], [y + dy*0.3, y + dy*0.3], color="#aaaaaa", lw=0.5, transform=ax.transAxes, clip_on=False)
    ax.text(0.98, y, f"{soma_p:.6f}", fontsize=7.5, va="top", ha="right",
            transform=ax.transAxes, family="monospace",
            color="#27ae60" if abs(soma_p - 1.0) < 1e-9 else "#e74c3c")
    ax.text(0.05, y, "Soma",         fontsize=7.5, va="top",
            transform=ax.transAxes, family="monospace")


def gerar_e_salvar_grafico(tempos_exp, sinal_exp, params, fixos,
                            titulo, caminho):
    """Gera figura e salva em disco sem exibir."""
    fig     = plt.figure(figsize=(11, 7), constrained_layout=True)
    gs      = gridspec.GridSpec(2, 1, height_ratios=[3, 1], hspace=0.06, figure=fig)
    ax1     = fig.add_subplot(gs[0])
    ax2     = fig.add_subplot(gs[1], sharex=ax1)
    _desenhar_nos_eixos(ax1, ax2, tempos_exp, sinal_exp, params, fixos, titulo)
    fig.savefig(caminho, dpi=150, bbox_inches="tight")
    plt.close(fig)


# Paleta para comparação de rodadas — cicla entre essas cores
_CORES_COMPARACAO = [
    "#e31a1c", "#1f78b4", "#33a02c", "#ff7f00",
    "#6a3d9a", "#b15928", "#a6cee3", "#fb9a99",
]


def _grafico_comparacao(registros, tempos_exp, sinal_exp, fixos, titulo_base):
    """Abre nova figura sobrepondo múltiplas rodadas — só a curva Total de cada uma."""
    fig = plt.figure(figsize=(12, 7), constrained_layout=True)
    gs  = gridspec.GridSpec(2, 1, height_ratios=[3, 1], hspace=0.06, figure=fig)
    ax1 = fig.add_subplot(gs[0])
    ax2 = fig.add_subplot(gs[1], sharex=ax1)

    ax1.plot(tempos_exp, sinal_exp, "ko", ms=4, zorder=6, label="Experimental")
    mask = np.abs(sinal_exp) > 1e-12
    ax2.axhline(0, color="k", lw=0.8)
    ax2.set_ylabel("Resíduo (teo−exp)/exp", fontsize=8)
    ax2.set_xlabel("Tempo (ns)", fontsize=9)

    estilos = ["-", "--", "-.", ":"]
    for i, reg in enumerate(registros):
        cor = _CORES_COMPARACAO[i % len(_CORES_COMPARACAO)]
        ls  = estilos[i % len(estilos)]
        rod  = reg.get("rodada", i+1)
        af4  = reg.get("af4",  0)
        rmse = reg.get("rmse", 0)

        fixos_reg = {**fixos,
                     **{k: reg[k] for k in ("e_g3","e_g4","e_g5") if k in reg}}
        sinal_teo = calcular_espectro(
            tempos_exp,
            reg.get("beta_exp", 0), reg.get("beta_g3", 0),
            reg.get("beta_g4",  0), reg.get("beta_g5", 0),
            reg.get("p_mb",  0), reg.get("p_exp", 0),
            reg.get("p_g3",  0), reg.get("p_g4",  0),
            reg.get("p_g5",  0), fixos_reg
        )

        ax1.plot(tempos_exp, sinal_teo, color=cor, lw=2.0, ls=ls,
                 label=f"Rodada {rod}  R²={af4:.4f}  RMSE={rmse:.4f}", zorder=5)

        E = np.where(mask, (sinal_teo - sinal_exp) / sinal_exp, 0.0)
        ax2.plot(tempos_exp, E, color=cor, lw=1.5, ls=ls)

    ax1.set_xscale("log"); ax1.set_yscale("log")
    spos = sinal_exp[sinal_exp > 0]
    if len(spos):
        ax1.set_ylim(spos.min() * 0.5, spos.max() * 2)
    ax1.legend(fontsize=8, framealpha=0.9)
    ax1.set_title(f"{titulo_base}  —  Comparação de {len(registros)} rodadas", fontsize=10)
    ax1.set_ylabel("Sinal normalizado", fontsize=9)
    ax2.set_xscale("log")
    ax2.set_ylim(-0.3, 0.3)
    plt.show()


class ViewerRodadas:
    """
    Viewer interativo.
    Teclas:
      ← / →   navegar entre rodadas
      M        marcar rodada atual para comparação
      C        comparar todas as rodadas marcadas num único gráfico
      X        limpar marcação
      Q        fechar
    """
    def __init__(self, historico, tempos_exp, sinal_exp, fixos, titulo="DETOF"):
        todos = sorted(
            [r for r in historico if r["status"] == "VÁLIDO"],
            key=lambda r: r["rmse"]
        )
        if VIEWER_TOP is not None and len(todos) > VIEWER_TOP:
            print(f"  (viewer limitado aos {VIEWER_TOP} melhores de {len(todos)} válidos)")
            todos = todos[:VIEWER_TOP]
        self.registros  = todos
        self.tempos_exp  = tempos_exp
        self.sinal_exp   = sinal_exp
        self.fixos       = fixos
        self.titulo_base = titulo
        self.idx         = 0
        self.selecionados = set()   # índices marcados para comparação
        self.fig  = None
        self.ax1  = None
        self.ax2  = None

    def mostrar(self):
        if not self.registros:
            print("Nenhum resultado válido para exibir.")
            return
        self.fig  = plt.figure(figsize=(11, 7), constrained_layout=True)
        self.fig.canvas.mpl_connect("key_press_event", self._on_key)
        gs        = gridspec.GridSpec(2, 1, height_ratios=[3, 1], hspace=0.06,
                                      figure=self.fig)
        self.ax1  = self.fig.add_subplot(gs[0])
        self.ax2  = self.fig.add_subplot(gs[1], sharex=self.ax1)
        self._desenhar()
        plt.show()

    def _desenhar(self):
        reg = self.registros[self.idx]
        n   = len(self.registros)
        sel_mark = " ★" if self.idx in self.selecionados else ""
        n_sel = len(self.selecionados)
        hints = "← → navegar | M marcar | C comparar | X limpar | Q fechar"
        if n_sel:
            hints = f"[{n_sel} selecionada(s)] " + hints
        titulo = (f"{self.titulo_base}  —  [{self.idx+1}/{n}]  "
                  f"Rodada {reg['rodada']}{sel_mark}  [{reg['status']}]  "
                  f"({hints})")
        _desenhar_nos_eixos(self.ax1, self.ax2,
                            self.tempos_exp, self.sinal_exp,
                            reg, self.fixos, titulo)
        self.fig.canvas.draw_idle()

    def _on_key(self, event):
        if event.key in ("right", "d"):
            self.idx = (self.idx + 1) % len(self.registros)
            self._desenhar()
        elif event.key in ("left", "a"):
            self.idx = (self.idx - 1) % len(self.registros)
            self._desenhar()
        elif event.key in ("m", "M"):
            if self.idx in self.selecionados:
                self.selecionados.discard(self.idx)
                print(f"  Rodada {self.registros[self.idx]['rodada']} desmarcada "
                      f"({len(self.selecionados)} selecionadas)")
            else:
                self.selecionados.add(self.idx)
                print(f"  Rodada {self.registros[self.idx]['rodada']} marcada ★ "
                      f"({len(self.selecionados)} selecionadas)")
            self._desenhar()
        elif event.key in ("c", "C"):
            if len(self.selecionados) < 2:
                print("  Selecione ao menos 2 rodadas com S antes de comparar.")
            else:
                regs = [self.registros[i]
                        for i in sorted(self.selecionados)]
                _grafico_comparacao(regs, self.tempos_exp, self.sinal_exp,
                                    self.fixos, self.titulo_base)
        elif event.key in ("x", "X"):
            self.selecionados.clear()
            print("  Seleção limpa.")
            self._desenhar()
        elif event.key in ("q", "Q"):
            plt.close(self.fig)


# =============================================================================
# SEÇÃO 12 — RELATÓRIO EXCEL (formato compatível com macro VBA)
# =============================================================================

def _prox_numero_relatorio(pasta):
    """Retorna o próximo número de relatório com base nos arquivos já existentes."""
    from pathlib import Path
    existentes = list(Path(pasta).glob("*_Relatorio*.xlsx"))
    nums = []
    for p in existentes:
        partes = p.stem.rsplit("_Relatorio", 1)
        if len(partes) == 2 and partes[1].isdigit():
            nums.append(int(partes[1]))
    return max(nums, default=0) + 1


def _eq_reta_str(m, b):
    """Formata equação da reta no estilo VBA: 'y = 3,1E-06x - 0,0141'."""
    m_fmt = f"{m:.1E}".replace("E+0", "E+").replace("E-0", "E-").replace(".", ",")
    # remove leading zeros no expoente: E+06 → E+6, mas E-06 → E-6
    import re
    m_fmt = re.sub(r"E([+-])0*(\d+)", lambda x: f"E{x.group(1)}{x.group(2)}", m_fmt)
    sinal  = "-" if b < 0 else "+"
    b_fmt  = f"{abs(b):.4f}".replace(".", ",")
    return f"y = {m_fmt}x {sinal} {b_fmt}"


def escrever_relatorio_excel(caminho_excel, resultado, melhor, fixos,
                              t_inicio, t_fim, historico_completo, pasta_exec=None):
    """
    Escreve uma nova aba 'RelatórioN' na planilha Excel existente,
    replicando exatamente o formato gerado pelo VBA.

    Parâmetros
    ----------
    caminho_excel      : caminho para o .xlsm
    resultado          : dict retornado por busca_aleatoria
    melhor             : dict do melhor resultado global
    fixos              : dict de parâmetros fixos
    t_inicio / t_fim   : datetime do início e fim da busca
    historico_completo : lista com todas as rodadas (aleatória + NM)
    """
    from openpyxl import Workbook
    from pathlib import Path
    import datetime as dt

    # Pasta de destino — usa pasta_exec se fornecida, senão RESULTADOS_DIR
    pasta_res = Path(pasta_exec) if pasta_exec else Path(RESULTADOS_DIR)
    pasta_res.mkdir(parents=True, exist_ok=True)
    num      = _prox_numero_relatorio(pasta_res)
    nome_base_xl = Path(caminho_excel).stem
    stamp_xl = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    destino  = pasta_res / f"{nome_base_xl}_Relatorio{num}_{stamp_xl}.xlsx"

    # Workbook completamente novo — nunca toca no .xlsm original
    wb  = Workbook()
    ws  = wb.active
    ws.title = f"Relatório{num}"

    dur_s   = resultado["duracao_s"]
    h, rem  = divmod(int(dur_s), 3600)
    m_, s_  = divmod(rem, 60)
    dur_time = dt.time(h % 24, m_, s_)   # datetime.time (como VBA)

    calc_por_s = resultado["total"] / max(dur_s, 1)
    t_medio_ms = dur_s / max(resultado["total"], 1) * 1000   # ms por cálculo

    # Todos os válidos ordenados por R² decrescente (igual ao VBA)
    todos = sorted(historico_completo, key=lambda r: -r.get("af4", 0))

    validos = [r for r in todos if r.get("status") == "VÁLIDO"]

    def cv(x):
        """Float para string com vírgula (para campos texto estilo VBA)."""
        return f"{float(x):.1f}".replace(".", ",")

    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    def titulo(ws, row, texto):
        ws.cell(row=row, column=1, value=texto).font = Font(bold=True)

    def par(ws, row, label, valor):
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=valor)

    r = 1  # cursor de linha

    # ── METADADOS ─────────────────────────────────────────────────────────────
    titulo(ws, r, "METADADOS DA OTIMIZAÇÃO"); r += 1
    par(ws, r, "Início",               t_inicio.strftime("%d/%m/%Y %H:%M:%S")); r += 1
    par(ws, r, "Fim",                  t_fim.strftime("%d/%m/%Y %H:%M:%S")); r += 1
    par(ws, r, "Duração",              dur_time); r += 1
    par(ws, r, "Iterações Beta",       ITERACOES_BETA); r += 1
    par(ws, r, "Iterações Pesos",      ITERACOES_PESOS); r += 1
    par(ws, r, "Cálculos realizados",  resultado["total"]); r += 1
    par(ws, r, "Calc/s",               round(calc_por_s, 1)); r += 1
    par(ws, r, f"Válidos (R² ≥ {SCORE_MINIMO})", resultado["validos"]); r += 1
    par(ws, r, f"Alta precisão (RMSE ≤ {DESVIO_ALVO})", resultado["alta_precisao"]); r += 1
    r += 1  # linha em branco

    # ── NELDER-MEAD ───────────────────────────────────────────────────────────
    titulo(ws, r, "REFINAMENTO NELDER-MEAD"); r += 1
    par(ws, r, "Ativo",           "Sim" if USAR_NELDER_MEAD else "Não"); r += 1
    if USAR_NELDER_MEAD:
        nm_regs = [x for x in todos if str(x.get("rodada","")).startswith("NM")]
        nm_val  = [x for x in nm_regs if x.get("status") == "VÁLIDO"]
        par(ws, r, "Candidatos refinados",  NM_TOP_CANDIDATOS); r += 1
        par(ws, r, "Max iterações",         NM_MAX_ITER); r += 1
        par(ws, r, "Max reinícios",         NM_MAX_REINICIAR); r += 1
        par(ws, r, "Tolerância",            NM_TOLERANCIA); r += 1
        par(ws, r, "Rodadas NM",            len(nm_regs)); r += 1
        par(ws, r, "Válidos NM",            len(nm_val)); r += 1
    r += 1

    # ── INTERVALOS DE BUSCA ───────────────────────────────────────────────────
    titulo(ws, r, "INTERVALOS DE BUSCA (Beta)"); r += 1
    for nome, (lo, hi) in INTERVALOS_BETA.items():
        par(ws, r, f"  {nome}", f"De {cv(lo)} até {cv(hi)}"); r += 1
    r += 1
    titulo(ws, r, "ENERGIAS DAS GAUSSIANAS"); r += 1
    for k in ("e_g3", "e_g4", "e_g5"):
        val_e  = ENERGIAS_VALOR.get(k)
        nm_liv = ENERGIA_NM_LIVRE.get(k, False)
        if nm_liv:
            lo, hi = INTERVALOS_ENERGIA[k]
            src = f"{cv(val_e)} eV" if val_e else "planilha"
            par(ws, r, f"  {k}", f"NM livre [{cv(lo)}–{cv(hi)}] eV  ponto={src}"); r += 1
        elif val_e is not None:
            par(ws, r, f"  {k}", f"Fixo {cv(val_e)} eV (config)"); r += 1
        else:
            par(ws, r, f"  {k}", "Fixo (planilha)"); r += 1
    r += 1

    # ── MELHOR RESULTADO GLOBAL ───────────────────────────────────────────────
    titulo(ws, r, "MELHOR RESULTADO GLOBAL"); r += 1
    par(ws, r, "Rodada",          melhor.get("rodada", "")); r += 1
    par(ws, r, "R²",              melhor.get("af4", 0)); r += 1
    par(ws, r, "RMSE",            melhor.get("rmse", 0)); r += 1
    par(ws, r, "Equação da reta", _eq_reta_str(melhor.get("m_reta",0), melhor.get("b_reta",0))); r += 1
    par(ws, r, "Beta Exponencial",    melhor.get("beta_exp", 0)); r += 1
    par(ws, r, "Beta Gaussiana 1",    melhor.get("beta_g3", 0)); r += 1
    par(ws, r, "Beta Gaussiana 2",    melhor.get("beta_g4", 0)); r += 1
    par(ws, r, "Beta Gaussiana 3",    melhor.get("beta_g5", 0)); r += 1
    par(ws, r, "Energia Gaussiana 1 (eV)", melhor.get("e_g3") or fixos.get("e_g3",0)); r += 1
    par(ws, r, "Energia Gaussiana 2 (eV)", melhor.get("e_g4") or fixos.get("e_g4",0)); r += 1
    par(ws, r, "Energia Gaussiana 3 (eV)", melhor.get("e_g5") or fixos.get("e_g5",0)); r += 1
    par(ws, r, "Peso Maxwell-Boltzmann",   melhor.get("p_mb", 0)); r += 1
    par(ws, r, "Peso Exponencial",         melhor.get("p_exp", 0)); r += 1
    par(ws, r, "Peso Gaussiana 1",         melhor.get("p_g3", 0)); r += 1
    par(ws, r, "Peso Gaussiana 2",         melhor.get("p_g4", 0)); r += 1
    par(ws, r, "Peso Gaussiana 3",         melhor.get("p_g5", 0)); r += 1
    r += 1



    # ── CABEÇALHO DA TABELA ───────────────────────────────────────────────────
    cabecalho = [
        "Rodada", "Status", "R²", "RMSE", "Equação da Reta",
        "Beta Exp", "Beta G1", "Beta G2", "Beta G3",
        "Energia G1 (eV)", "Energia G2 (eV)", "Energia G3 (eV)",
        "Peso MB", "Peso Exp", "Peso G1", "Peso G2", "Peso G3",
    ]
    cab_fill = PatternFill("solid", fgColor="2C3E50")
    cab_font = Font(bold=True, color="FFFFFF")
    for col, titulo_col in enumerate(cabecalho, start=1):
        c = ws.cell(row=r, column=col, value=titulo_col)
        c.font = cab_font
        c.fill = cab_fill
        c.alignment = Alignment(horizontal="center")
    r += 1

    # ── DADOS ─────────────────────────────────────────────────────────────────
    fill_valido  = PatternFill("solid", fgColor="D5F5E3")
    fill_invalido = PatternFill("solid", fgColor="FADBD8")
    for reg in todos:
        valido = reg.get("status") == "VÁLIDO"
        fill   = fill_valido if valido else fill_invalido
        eq     = _eq_reta_str(reg.get("m_reta", 0), reg.get("b_reta", 0))
        vals   = [
            reg.get("rodada", ""),
            reg.get("status", ""),
            reg.get("af4", 0),
            reg.get("rmse", 0),
            eq,
            reg.get("beta_exp", 0),
            reg.get("beta_g3",  0),
            reg.get("beta_g4",  0),
            reg.get("beta_g5",  0),
            reg.get("e_g3") or fixos.get("e_g3", 0),
            reg.get("e_g4") or fixos.get("e_g4", 0),
            reg.get("e_g5") or fixos.get("e_g5", 0),
            reg.get("p_mb",  0),
            reg.get("p_exp", 0),
            reg.get("p_g3",  0),
            reg.get("p_g4",  0),
            reg.get("p_g5",  0),
        ]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=col, value=v)
            c.fill = fill
        r += 1

    # Ajustar largura das colunas automaticamente
    for col in ws.columns:
        max_w = max((len(str(c.value)) for c in col if c.value), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_w + 2, 30)

    try:
        wb.save(str(destino))
        print(f"\nRelatório Excel salvo em:\n  {destino}")
    except Exception as e:
        print(f"\n  AVISO: não foi possível salvar relatório Excel: {e}")
    finally:
        wb.close()


# =============================================================================
# SEÇÃO 13 — EXECUÇÃO PRINCIPAL
# =============================================================================

def _rodar_analise(excel_path_run, curvas_ativas_run=None, modo_sequencia=False):
    """
    Executa a análise completa para uma planilha.
    curvas_ativas_run sobrescreve CURVAS_PESO_ATIVAS quando fornecido.
    modo_sequencia=True suprime o viewer interativo (não bloqueia a sequência).
    """
    global CURVAS_PESO_ATIVAS
    if curvas_ativas_run is not None:
        CURVAS_PESO_ATIVAS = curvas_ativas_run

    # ── 1. Leitura ────────────────────────────────────────────────────────────
    _ext = os.path.splitext(excel_path_run)[1].lower()
    _fmt_csv = (_ext == ".csv")
    _fmt_label = "CSV+TOML" if _fmt_csv else "Excel"
    print(f"Lendo dados ({_fmt_label})...")
    try:
        if _fmt_csv:
            fixos                 = ler_parametros_toml(excel_path_run)
            tempos_exp, sinal_exp = ler_dados_csv(excel_path_run)
            # nome da pasta pai como nome base do experimento
            nome_base = os.path.basename(os.path.dirname(os.path.abspath(excel_path_run)))
            if not nome_base:
                nome_base = os.path.splitext(os.path.basename(excel_path_run))[0]
        else:
            fixos                 = ler_parametros_fixos(excel_path_run)
            tempos_exp, sinal_exp = ler_dados_experimentais(excel_path_run)
            nome_base = os.path.splitext(os.path.basename(excel_path_run))[0]

        print(f"  OK — {len(tempos_exp)} pontos  "
              f"({tempos_exp[0]:.0f}–{tempos_exp[-1]:.0f} ns)")
        if len(tempos_exp) < 5:
            raise ValueError(f"Poucos pontos experimentais lidos ({len(tempos_exp)}). "
                             f"Verifique o arquivo de entrada.")
    except FileNotFoundError:
        print("  Arquivo não encontrado — usando dados de teste embutidos")
        fixos     = dict(FIXOS_FALLBACK)
        nome_base = "teste"
        tempos_exp = np.array([510,560,610,660,710,760,810,860,910,960,
                                1010,1060,1110,1160,1210,1260], dtype=float)
        sinal_exp  = np.array([1.0,0.985348,0.950200,0.928889,0.906022,
                                0.881524,0.855517,0.827979,0.798790,0.768289,
                                0.736600,0.705140,0.674342,0.644488,0.616251,
                                0.588994])

    # Preencher energias:
    #   Modo CSV: tenta toml → ENERGIAS_VALOR → erro (não usa fallback silencioso)
    #   Modo Excel: tenta célula → ENERGIAS_VALOR → fallback 1.0 eV (comportamento original)
    _energia_para_curva = {"e_g3": "p_g3", "e_g4": "p_g4", "e_g5": "p_g5"}
    _energias_faltando = []
    for ke in ("e_g3", "e_g4", "e_g5"):
        if fixos.get(ke) is None:
            val_cfg = ENERGIAS_VALOR.get(ke)
            if val_cfg is not None:
                fixos[ke] = val_cfg
                _origem = "parametros.toml" if _fmt_csv else "planilha"
                print(f"  INFO: {ke} ausente no {_origem} — usando ENERGIAS_VALOR = {val_cfg} eV")
            elif _fmt_csv:
                curva = _energia_para_curva[ke]
                if CURVAS_PESO_ATIVAS.get(curva, False):
                    _energias_faltando.append(ke)
                else:
                    fixos[ke] = 1.0   # valor irrelevante — curva inativa
            else:
                fixos[ke] = 1.0   # fallback Excel (comportamento original)
                print(f"  AVISO: {ke} vazio na planilha e sem valor em ENERGIAS_VALOR "
                      f"— usando fallback = 1.0 eV")
    if _energias_faltando:
        nomes = {'e_g3': 'Gaussiana 1', 'e_g4': 'Gaussiana 2', 'e_g5': 'Gaussiana 3'}
        faltando_str = ", ".join(nomes[k] for k in _energias_faltando)
        raise ValueError(
            f"Energias não definidas para: {faltando_str}.\n"
            "Preencha os campos de energia na interface do Natalia Time, "
            "ou grave os valores no parametros.toml usando o checkbox Gravar."
        )

    # Aplicar fallback para parâmetros geométricos ausentes
    for kg, fb in [("DE", 350.0), ("D", 6.8), ("LL", 6.8), ("I7_NORM", 500.0)]:
        if fixos.get(kg) is None:
            fixos[kg] = fb
            print(f"  INFO: {kg} não encontrado — usando fallback {fb}")

    print(f"\n  Parâmetros fixos:")
    for k, v in fixos.items():
        print(f"    {k:10s} = {v}")

    # ── 2. Inicializar grade ──────────────────────────────────────────────────
    inicializar_grade(fixos, t_min_exp=float(tempos_exp.min()),
                             t_max_exp=float(tempos_exp.max()))

    # ── 2b. Resumo das curvas configuradas ───────────────────────────────────
    W_R = 72
    print(f"\n{'═'*W_R}")
    print("CONFIGURAÇÃO DAS CURVAS")
    print(f"{'═'*W_R}")
    print(f"  {'Curva':<22} {'Status':<10} {'β inicial':>10}  {'Intervalo β':>18}  Energia")
    print(f"  {'─'*68}")

    # MB — sempre ativa, beta fixo
    pmin_mb_str = f"  (p_min={PESO_MIN['p_mb']:.2f})" if PESO_MIN.get('p_mb') is not None else ""
    print(f"  {'MB':<22} {'ATIVA':<10} {fixos['beta_mb']:>10.1f}  {'(fixo da planilha)':>18}  —{pmin_mb_str}")

    # EXP
    ativo_exp = CURVAS_PESO_ATIVAS.get('p_exp', True)
    lo_e, hi_e = INTERVALOS_BETA['beta_exp']
    intervalo_exp = f"[{lo_e:.0f} – {hi_e:.0f}]"
    print(f"  {'Exponencial':<22} {'ATIVA' if ativo_exp else 'INATIVA':<10} "
          f"{'—':>10}  {intervalo_exp:>18}  —")

    # Gaussianas
    for kp, kb, ke in [('p_g3','beta_g3','e_g3'),
                       ('p_g4','beta_g4','e_g4'),
                       ('p_g5','beta_g5','e_g5')]:
        nome   = NOME_CURVA.get(kp, kp)
        ativo  = CURVAS_PESO_ATIVAS.get(kp, True)
        lo_b, hi_b = INTERVALOS_BETA.get(kb, (0,0))
        intervalo_b = f"[{lo_b:.0f} – {hi_b:.0f}]"

        val_e  = ENERGIAS_VALOR.get(ke)
        nm_liv = ENERGIA_NM_LIVRE.get(ke, False)
        if val_e is not None and not nm_liv:
            en_str = f"fixo {val_e:.2f} eV (config)"
        elif val_e is not None and nm_liv:
            lo_en, hi_en = INTERVALOS_ENERGIA[ke]
            en_str = f"NM livre [{lo_en:.1f}–{hi_en:.1f}] eV  ponto={val_e:.2f}"
        elif nm_liv:
            lo_en, hi_en = INTERVALOS_ENERGIA[ke]
            en_str = f"NM livre [{lo_en:.1f}–{hi_en:.1f}] eV  ponto=planilha"
        else:
            en_str = f"fixo (planilha)"

        pmin = PESO_MIN.get(kp)
        pmin_str = f"  (p_min={pmin:.2f})" if pmin is not None else ""

        print(f"  {nome:<22} {'ATIVA' if ativo else 'INATIVA':<10} "
              f"{'—':>10}  {intervalo_b:>18}  {en_str}{pmin_str}")

    print(f"{'═'*W_R}\n")

    # ── 3. Pasta desta execução ───────────────────────────────────────────────
    agora      = datetime.datetime.now()
    stamp      = agora.strftime("%Y%m%d_%H%M%S")
    pasta_exec = os.path.join(RESULTADOS_DIR, f"{nome_base}_{stamp}")
    pasta_graf = os.path.join(pasta_exec, "graficos_rodadas")
    os.makedirs(pasta_exec, exist_ok=True)

    # Remove PARAR.txt residual de execução anterior
    _parar_path = os.path.join(pasta_exec, ARQUIVO_PARAR)
    if os.path.exists(_parar_path):
        os.remove(_parar_path)
    print(f"  (Para interromper e salvar: crie '{ARQUIVO_PARAR}' em:\n"
          f"   {pasta_exec})")

    # ── 4. Busca ou validação manual ─────────────────────────────────────────
    if MODO_VALIDACAO:
        # ── MODO VALIDAÇÃO: calcula direto com parâmetros conhecidos ──────────
        print("\n" + "=" * 70)
        print("MODO VALIDAÇÃO MANUAL")
        print("=" * 70)

        vp = VALIDACAO_PARAMS
        # Montar fixos_val: ENERGIAS_VALOR sobrescreve planilha; VALIDACAO_PARAMS sobrescreve tudo
        vp_fixos = dict(fixos)
        for k in ("e_g3", "e_g4", "e_g5"):
            val_e = ENERGIAS_VALOR.get(k)
            if val_e is not None:
                vp_fixos[k] = val_e
            if k in vp:
                vp_fixos[k] = vp[k]

        sinal_teo  = calcular_espectro(
            tempos_exp,
            vp["beta_exp"], vp["beta_g3"], vp["beta_g4"], vp["beta_g5"],
            vp["p_mb"], vp["p_exp"], vp["p_g3"], vp["p_g4"], vp["p_g5"],
            vp_fixos
        )
        af4              = calcular_af4(sinal_exp, sinal_teo)
        rmse, m_r, b_r   = calcular_rmse(tempos_exp, sinal_exp, sinal_teo)
        soma_p = sum(vp[k] for k in ["p_mb","p_exp","p_g3","p_g4","p_g5"])
        valido = af4 >= SCORE_MINIMO

        print(f"\n  R²    = {af4:.6f}  {'✓ VÁLIDO' if valido else '✗ abaixo do critério'}")
        print(f"  RMSE  = {rmse:.6f}")
        print(f"  Reta  = {m_r:.2e}·t + {b_r:.4f}")
        print(f"\n  {'Curva':<22} {'Largura':>10}  {'Peso':>10}  {'E central':>10}")
        print(f"  {'─'*56}")
        for kb, kp, e_ev in [
                ("beta_mb",  "p_mb",  0.0),
                ("beta_exp", "p_exp", 0.0),
                ("beta_g3",  "p_g3",  fixos["e_g3"]),
                ("beta_g4",  "p_g4",  fixos["e_g4"]),
                ("beta_g5",  "p_g5",  fixos["e_g5"])]:
            b_v = fixos.get(kb) or vp.get(kb, 0)
            p_v = vp.get(kp, 0)
            print(f"  {NOME_CURVA.get(kp,kp):<22} {b_v:>10.2f}  {p_v:>10.6f}  {e_ev:>8.2f} eV")
        print(f"  {'─'*56}")
        print(f"  {'Soma pesos':<22} {'':>10}  {soma_p:>10.6f}")
        print("=" * 70)

        # Construir registro compatível com o viewer e relatório
        melhor = {
            **vp,
            "beta_mb":  fixos["beta_mb"],
            "af4":      af4,
            "rmse":     rmse,
            "m_reta":   m_r,
            "b_reta":   b_r,
            "valido":   valido,
            "status":   "VÁLIDO ✓" if valido else "Abaixo do critério",
            "rodada":   "VAL",
        }
        resultado = {
            "melhor":        melhor,
            "historico":     [melhor],
            "validos":       1 if valido else 0,
            "alta_precisao": 1 if (valido and rmse <= DESVIO_ALVO) else 0,
            "total":         1,
            "duracao_s":     0.0,
            "calc_por_s":    0.0,
        }
        t_inicio_busca = t_fim_busca = datetime.datetime.now()
        resultados_nm  = []
        historico_completo = [melhor]

        # ── NM a partir dos parâmetros manuais ───────────────────────────────
        if VALIDACAO_REFINAR_NM:
            print(f"\n{'─'*70}")
            print("REFINAMENTO NM A PARTIR DOS PARÂMETROS MANUAIS")
            print(f"{'─'*70}")
            resultados_nm, melhor_nm = refinar_nelder_mead(
                tempos_exp, sinal_exp, vp_fixos, [melhor]
            )
            if (melhor_nm["valido"] and not melhor.get("valido", False)) or                (melhor_nm["valido"] == melhor.get("valido", False) and
                    melhor_nm["rmse"] < melhor.get("rmse", 99999)):
                melhor = melhor_nm
                print("  → NM melhorou o resultado manual")
            historico_completo = [resultado["historico"][0]] + resultados_nm

    else:
        # ── MODO NORMAL: busca aleatória + NM ────────────────────────────────
        t_inicio_busca = datetime.datetime.now()

        if BUSCA_DUAS_FASES:
            resultado = busca_duas_fases(tempos_exp, sinal_exp, fixos,
                                         pasta_parar=pasta_exec)
        else:
            resultado = busca_aleatoria(tempos_exp, sinal_exp, fixos,
                                        pasta_parar=pasta_exec)

        if ADAPTAR_INTERVALOS and not BUSCA_DUAS_FASES:
            resultado = adaptar_e_rebuscar(tempos_exp, sinal_exp, fixos, resultado,
                                           pasta_parar=pasta_exec)

        t_fim_busca = datetime.datetime.now()
        melhor      = resultado["melhor"]

    # ── 4b. Refinamento Nelder-Mead (modo normal apenas) ────────────────────────
    if not MODO_VALIDACAO:
        resultados_nm = []
    if not MODO_VALIDACAO and USAR_NELDER_MEAD:
        # Seleciona candidatos por clusters ou top-N por RMSE
        df_tmp      = pd.DataFrame(resultado["historico"])
        validos_tmp = df_tmp[df_tmp["status"] == "VÁLIDO"].sort_values("rmse")
        if len(validos_tmp) == 0:
            # Sem válidos: usa os de maior R²
            candidatos_nm = (df_tmp.sort_values("af4", ascending=False)
                             .head(NM_CLUSTERS_MAX if NM_USAR_CLUSTERS else NM_TOP_CANDIDATOS)
                             .to_dict("records"))
        elif NM_USAR_CLUSTERS:
            candidatos_nm = selecionar_candidatos_por_clusters(validos_tmp)
            if len(candidatos_nm) == 0:
                candidatos_nm = validos_tmp.head(NM_TOP_CANDIDATOS).to_dict("records")
        else:
            candidatos_nm = validos_tmp.head(NM_TOP_CANDIDATOS).to_dict("records")

        resultados_nm, melhor_nm = refinar_nelder_mead(
            tempos_exp, sinal_exp, fixos, candidatos_nm
        )
        # Substitui melhor global se NM encontrou algo melhor
        if (melhor_nm["valido"] and not melhor.get("valido", False)) or \
           (melhor_nm["valido"] == melhor.get("valido", False) and
                melhor_nm["rmse"] < melhor.get("rmse", 99999)):
            melhor = melhor_nm
            print("  → Melhor global atualizado pelo Nelder-Mead")

    # historico_completo (no modo normal, monta aqui; no validação já foi definido)
    if not MODO_VALIDACAO:
        historico_completo = resultado["historico"] + resultados_nm

    # ── 5/6. Resumo no terminal ──────────────────────────────────────────────
    df         = pd.DataFrame(historico_completo)
    validos_df = df[df["status"] == "VÁLIDO"].sort_values("rmse")
    n_show     = min(20, len(validos_df))

    def p2(x, fmt): return format(float(x), fmt).replace(".", ",")

    W2 = 90
    print(f"\n{'='*W2}")
    print("MELHOR RESULTADO GLOBAL")
    print(f"{'='*W2}")
    print(f"  AF4         = {p2(melhor['af4'],'.6f')}")
    print(f"  RMSE        = {p2(melhor['rmse'],'.6f')}")
    m_r = melhor.get('m_reta', 0); b_r = melhor.get('b_reta', 0)
    m_r_str = f"{m_r:.2e}".replace(".", ",")
    b_r_str = f"{abs(b_r):.4f}".replace(".", ",")
    eq_r = f"{m_r_str}·t {'+' if b_r>=0 else '-'} {b_r_str}"
    print(f"  Reta res.   = {eq_r}")
    print(f"  Status      = {'VÁLIDO ✓' if melhor['valido'] else 'Abaixo do critério'}")
    print(f"\n  {'Curva':<22} {'Largura':>10}  {'Peso':>10}  {'E central':>10}")
    print(f"  {'─'*60}")
    for kb, kp, e_key in [("beta_mb","p_mb",None),("beta_exp","p_exp",None),
                          ("beta_g3","p_g3","e_g3"),
                          ("beta_g4","p_g4","e_g4"),
                          ("beta_g5","p_g5","e_g5")]:
        nome_c = NOME_CURVA.get(kp, kp)
        b   = fixos.get(kb) or melhor.get(kb, 0)
        p   = melhor.get(kp, 0)
        e   = 0.0 if e_key is None else (melhor.get(e_key) or fixos.get(e_key, 0))
        print(f"  {nome_c:<22} {p2(b,'>10.2f')}  {p2(p,'>10.6f')}   {p2(e,'>8.2f')} eV")

    _ds = resultado['duracao_s']
    _h, _rem = divmod(int(_ds), 3600); _m, _s = divmod(_rem, 60)
    _cps = resultado['total'] / max(_ds, 1)
    print(f"\n  Válidos     : {resultado['validos']}  |  "
          f"Alta precisão: {resultado['alta_precisao']}  |  "
          f"Total: {resultado['total']}  |  "
          f"Duração: {_h:02d}:{_m:02d}:{_s:02d} ({p2(_ds,'.1f')}s)  |  "
          f"{_cps:.0f} calc/s")
    print(f"{'='*W2}")

    if n_show > 0:
        print(f"\n{'─'*W2}")
        print(f"TOP {n_show} RODADAS VÁLIDAS (ordenadas por RMSE)")
        print(f"{'─'*W2}")
        tem_e_g3 = ENERGIA_NM_LIVRE.get("e_g3", False)
        tem_e_g4 = ENERGIA_NM_LIVRE.get("e_g4", False)
        tem_e_g5 = ENERGIA_NM_LIVRE.get("e_g5", False)
        hdr = (f"  {'Rod':>5}  {'R²':>8}  {'RMSE':>8}  "
               f"{'β_exp':>7}  {'β_G1':>7}  {'β_G2':>7}")
        if tem_e_g3: hdr += f"  {'E_G1(eV)':>8}"
        if tem_e_g4: hdr += f"  {'E_G2(eV)':>8}"
        if tem_e_g5: hdr += f"  {'E_G3(eV)':>8}"
        hdr += f"  {'p_mb':>8}  {'p_exp':>8}  {'p_G1':>8}  {'p_G2':>8}"
        print(hdr)
        print(f"  {'─'*(len(hdr)-2)}")
        for _, row in validos_df.head(n_show).iterrows():
            rod_str = str(row['rodada'])
            linha = (f"  {rod_str:>5}  "
                     f"{p2(row['af4'],'.6f'):>8}  "
                     f"{p2(row['rmse'],'.6f'):>8}  "
                     f"{p2(row['beta_exp'],'.1f'):>7}  "
                     f"{p2(row['beta_g3'],'.1f'):>7}  "
                     f"{p2(row['beta_g4'],'.1f'):>7}")
            if tem_e_g3: linha += f"  {p2(row.get('e_g3', fixos.get('e_g3',0)),'.3f'):>8}"
            if tem_e_g4: linha += f"  {p2(row.get('e_g4', fixos.get('e_g4',0)),'.3f'):>8}"
            if tem_e_g5: linha += f"  {p2(row.get('e_g5', fixos.get('e_g5',0)),'.3f'):>8}"
            linha += (f"  {p2(row['p_mb'],'.6f'):>8}  "
                      f"{p2(row['p_exp'],'.6f'):>8}  "
                      f"{p2(row['p_g3'],'.6f'):>8}  "
                      f"{p2(row['p_g4'],'.6f'):>8}")
            print(linha)
        print(f"{'─'*W2}")
    else:
        print("\n  Nenhum resultado válido encontrado.")

    # ── 7. Relatório de texto ─────────────────────────────────────────────────
    def vc(x, fmt=""):
        """Float formatado com vírgula decimal."""
        return format(float(x), fmt).replace(".", ",") if fmt else str(x).replace(".", ",")

    caminho_rel = os.path.join(pasta_exec, f"{nome_base}_relatorio.txt")
    W = 72
    with open(caminho_rel, "w", encoding="utf-8") as f:
        f.write("=" * W + "\n")
        f.write("RELATÓRIO DE OTIMIZAÇÃO TOF\n")
        f.write("=" * W + "\n")
        f.write(f"  Data/hora   : {agora:%d/%m/%Y %H:%M:%S}\n")
        f.write(f"  Planilha    : {os.path.basename(excel_path_run)}\n")
        _ds = resultado['duracao_s']
        _h, _rem = divmod(int(_ds), 3600); _m, _s = divmod(_rem, 60)
        _cps = resultado['total'] / max(_ds, 1)
        f.write(f"  Duração     : {_h:02d}:{_m:02d}:{_s:02d}  ({vc(_ds,'.1f')} s)\n")
        f.write(f"  Calc/s      : {_cps:.1f}\n")
        f.write(f"  Cálculos    : {resultado['total']}\n")
        f.write(f"  Válidos     : {resultado['validos']}  (R² >= {SCORE_MINIMO})\n")
        f.write(f"  Alta prec.  : {resultado['alta_precisao']}  (RMSE <= {DESVIO_ALVO})\n")
        taxa_val = resultado['validos'] / max(resultado['total'], 1) * 100
        f.write(f"  Taxa válidos: {taxa_val:.1f}%\n")

        f.write("\n" + "─" * W + "\n")
        f.write("PARÂMETROS FIXOS\n")
        f.write("─" * W + "\n")
        f.write(f"  {'beta_mb':<12} = {vc(fixos['beta_mb'],'.4f')}\n")
        f.write(f"  {'m_frag':<12} = {vc(fixos['m_frag'],'.1f')}\n")
        f.write(f"  {'m_mol':<12} = {vc(fixos['m_mol'],'.1f')}\n")
        f.write(f"  {'e_g3':<12} = {vc(fixos['e_g3'],'.2f')} eV\n")
        f.write(f"  {'e_g4':<12} = {vc(fixos['e_g4'],'.2f')} eV\n")
        f.write(f"  {'e_g5':<12} = {vc(fixos['e_g5'],'.2f')} eV\n")

        f.write("\n" + "─" * W + "\n")
        f.write("INTERVALOS DE BUSCA\n")
        f.write("─" * W + "\n")
        for k, (lo, hi) in INTERVALOS_BETA.items():
            f.write(f"  {k:<12}: {vc(lo,'.1f')} – {vc(hi,'.1f')}\n")
        f.write(f"  Iterações   : {ITERACOES_BETA} betas × {ITERACOES_PESOS} pesos\n")
        curvas_on  = [NOME_CURVA.get(k,k) for k,v in CURVAS_PESO_ATIVAS.items() if v]
        curvas_off = [NOME_CURVA.get(k,k) for k,v in CURVAS_PESO_ATIVAS.items() if not v]
        f.write(f"  Curvas ativas: {', '.join(curvas_on)}\n")
        if curvas_off:
            f.write(f"  Desativadas  : {', '.join(curvas_off)}\n")
        for ke in ("e_g3","e_g4","e_g5"):
            val_e  = ENERGIAS_VALOR.get(ke)
            nm_liv = ENERGIA_NM_LIVRE.get(ke, False)
            if nm_liv:
                lo_en, hi_en = INTERVALOS_ENERGIA[ke]
                src = f"{vc(val_e,'.2f')} eV" if val_e else "planilha"
                f.write(f"  {ke:<12} : NM livre [{vc(lo_en,'.1f')}–{vc(hi_en,'.1f')}] eV  ponto={src}\n")
            elif val_e is not None:
                f.write(f"  {ke:<12} : fixo {vc(val_e,'.2f')} eV (config)\n")
            else:
                f.write(f"  {ke:<12} : fixo (planilha)\n")
        pesos_min = {k: v for k, v in PESO_MIN.items() if v is not None}
        if pesos_min:
            pm_str = "  ".join(f"{NOME_CURVA_CURTO.get(k,k)}≥{v:.2f}"
                               for k, v in pesos_min.items())
            f.write(f"  Pesos mín.   : {pm_str}\n")

        # Seção NM
        if USAR_NELDER_MEAD:
            f.write("\n" + "─" * W + "\n")
            f.write("REFINAMENTO NELDER-MEAD\n")
            f.write("─" * W + "\n")
            f.write(f"  Ativo       : Sim\n")
            if NM_USAR_CLUSTERS:
                f.write(f"  Candidatos  : clusters (limiar={NM_CLUSTER_LIMIAR:.2f}  max={NM_CLUSTERS_MAX})\n")
            else:
                f.write(f"  Candidatos  : top {NM_TOP_CANDIDATOS} por RMSE\n")
            f.write(f"  Max iter    : {NM_MAX_ITER}\n")
            f.write(f"  Tolerância  : {NM_TOLERANCIA}\n")
            nm_regs = [r for r in historico_completo
                       if str(r.get("rodada","")).startswith("NM")]
            if nm_regs:
                nm_validos = [r for r in nm_regs if r.get("status") == "VÁLIDO"]
                f.write(f"  Rodadas NM  : {len(nm_regs)}\n")
                f.write(f"  Válidos NM  : {len(nm_validos)}\n")
                melhor_nm_r = min(nm_regs, key=lambda r: r.get("rmse", 99))
                nm_origem   = melhor_nm_r.get("rodada","?")
                nm_iters    = melhor_nm_r.get("nm_iter","?")
                nm_conv     = "Sim" if melhor_nm_r.get("nm_convergiu", False) else "Não"
                f.write(f"  Melhor NM   : {nm_origem}  "
                        f"R²={vc(melhor_nm_r['af4'],'.6f')}  "
                        f"RMSE={vc(melhor_nm_r['rmse'],'.6f')}\n")
                f.write(f"  Iterações   : {nm_iters}  Convergiu: {nm_conv}\n")
                nm_melhorou = (melhor.get("rodada","") == melhor_nm_r.get("rodada",""))
                f.write(f"  NM melhorou global: {'Sim ✓' if nm_melhorou else 'Não'}\n")
        else:
            f.write("\n" + "─" * W + "\n")
            f.write("REFINAMENTO NELDER-MEAD\n")
            f.write("─" * W + "\n")
            f.write(f"  Ativo       : Não (USAR_NELDER_MEAD = False)\n")

        f.write("\n" + "=" * W + "\n")
        f.write("MELHOR RESULTADO GLOBAL\n")
        f.write("=" * W + "\n")
        f.write(f"  R²          = {vc(melhor['af4'],'.6f')}\n")
        f.write(f"  RMSE        = {vc(melhor['rmse'],'.6f')}\n")
        m_r_ = melhor.get("m_reta", 0); b_r_ = melhor.get("b_reta", 0)
        sinal_b = "+" if b_r_ >= 0 else "-"
        m_str = f"{m_r_:.2e}".replace(".", ",")
        b_str = f"{abs(b_r_):.4f}".replace(".", ",")
        f.write(f"  Reta res.   = {m_str}·t {sinal_b} {b_str}\n")
        f.write(f"  Status      = {'VÁLIDO' if melhor['valido'] else 'Abaixo do critério'}\n")
        f.write("\n")
        f.write(f"  {'Curva':<22} {'Largura':>10}  {'Peso':>10}  {'E central':>10}\n")
        f.write(f"  {'─'*60}\n")
        for kb, kp, e_key in [
                ("beta_mb",  "p_mb",  None),
                ("beta_exp", "p_exp", None),
                ("beta_g3",  "p_g3",  "e_g3"),
                ("beta_g4",  "p_g4",  "e_g4"),
                ("beta_g5",  "p_g5",  "e_g5")]:
            nome_c = NOME_CURVA.get(kp, kp)
            b_v  = fixos.get(kb) or melhor.get(kb, 0)
            p_v  = melhor.get(kp, 0)
            e_ev = 0.0 if e_key is None else (melhor.get(e_key) or fixos.get(e_key, 0))
            livre = (e_key is not None and ENERGIA_NM_LIVRE.get(e_key, False))
            tag   = " *" if livre else ""
            f.write(f"  {nome_c:<22} {vc(b_v,'>10.2f')}  {vc(p_v,'>10.6f')} "
                    f"  {vc(e_ev,'>8.2f')} eV{tag}\n")
        if any(ENERGIA_NM_LIVRE.get(k, False) for k in ("e_g3","e_g4","e_g5")):
            f.write(f"  (* energia livre — otimizada na busca)\n")

        if len(validos_df) > 0:
            n_top   = min(20, len(validos_df))
            e_livre = {k: ENERGIA_NM_LIVRE.get(k, False)
                       for k in ("e_g3","e_g4","e_g5")}
            f.write("\n" + "─" * W + "\n")
            f.write(f"TOP {n_top} RODADAS VÁLIDAS (ordenadas por RMSE crescente)\n")
            f.write("─" * W + "\n")
            # Cabeçalho dinâmico — colunas de energia só aparecem quando livres
            hdr = (f"  {'Rod':>5}  {'R²':>8}  {'RMSE':>8}  "
                   f"{'β_exp':>7}  {'β_G1':>7}  {'β_G2':>7}")
            if e_livre["e_g3"]: hdr += f"  {'E_G1(eV)':>9}"
            if e_livre["e_g4"]: hdr += f"  {'E_G2(eV)':>9}"
            if e_livre["e_g5"]: hdr += f"  {'E_G3(eV)':>9}"
            hdr += f"  {'p_mb':>8}  {'p_exp':>8}  {'p_G1':>8}  {'p_G2':>8}  Reta\n"
            f.write(hdr)
            f.write("  " + "─" * (len(hdr)-3) + "\n")
            for _, row in validos_df.head(n_top).iterrows():
                m_r = row.get("m_reta", 0)
                b_r = row.get("b_reta", 0)
                eq_m = f"{m_r:.2e}".replace(".", ",")
                eq_b = f"{abs(b_r):.4f}".replace(".", ",")
                eq   = f"{eq_m}·t {'+' if b_r>=0 else '-'} {eq_b}"
                linha = (f"  {str(row['rodada']):>5}  "
                         f"{vc(row['af4'],'.6f'):>8}  "
                         f"{vc(row['rmse'],'.6f'):>8}  "
                         f"{vc(row['beta_exp'],'.1f'):>7}  "
                         f"{vc(row['beta_g3'],'.1f'):>7}  "
                         f"{vc(row['beta_g4'],'.1f'):>7}")
                if e_livre["e_g3"]:
                    linha += f"  {vc(row.get('e_g3', fixos.get('e_g3',0)),'.3f'):>9}"
                if e_livre["e_g4"]:
                    linha += f"  {vc(row.get('e_g4', fixos.get('e_g4',0)),'.3f'):>9}"
                if e_livre["e_g5"]:
                    linha += f"  {vc(row.get('e_g5', fixos.get('e_g5',0)),'.3f'):>9}"
                linha += (f"  {vc(row['p_mb'],'.6f'):>8}  "
                          f"{vc(row['p_exp'],'.6f'):>8}  "
                          f"{vc(row['p_g3'],'.6f'):>8}  "
                          f"{vc(row['p_g4'],'.6f'):>8}  "
                          f"{eq}\n")
                f.write(linha)
    print(f"Relatório salvo em:\n  {caminho_rel}")

    # ── 8. Gráfico do melhor global ───────────────────────────────────────────
    caminho_melhor = os.path.join(pasta_exec, f"{nome_base}_melhor_global.png")
    gerar_e_salvar_grafico(tempos_exp, sinal_exp, melhor, fixos,
                           "DETOF — Melhor Global", caminho_melhor)
    print(f"Gráfico do melhor salvo em:\n  {caminho_melhor}")

    # ── 9. Gráficos por rodada válida ─────────────────────────────────────────
    validas_todas   = sorted(
        [r for r in historico_completo if r["status"] == "VÁLIDO"],
        key=lambda r: r["rmse"]
    )
    n_validas_total = len(validas_todas)

    # ── 9b. Relatório na planilha Excel ───────────────────────────────────────
    escrever_relatorio_excel(
        excel_path_run, resultado, melhor, fixos,
        t_inicio_busca, t_fim_busca, historico_completo,
        pasta_exec=pasta_exec
    )

    if SALVAR_GRAFICOS_TOP is not None and n_validas_total > 0:
        n_gerar = min(SALVAR_GRAFICOS_TOP, n_validas_total)
        os.makedirs(pasta_graf, exist_ok=True)
        print(f"\nGerando {n_gerar} gráfico(s) (top {SALVAR_GRAFICOS_TOP} de "
              f"{n_validas_total} válidas)...")
        for reg in validas_todas[:n_gerar]:
            rod_str = str(reg["rodada"])
            rod_fmt = rod_str.zfill(4) if rod_str.isdigit() else rod_str
            fname   = (f"rod{rod_fmt}_"
                       f"R2_{reg['af4']:.5f}_"
                       f"RMSE_{reg['rmse']:.5f}.png")
            caminho = os.path.join(pasta_graf, fname)
            titulo  = f"DETOF — Rodada {rod_str}  [{reg['status']}]"
            fixos_reg = {**fixos,
                         **{k: reg[k] for k in ("e_g3","e_g4","e_g5") if k in reg}}
            gerar_e_salvar_grafico(tempos_exp, sinal_exp, reg, fixos_reg,
                                   titulo, caminho)
        print(f"  Gráficos em: {pasta_graf}")
    elif SALVAR_GRAFICOS_TOP is None and n_validas_total > 0:
        os.makedirs(pasta_graf, exist_ok=True)
        print(f"\nGerando {n_validas_total} gráfico(s) de rodadas válidas...")
        for reg in validas_todas:
            rod_str = str(reg["rodada"])
            rod_fmt = rod_str.zfill(4) if rod_str.isdigit() else rod_str
            fname   = (f"rod{rod_fmt}_"
                       f"R2_{reg['af4']:.5f}_"
                       f"RMSE_{reg['rmse']:.5f}.png")
            caminho = os.path.join(pasta_graf, fname)
            titulo  = f"DETOF — Rodada {rod_str}  [{reg['status']}]"
            fixos_reg = {**fixos,
                         **{k: reg[k] for k in ("e_g3","e_g4","e_g5") if k in reg}}
            gerar_e_salvar_grafico(tempos_exp, sinal_exp, reg, fixos_reg,
                                   titulo, caminho)
        print(f"  Gráficos em: {pasta_graf}")

    # ── 10. Viewer interativo ─────────────────────────────────────────────────
    if modo_sequencia:
        print("  (Viewer suprimido em modo sequência)")
    elif n_validas_total > 0:
        print(f"\nAbrindo viewer  ({n_validas_total} rodadas válidas, "
              f"ordenadas por RMSE)\n  ← →  navegar   |   Q  fechar")
        ViewerRodadas(historico_completo,
                      tempos_exp, sinal_exp, fixos,
                      titulo="DETOF").mostrar()
    else:
        print("\nNenhum resultado válido (busca + NM) — exibindo melhor encontrado...")
        fig  = plt.figure(figsize=(11, 7), constrained_layout=True)
        gs   = gridspec.GridSpec(2, 1, height_ratios=[3, 1], hspace=0.06, figure=fig)
        ax1  = fig.add_subplot(gs[0])
        ax2  = fig.add_subplot(gs[1], sharex=ax1)
        _desenhar_nos_eixos(ax1, ax2, tempos_exp, sinal_exp, melhor, fixos,
                            "DETOF — Melhor encontrado (abaixo do critério)")
        plt.show()


if __name__ == "__main__":
    mp.freeze_support()   # necessário no Windows com PyInstaller / multiprocessing

    if SEQUENCIA:
        # ── Modo sequência: processa cada planilha em ordem ───────────────────
        n_total = len(SEQUENCIA)
        print(f"{'═'*72}")
        print(f"SEQUÊNCIA DE {n_total} PLANILHA(S)")
        print(f"{'═'*72}")
        curvas_orig = dict(CURVAS_PESO_ATIVAS)
        for i, entrada in enumerate(SEQUENCIA):
            excel_i   = entrada["planilha"]
            curvas_i  = entrada.get("curvas_ativas", None)
            print(f"\n{'─'*72}")
            print(f"  [{i+1}/{n_total}] {os.path.basename(excel_i)}")
            if curvas_i:
                ativas = [k for k, v in curvas_i.items() if v]
                print(f"  Curvas: {', '.join(NOME_CURVA.get(k,k) for k in ativas)}")
            print(f"{'─'*72}\n")
            _rodar_analise(excel_i, curvas_i, modo_sequencia=True)
            CURVAS_PESO_ATIVAS = curvas_orig   # restaura para a próxima planilha
        print(f"\n{'═'*72}")
        print(f"SEQUÊNCIA CONCLUÍDA — {n_total} planilha(s) processada(s)")
        print(f"{'═'*72}")
    else:
        # ── Modo simples: processa apenas EXCEL_PATH ──────────────────────────
        _rodar_analise(EXCEL_PATH)
